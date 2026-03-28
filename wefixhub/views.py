# 1. Bibliotecas Padrão do Python
import os
import re
import json
import locale
import calendar
import unicodedata
from datetime import datetime, date, timedelta
from decimal import Decimal
from io import BytesIO
from urllib.parse import quote

# 2. Bibliotecas de Terceiros
import stripe
import pandas as pd
import numpy as np
import openpyxl
import xlsxwriter
import pdfplumber
import requests
from fpdf import FPDF
import pdb 

# 3. Núcleo e Utilidades do Django
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.admin.views.decorators import staff_member_required
from django.core.cache import cache
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction, connection
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone, formats
from django.utils.crypto import get_random_string
from django.views.decorators.http import require_POST
from django import template

#Utils

from .utils import ( gerar_dados_dashboard_analise, gerar_excel_vendas_reais,
                    processar_status_pdf, processar_giro_cliente,
                    calcular_evolucao_clientes )


# 4. ORM e Banco de Dados do Django
from django.db.models import (
    Count, Sum, Max, F, Q, OuterRef, Subquery, 
    DecimalField, IntegerField, ExpressionWrapper, Value
)
from django.db.models.functions import (
    ExtractMonth, TruncMonth, TruncDate, TruncDay, Coalesce, Cast
)

# 5. Aplicações Locais (Models e Forms)
from .models import (
    Product, Pedido, ItemPedido, WfClient, Endereco,
    ItemPedidoIgnorado, VendaReal, StatusPedidoERP,
    Carrinho, ItemCarrinho, SugestaoCompraERP, HistoricoPreco,
    Empresa, PerfilUsuario,
    Tarefa, TagTarefa, ComentarioTarefa, NotificacaoTarefa,
    AtividadeTarefa, ChecklistItem,
    NotificacaoPedido,
    ComentarioPedido,
    AnexoTarefa,
    LogAuditoria,
)
from .forms import (
    WfClientForm, EnderecoForm, GerarPedidoForm,
    UploadPedidoForm, SelectClientForm, CadastroEmpresaForm
)

# ==========================================
# SAAS — HELPER DE ISOLAMENTO POR EMPRESA
# ==========================================

def por_empresa(queryset, request, campo='empresa'):
    """Filtra o queryset pela empresa do usuário. Superuser vê tudo (empresa=None).
    Para modelos com campo 'cliente', também inclui registros sem empresa definida
    cujo cliente pertença à empresa (compatibilidade com registros antigos).
    """
    from django.db.models import Q
    if request.empresa:
        model = queryset.model
        fields = [f.name for f in model._meta.get_fields()]
        if 'cliente' in fields:
            return queryset.filter(
                Q(**{campo: request.empresa}) |
                Q(**{campo + '__isnull': True, 'cliente__empresa': request.empresa})
            )
        return queryset.filter(**{campo: request.empresa})
    return queryset


def get_empresa_or_404(model, request, **kwargs):
    """get_object_or_404 + filtra por empresa para usuários não-superuser.
    Para modelos com campo 'cliente', também encontra registros antigos sem empresa.
    """
    from django.db.models import Q
    if request.empresa:
        fields = [f.name for f in model._meta.get_fields()]
        if 'cliente' in fields:
            qs = model.objects.filter(
                Q(empresa=request.empresa) |
                Q(empresa__isnull=True, cliente__empresa=request.empresa),
                **kwargs
            )
            if not qs.exists():
                from django.http import Http404
                raise Http404
            return qs.first()
        kwargs['empresa'] = request.empresa
    return get_object_or_404(model, **kwargs)



# Configuração do locale para o formato brasileiro
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    print("Aviso: Locale 'pt_BR.UTF-8' não encontrado. Usando 'C.UTF-8' como fallback.")
    locale.setlocale(locale.LC_ALL, 'C.UTF-8')

try:
    # Tenta definir o locale ideal
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    # Se falhar, usa um locale padrão que suporta UTF-8
    print("Aviso: Locale 'pt_BR.UTF-8' não encontrado. Usando 'C.UTF-8' como fallback.")
    locale.setlocale(locale.LC_ALL, 'C.UTF-8')

# O resto do seu código continua aqui...
#from unidecode import unidecode 



# View para a página inicial com filtros e paginação
@login_required
def home(request):
    # Superuser vai para o dashboard SaaS
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect('saas_dashboard')

    # 1. Parâmetros de Filtro e Busca
    codigo = request.GET.get('codigo', None)
    descricao = request.GET.get('descricao', None)
    grupo = request.GET.get('grupo', None)
    marca = request.GET.get('marca', None)
    pedidos_rascunho_count = por_empresa(Pedido.objects.filter(status='RASCUNHO'), request).count()

    # 2. Obter o registro mais recente para cada produto
    is_staff = request.user.is_authenticated and request.user.is_staff

    empresa_key = request.empresa.slug if request.empresa else 'global'
    cache_key = f'home_products_{empresa_key}'

    products = cache.get(cache_key) if is_staff and not any([codigo, descricao, grupo, marca]) else None

    if products is None:
        # Usa Max em vez de subquery correlacionada — muito mais rápido
        from django.db.models import Max
        base_qs = por_empresa(Product.objects, request)
        latest_date = base_qs.aggregate(max_date=Max('date_product'))['max_date']

        products_qs = base_qs.filter(date_product=latest_date).order_by('product_code')

        # Subqueries de preço anterior — apenas para clientes (indicadores de variação)
        if not is_staff:
            prev_sp_sq = HistoricoPreco.objects.filter(
                product_code=OuterRef('product_code')
            ).order_by('-data_registro').values('product_value_sp')[1:2]

            prev_es_sq = HistoricoPreco.objects.filter(
                product_code=OuterRef('product_code')
            ).order_by('-data_registro').values('product_value_es')[1:2]

            products_qs = products_qs.annotate(
                prev_value_sp=Subquery(prev_sp_sq),
                prev_value_es=Subquery(prev_es_sq),
            )

        products = products_qs

        if is_staff and not any([codigo, descricao, grupo, marca]):
            cache.set(cache_key, products, 300)

    # 3. Aplicação dos Filtros de Busca
    if codigo:
        products = products.filter(product_code__icontains=codigo)
    if descricao:
        products = products.filter(product_description__icontains=descricao)
    if grupo:
        products = products.filter(product_group__icontains=grupo)
    if marca:
        products = products.filter(product_brand__icontains=marca)
    
    preco_exibido = None
    cliente_logado = None
    itens_frequentes = [] 
    produtos_wishlist_cliente = [] 
    total_carrinho_real = Decimal('0.00') # Inicializa o total como zero

    # 4. Lógica de Preços, Recomendações, WISHLIST e SINCRONIZAÇÃO DE CARRINHO
    if request.user.is_authenticated:
        if request.user.is_staff:
            preco_exibido = 'todos'
        else:
            try:
                cliente_logado = request.user.wfclient
                itens_frequentes = cliente_logado.get_frequent_items(limit=6)
                estado_cliente = cliente_logado.client_state.uf_name
                
                if estado_cliente == 'SP':
                    preco_exibido = 'sp'
                    products = products.exclude(product_value_sp=0, status_estoque='DISPONIVEL')
                elif estado_cliente == 'ES':
                    preco_exibido = 'es'
                    products = products.exclude(product_value_es=0, status_estoque='DISPONIVEL')

                # --- SINCRONIZAÇÃO DO TOTAL E QUANTIDADES DO CARRINHO (NOVO) ---
                # Importação local para evitar importação circular no topo do arquivo
                from .models import Carrinho, ItemCarrinho 
                carrinho = Carrinho.objects.filter(cliente=cliente_logado).first()
                
                if carrinho:
                    total_carrinho_real = carrinho.get_total_carrinho()
                    
                    # CUIDADO APLICADO: Cria uma subquery que busca exatamente a quantidade 
                    # do produto atual dentro do carrinho do cliente logado.
                    sq_qtd = ItemCarrinho.objects.filter(
                        carrinho=carrinho,
                        produto_id=OuterRef('product_id')
                    ).values('quantidade')[:1]
                    
                    # Anota (injeta) a quantidade no QuerySet de produtos
                    products = products.annotate(qtd_carrinho=Subquery(sq_qtd))

                # --- ALERTA DE WISHLIST NO FRONT-END ---
                data_limite = timezone.localdate() - timedelta(days=30)
                itens_pendentes = ItemPedidoIgnorado.objects.filter(
                    cliente=cliente_logado,
                    descartado_pelo_cliente=False,
                    motivo_erro__icontains="estoque",
                    data_tentativa__gte=data_limite
                )

                if itens_pendentes.exists():
                    codigos_pendentes = itens_pendentes.values_list('codigo_produto', flat=True).distinct()
                    produtos_dict = {p.product_code: p for p in Product.objects.filter(product_code__in=codigos_pendentes)}

                    for item in itens_pendentes:
                        produto = produtos_dict.get(item.codigo_produto)
                        if not produto: continue

                        preco_atual = getattr(produto, 'product_value_sp' if estado_cliente == 'SP' else 'product_value_es')

                        if preco_atual and preco_atual > 0 and produto.status_estoque != 'SEM_ESTOQUE':
                            if not any(p['codigo'] == produto.product_code for p in produtos_wishlist_cliente):
                                produtos_wishlist_cliente.append({
                                    'codigo': produto.product_code,
                                    'descricao': produto.product_description,
                                    'preco': f"{preco_atual.quantize(Decimal('0.01'))}".replace('.', ','),
                                    'product_id': produto.product_id,
                                })

                # Produtos que o cliente já solicitou aviso
                codigos_avise_me = set(
                    ItemPedidoIgnorado.objects.filter(
                        cliente=cliente_logado,
                        notificado=False,
                        motivo_erro__icontains='estoque'
                    ).values_list('codigo_produto', flat=True)
                )

            except WfClient.DoesNotExist:
                products = Product.objects.none()
                codigos_avise_me = set()

    if not request.user.is_authenticated:
        products = Product.objects.none()
    
    # 5. Paginação — antes da formatação para não carregar tudo na memória
    paginator = Paginator(products, 30)
    page = request.GET.get('page')

    try:
        product_list = paginator.page(page)
    except PageNotAnInteger:
        product_list = paginator.page(1)
    except EmptyPage:
        product_list = paginator.page(paginator.num_pages)

    # 6. Formatação apenas dos produtos da página atual (30 itens)
    for product in product_list:
        # Produto sem estoque mostra preço como 0,00 para o cliente
        sem_estoque = product.status_estoque == 'SEM_ESTOQUE'
        sp = product.product_value_sp if not sem_estoque else None
        es = product.product_value_es if not sem_estoque else None
        product.valor_sp_formatado = f"{sp.quantize(Decimal('0.01'))}".replace('.', ',') if sp else "0,00"
        product.valor_es_formatado = f"{es.quantize(Decimal('0.01'))}".replace('.', ',') if es else "0,00"

        prev_sp = getattr(product, 'prev_value_sp', None)
        product.desconto_sp = None
        if prev_sp and product.product_value_sp and prev_sp > 0 and product.product_value_sp < prev_sp:
            product.desconto_sp = round((prev_sp - product.product_value_sp) / prev_sp * 100, 1)

        prev_es = getattr(product, 'prev_value_es', None)
        product.desconto_es = None
        if prev_es and product.product_value_es and prev_es > 0 and product.product_value_es < prev_es:
            product.desconto_es = round((prev_es - product.product_value_es) / prev_es * 100, 1)
        
    # 7. Contexto do Template
    context = {
        'product_list': product_list,
        'itens_frequentes': itens_frequentes, 
        'produtos_wishlist_cliente': produtos_wishlist_cliente,
        'cliente_logado': cliente_logado,
        'preco_exibido': preco_exibido,
        'data_hoje': products[0].date_product if products else date.today(),
        'pedidos_rascunho_count': pedidos_rascunho_count,
        # INJETADO: Valor numérico puro para o data-attribute do HTML
        'total_carrinho_real': float(total_carrinho_real),
        'codigos_avise_me': codigos_avise_me if 'codigos_avise_me' in locals() else set(),
    }
            
    return render(request, 'home.html', context)

"""
    # Lógica de Paginação:
    paginator = Paginator(product_list, 10)
    page = request.GET.get('page')
    try:
        produtos_na_pagina = paginator.page(page)
    except PageNotAnInteger:
        produtos_na_pagina = paginator.page(1)
    except EmptyPage:
        produtos_na_pagina = paginator.page(paginator.num_pages)
    
    # NOVO: Puxa o objeto WfClient do usuário logado
    cliente_logado = None
    if request.user.is_authenticated:
        try:
            cliente_logado = request.user.wfclient
        except WfClient.DoesNotExist:
            cliente_logado = None

    carrinho_da_sessao = request.session.get('carrinho', {})
    contexto = {
        'titulo': 'Página Inicial',
        'product_list': produtos_na_pagina,
        'carrinho': carrinho_da_sessao,
        'cliente_logado': cliente_logado, # NOVO: Adiciona o cliente ao contexto
    }
    return render(request, 'home.html', contexto)"""

# Inicio gerar pedido
# Importa o módulo do debugger


# Inicio do carrinho


@login_required
def carrinho(request):
    # 1. Identifica o cliente logado
    try:
        cliente_logado = request.user.wfclient
    except WfClient.DoesNotExist:
        messages.error(request, "Cliente não encontrado.")
        return redirect('home')

    # 2. Busca o carrinho do banco de dados
    carrinho_obj = Carrinho.objects.filter(cliente=cliente_logado).first()
    
    itens_carrinho = []
    total_carrinho = Decimal('0.00')

    # 3. Se existir um carrinho, pega os itens
    if carrinho_obj:
        # Pega o estado para definir qual preço mostrar (SP ou ES)
        uf_cliente = cliente_logado.client_state.uf_name if cliente_logado.client_state else 'SP'
        
        # Otimização: select_related traz o Produto junto, evitando consultas extras ao banco
        itens_db = carrinho_obj.itens.select_related('produto').all()
        
        for item in itens_db:
            preco_unitario = item.produto.product_value_sp if uf_cliente == 'SP' else item.produto.product_value_es
            preco_unitario = preco_unitario or Decimal('0.00')
            subtotal = preco_unitario * item.quantidade
            
            # Monta o dicionário para a tela (HTML) ler
            itens_carrinho.append({
                'item_id': item.id,
                'produto': item.produto,
                'quantidade': item.quantidade,
                'preco_unitario': preco_unitario,
                'subtotal': subtotal
            })
            total_carrinho += subtotal

    context = {
        'itens_carrinho': itens_carrinho,
        'total_carrinho': total_carrinho,
        'cliente_logado': cliente_logado,
    }
    
    return render(request, 'carrinho.html', context)


# Fim do carrinho


@login_required
def remover_item(request, product_id):
    """Remove um item específico do carrinho no Banco de Dados."""
    try:
        cliente_logado = request.user.wfclient
        # Deleta apenas o item especificado atrelado ao cliente logado
        ItemCarrinho.objects.filter(
            carrinho__cliente=cliente_logado, 
            produto__product_id=product_id
        ).delete()
        messages.success(request, "Item removido do carrinho.")
    except Exception as e:
        messages.error(request, f"Erro ao remover item: {e}")
        
    return redirect('carrinho')

@login_required
def limpar_carrinho(request):
    """Esvazia o carrinho no Banco de Dados."""
    try:
        # Apagar o Carrinho inteiro limpa todos os itens por causa do CASCADE
        Carrinho.objects.filter(cliente=request.user.wfclient).delete()
        messages.success(request, "Carrinho limpo com sucesso.")
    except Exception as e:
        messages.error(request, "Erro ao limpar o carrinho.")
        
    return redirect('carrinho')

@login_required
def atualizar_carrinho(request):
    """Atualiza as quantidades se o usuário digitar os números e clicar em 'Atualizar'."""
    if request.method == 'POST':
        try:
            cliente_logado = request.user.wfclient
            carrinho_obj = Carrinho.objects.filter(cliente=cliente_logado).first()
            
            if carrinho_obj:
                for key, value in request.POST.items():
                    # Procura pelos inputs de quantidade que nomeamos no HTML
                    if key.startswith('quantidade_'):
                        produto_id = key.split('_')[1]
                        try:
                            nova_qtd = int(value)
                        except (ValueError, TypeError):
                            messages.warning(request, f'Quantidade inválida para o produto {produto_id}, ignorado.')
                            continue

                        if nova_qtd > 0:
                            ItemCarrinho.objects.filter(
                                carrinho=carrinho_obj, 
                                produto__product_id=produto_id
                            ).update(quantidade=nova_qtd)
                        else:
                            # Se ele botar zero, deleta o item
                            ItemCarrinho.objects.filter(
                                carrinho=carrinho_obj, 
                                produto__product_id=produto_id
                            ).delete()
                            
                messages.success(request, "Carrinho atualizado com sucesso!")
        except Exception as e:
            messages.error(request, "Erro ao atualizar o carrinho.")
            
    return redirect('carrinho')


@login_required
def atualizar_item_qtd(request):
    """Atualiza a quantidade de um item via JSON (sem redirect). Usado pelo carrinho Alpine."""
    if request.method == 'POST':
        import json as _json
        try:
            data = _json.loads(request.body)
            product_id = data.get('product_id')
            quantidade = int(data.get('quantidade', 0))
            carrinho_obj = Carrinho.objects.filter(cliente=request.user.wfclient).first()
            if carrinho_obj:
                if quantidade > 0:
                    ItemCarrinho.objects.filter(
                        carrinho=carrinho_obj,
                        produto__product_id=product_id
                    ).update(quantidade=quantidade)
                else:
                    ItemCarrinho.objects.filter(
                        carrinho=carrinho_obj,
                        produto__product_id=product_id
                    ).delete()
            return JsonResponse({'ok': True})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False}, status=405)


@login_required
def remover_item_ajax(request, product_id):
    """Remove um item do carrinho via POST/JSON (sem redirect). Usado pelo carrinho Alpine."""
    if request.method == 'POST':
        try:
            ItemCarrinho.objects.filter(
                carrinho__cliente=request.user.wfclient,
                produto__product_id=product_id
            ).delete()
            return JsonResponse({'ok': True})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': False}, status=405)


# Inicio Checkout





# views.py

# ... (seus outros imports)

# --- SUBSTITUA A VIEW CHECKOUT EXISTENTE POR ESTA VERSÃO OTIMIZADA ---

@login_required
def checkout(request, pedido_id_rascunho=None):
    if request.method == 'POST':
        endereco_id = request.POST.get('endereco_selecionado')
        data_envio_str = request.POST.get('data_expedicao')
        frete_option = request.POST.get('frete_option')
        nota_fiscal = request.POST.get('nota_fiscal')
        observacao = request.POST.get('observacao')

        _frete_validos = [c[0] for c in Pedido.FRETE_CHOICES]
        _nota_validos = [c[0] for c in Pedido.NOTA_FISCAL_CHOICES]
        if frete_option not in _frete_validos:
            messages.error(request, 'Opção de frete inválida.')
            return redirect('checkout')
        if nota_fiscal not in _nota_validos:
            messages.error(request, 'Opção de nota fiscal inválida.')
            return redirect('checkout')

        id_do_pedido = pedido_id_rascunho or request.POST.get('pedido_id_rascunho')

        endereco_selecionado = None
        fretes_sem_endereco = ['ONIBUS', 'RETIRADA']
        data_envio_obj = None

        try:
            if id_do_pedido:
                if request.empresa:
                    from django.db.models import Q
                    pedido_rascunho = get_object_or_404(
                        Pedido,
                        Q(empresa=request.empresa) | Q(empresa__isnull=True, cliente__empresa=request.empresa),
                        id=id_do_pedido,
                    )
                else:
                    pedido_rascunho = get_object_or_404(Pedido, id=id_do_pedido)
                cliente_logado = pedido_rascunho.cliente
            else:
                if request.user.is_staff:
                    messages.error(request, 'Usuários administrativos não podem criar pedidos por esta página.')
                    return redirect('home')
                cliente_logado = request.user.wfclient

            if frete_option not in fretes_sem_endereco:
                if not endereco_id:
                    messages.error(request, 'Por favor, selecione um endereço válido.')
                    return redirect('checkout')
                endereco_selecionado = get_object_or_404(Endereco, id=endereco_id, cliente=cliente_logado)
            
            if data_envio_str:
                data_envio_obj = datetime.strptime(data_envio_str, '%Y-%m-%d').date()
            
        except (WfClient.DoesNotExist, Endereco.DoesNotExist, ValueError) as e:
            messages.error(request, f'Dados inválidos: {e}')
            return redirect('home')

        with transaction.atomic():
            if id_do_pedido:
                # Fluxo de Atualização de Rascunho
                pedido_rascunho.endereco = endereco_selecionado
                pedido_rascunho.data_envio_solicitada = data_envio_obj
                pedido_rascunho.frete_option = frete_option
                pedido_rascunho.nota_fiscal = nota_fiscal
                pedido_rascunho.observacao = observacao
                pedido_rascunho.status = 'PENDENTE'
                
                # RECALCULA O TOTAL ANTES DE SALVAR O RASCUNHO FINALIZADO
                pedido_rascunho.valor_total = pedido_rascunho.get_total_geral()
                pedido_rascunho.save()
                pedido_final = pedido_rascunho
            else:
                # Fluxo de Criação de Novo Pedido (Carrinho Manual)
                carrinho_da_sessao = request.session.get('carrinho', {})
                if not carrinho_da_sessao:
                    messages.error(request, 'Seu carrinho está vazio.')
                    return redirect('home')

                pedido = Pedido.objects.create(
                    cliente=cliente_logado,
                    empresa=cliente_logado.empresa,
                    endereco=endereco_selecionado,
                    data_envio_solicitada=data_envio_obj,
                    frete_option=frete_option,
                    nota_fiscal=nota_fiscal,
                    observacao=observacao,
                    criado_por=request.user,
                )
                
                # --- OTIMIZAÇÃO APLICADA: 1 SELECT IN (...) ---
                product_ids = [int(pid) for pid in carrinho_da_sessao.keys()]
                produtos_dict = Product.objects.in_bulk(product_ids, field_name='product_id')
                
                itens_para_criar = []
                for product_id_str, quantidade in carrinho_da_sessao.items():
                    product = produtos_dict.get(int(product_id_str))
                    if product:
                        # Prepara o objeto em memória (Não salva no banco ainda)
                        itens_para_criar.append(ItemPedido(
                            pedido=pedido,
                            produto=product,
                            quantidade=quantidade,
                            valor_unitario_sp=product.product_value_sp,
                            valor_unitario_es=product.product_value_es,
                        ))
                
                # --- OTIMIZAÇÃO APLICADA: 1 INSERT MÚLTIPLO ---
                if itens_para_criar:
                    ItemPedido.objects.bulk_create(itens_para_criar)
                
                # Calcula o total usando a função do model
                total_calculado = pedido.get_total_geral()
                
                # Salva o valor no banco de dados
                pedido.valor_total = total_calculado
                pedido.save()
                pedido_final = pedido
                
                if 'carrinho' in request.session:
                    del request.session['carrinho']

                # Limpa o carrinho do banco de dados
                carrinho_bd = Carrinho.objects.filter(cliente=cliente_logado).first()
                if carrinho_bd:
                    carrinho_bd.itens.all().delete()

        _notificar_novo_pedido(pedido_final)
        registrar_log(request, 'PEDIDO_CRIADO', f'Pedido #{pedido_final.id} criado via checkout', 'Pedido', pedido_final.id)
        return redirect('pedido_concluido', pedido_id=pedido_final.id)


    # Lógica para o método GET (primeira vez que a página é acessada)
    cliente_logado = None
    initial_data = {}
    
    if pedido_id_rascunho:
        if request.empresa:
            from django.db.models import Q
            pedido_para_finalizar = get_object_or_404(
                Pedido,
                Q(empresa=request.empresa) | Q(empresa__isnull=True, cliente__empresa=request.empresa),
                id=pedido_id_rascunho,
            )
        else:
            pedido_para_finalizar = get_object_or_404(Pedido, id=pedido_id_rascunho)
        cliente_logado = pedido_para_finalizar.cliente
        carrinho_detalhes = [
            {
                'product': item.produto,
                'quantidade': item.quantidade,
                'valor_unitario': item.valor_unitario_sp if cliente_logado.client_state.uf_name == 'SP' else item.valor_unitario_es,
                'valor_total': item.get_total(),
            }
            for item in pedido_para_finalizar.itens.all()
        ]
        total_geral = pedido_para_finalizar.get_total_geral()
        initial_data = {
            'endereco_selecionado': pedido_para_finalizar.endereco.id if pedido_para_finalizar.endereco else None,
            'data_envio': pedido_para_finalizar.data_envio_solicitada,
            'frete_option': pedido_para_finalizar.frete_option,
            'nota_fiscal': pedido_para_finalizar.nota_fiscal,
            'observacao': pedido_para_finalizar.observacao or cliente_logado.observacao_preferencia or '',
        }
    else:
        carrinho_da_sessao = request.session.get('carrinho', {})
        if not carrinho_da_sessao:
            messages.error(request, 'Seu carrinho está vazio.')
            return redirect('home')

        if not request.user.is_staff:
            try:
                cliente_logado = request.user.wfclient
            except WfClient.DoesNotExist:
                messages.error(request, 'Usuário não tem um cliente associado.')
                return redirect('home')
        else:
            messages.error(request, 'Usuários administrativos não podem criar pedidos por esta página.')
            return redirect('home')

        carrinho_detalhes = []
        total_geral = Decimal('0.00')

        # Batch fetch: 1 query para todos os produtos do carrinho
        product_ids = list(carrinho_da_sessao.keys())
        produtos_dict = {str(p.product_id): p for p in Product.objects.filter(product_id__in=product_ids)}

        for product_id, quantidade in carrinho_da_sessao.items():
            product = produtos_dict.get(str(product_id))
            if not product:
                messages.warning(request, f'Um produto (ID {product_id}) foi removido do catálogo e não pôde ser incluído no pedido.')
                continue
            uf_cliente = cliente_logado.client_state.uf_name if cliente_logado.client_state else 'SP'
            if uf_cliente == 'SP':
                valor_unitario = product.product_value_sp
            elif uf_cliente == 'ES':
                valor_unitario = product.product_value_es
            else:
                valor_unitario = product.product_value_sp
            if valor_unitario is None: valor_unitario = Decimal('0.00')
            valor_total_item = valor_unitario * quantidade
            total_geral += valor_total_item
            carrinho_detalhes.append({
                'product': product,
                'quantidade': quantidade,
                'valor_unitario': valor_unitario,
                'valor_total': valor_total_item,
            })

    _uf_display = cliente_logado.client_state.uf_name if (cliente_logado and cliente_logado.client_state) else None
    if _uf_display == 'SP':
        preco_exibido = 'sp'
    elif _uf_display == 'ES':
        preco_exibido = 'es'
    else:
        preco_exibido = 'sp'

    enderecos = Endereco.objects.filter(cliente=cliente_logado) if cliente_logado else Endereco.objects.none()

    # Preenche observação padrão se ainda não definida
    if not initial_data.get('observacao') and cliente_logado and cliente_logado.observacao_preferencia:
        initial_data['observacao'] = cliente_logado.observacao_preferencia

    contexto = {
        'titulo': 'Confirmação de Compra',
        'carrinho_detalhes': carrinho_detalhes,
        'total_geral': total_geral,
        'enderecos': enderecos,
        'preco_exibido': preco_exibido,
        'cliente_logado': cliente_logado,
        'initial_data': initial_data,
        'pedido_id_rascunho': pedido_id_rascunho,
    }

    return render(request, 'checkout.html', contexto)













# Fim Checkout

# Inicio salvar pedido

@login_required
def salvar_pedido(request):
    if request.method == 'POST':
        endereco_id = request.POST.get('endereco')
        data_envio_str = request.POST.get('data_envio') # Renomeei para evitar conflito

        try:
            cliente = request.user.wfclient
            endereco_selecionado = Endereco.objects.get(id=endereco_id, cliente=cliente)
            data_envio = datetime.datetime.strptime(data_envio_str, '%Y-%m-%d').date()
            
            carrinho_obj = Carrinho.objects.filter(cliente=cliente).first()
            itens_carrinho = carrinho_obj.itens.select_related('produto').all() if carrinho_obj else []

            with transaction.atomic():
                # 1. Cria o Pedido primeiro
                novo_pedido = Pedido.objects.create(
                    cliente=cliente,
                    empresa=cliente.empresa,
                    endereco=endereco_selecionado,
                    data_envio_solicitada=data_envio,
                    status='RASCUNHO',
                    criado_por=request.user,
                )

                # 2. Cria os Itens do Pedido (Loop)
                for item in itens_carrinho:
                    ItemPedido.objects.create(
                        pedido=novo_pedido,
                        produto=item.produto,
                        quantidade=item.quantidade,
                        valor_unitario_sp=item.produto.product_value_sp,
                        valor_unitario_es=item.produto.product_value_es,
                    )

                # 3. AGORA SIM: Atualiza o total e salva de novo
                # O get_total_geral() vai somar os ItemPedido que acabamos de criar acima
                novo_pedido.valor_total = novo_pedido.get_total_geral()
                novo_pedido.save() 

                # 4. Limpa o carrinho após salvar
                if carrinho_obj:
                    carrinho_obj.itens.all().delete()

            messages.success(request, 'Pedido realizado com sucesso!')
            return redirect('pedido_concluido') # Ou para a tela de checkout final

        except (Endereco.DoesNotExist, ValueError):
            messages.error(request, 'Endereço ou data de envio inválidos.')
            return redirect('checkout')
        except Exception as e:
            messages.error(request, f'Erro ao salvar pedido: {e}')
            return redirect('checkout')

    return redirect('checkout')

# Fim Salvar Pedido

@login_required
def pedido_concluido(request, pedido_id):
    try:
        cliente = request.user.wfclient
    except WfClient.DoesNotExist:
        return redirect('home')

    pedido = get_object_or_404(Pedido, id=pedido_id, cliente=cliente)
    itens = pedido.itens.select_related('produto').all()

    estado = cliente.client_state.uf_name if cliente.client_state else 'SP'
    itens_detalhados = []
    for item in itens:
        valor = item.valor_unitario_sp if estado == 'SP' else item.valor_unitario_es
        valor = valor or Decimal('0.00')
        itens_detalhados.append({
            'codigo': item.produto.product_code,
            'descricao': item.produto.product_description,
            'quantidade': item.quantidade,
            'valor_unitario': valor,
            'subtotal': valor * item.quantidade,
        })

    return render(request, 'pedido_concluido.html', {
        'pedido': pedido,
        'itens': itens_detalhados,
        'cliente': cliente,
        'estado': estado,
    })

@login_required
def historico_pedidos(request):
    try:
        cliente_logado = request.user.wfclient
    except WfClient.DoesNotExist:
        return redirect('home')

    pedidos_qs = Pedido.objects.filter(cliente=cliente_logado).order_by('-data_criacao')

    # Lógica de Paginação
    paginator = Paginator(pedidos_qs, 10)
    page = request.GET.get('page')

    try:
        pedidos_page = paginator.page(page)
    except PageNotAnInteger:
        pedidos_page = paginator.page(1)
    except EmptyPage:
        pedidos_page = paginator.page(paginator.num_pages)

    # A view agora só passa a QuerySet paginada para o template
    contexto = {
        'pedidos': pedidos_page, # Passamos o objeto paginado diretamente
        'titulo': 'Histórico de Pedidos',
    }
    return render(request, 'historico_pedidos.html', contexto)
#Inicio detalhes pedido

@login_required
def detalhes_pedido(request, pedido_id):
    try:
        # Acesso ao cliente logado
        cliente_logado = request.user.wfclient
    except WfClient.DoesNotExist:
        messages.error(request, "Erro: Usuário não tem um cliente associado.")
        return redirect('home')

    # Garante que o usuário logado só possa ver o próprio pedido, a menos que seja um admin.
    # A lógica aqui é mais robusta. Admins podem ver qualquer pedido.
    if request.user.is_staff:
        pedido = get_empresa_or_404(Pedido, request, id=pedido_id)
    else:
        pedido = get_object_or_404(Pedido, id=pedido_id, cliente=cliente_logado)

    # Obtém o estado do cliente do pedido para saber qual preço exibir
    estado_cliente = pedido.cliente.client_state.uf_name

    # Define qual preço será exibido
    preco_exibido = 'sp' if estado_cliente == 'SP' else 'es'

    itens_detalhes = []
    total_geral = Decimal('0.00')
    itens = ItemPedido.objects.filter(pedido=pedido)

    for item in itens:
        # Seleciona o valor unitário correto com base no estado do cliente
        if preco_exibido == 'sp':
            valor_unitario = item.valor_unitario_sp
        else:
            valor_unitario = item.valor_unitario_es

        # Garante que o valor não seja None
        if valor_unitario is None:
            valor_unitario = Decimal('0.00')

        valor_total_item = valor_unitario * item.quantidade
        total_geral += valor_total_item

        itens_detalhes.append({
            'item': item,
            'valor_unitario': valor_unitario,
            'valor_total': valor_total_item,
        })

    comentarios = ComentarioPedido.objects.filter(
        pedido=pedido, interno=False
    ).select_related('autor')

    contexto = {
        'titulo': f"Detalhes do Pedido #{pedido.id}",
        'pedido': pedido,
        'itens_detalhes': itens_detalhes,
        'total_geral': total_geral,
        'preco_exibido': preco_exibido,
        'comentarios': comentarios,
    }

    return render(request, 'detalhes_pedido.html', contexto)


# Fim detalhes pedido

# Configuração do locale para o formato brasileiro
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')


@staff_member_required
def dashboard_admin(request):
    filtro = request.GET.get('filtro')
    hoje = timezone.localdate()

    # Cache key única por empresa para evitar vazamento entre tenants
    empresa_key = request.empresa.slug if request.empresa else 'superuser'

    # =========================================================
    # 1. CÁLCULO DO FATURAMENTO REAL (CACHE: 15 MINUTOS)
    # =========================================================
    cache_fat_key = f'dashboard_faturamento_{empresa_key}'
    faturamento_formatado = cache.get(cache_fat_key)

    if not faturamento_formatado:
        total_faturamento_real = por_empresa(VendaReal.objects.all(), request).aggregate(
            total=Sum('Total')
        )['total'] or Decimal('0.00')
        faturamento_formatado = "{:,.2f}".format(float(total_faturamento_real)).replace(",", "X").replace(".", ",").replace("X", ".")
        cache.set(cache_fat_key, faturamento_formatado, 900)

    # =========================================================
    # 2. MÉTRICAS DO SITE E STATUS ERP (CACHE: 5 MINUTOS)
    # =========================================================
    cache_met_key = f'dashboard_metricas_{empresa_key}'
    metricas = cache.get(cache_met_key)

    if not metricas:
        pedidos_qs = por_empresa(Pedido.objects.all(), request)
        erp_qs = por_empresa(StatusPedidoERP.objects.all(), request)
        metricas = {
            'total_clientes': por_empresa(WfClient.objects.all(), request).count(),
            'total_pedidos': pedidos_qs.count(),
            'pendentes': pedidos_qs.filter(status='PENDENTE').count(),
            'concluidos': pedidos_qs.filter(status='FINALIZADO').count(),
            'orcamento': pedidos_qs.filter(status='ORCAMENTO').count(),
            'adc': pedidos_qs.filter(status='FINANCEIRO').count(),
            'separacao': pedidos_qs.filter(status='SEPARACAO').count(),
            'expedicao': pedidos_qs.filter(status='EXPEDICAO').count(),
            'atrasados': pedidos_qs.filter(status='ATRASADO').count(),
            'erp_total': erp_qs.count(),
            'erp_credito': erp_qs.filter(situacao__icontains='Crédito').count(),
            'erp_preco': erp_qs.filter(situacao__icontains='Preço').count(),
            'erp_separacao': erp_qs.filter(Q(situacao__icontains='Separação') | Q(situacao__icontains='Bloqueado')).count(),
            'erp_faturados': erp_qs.filter(situacao__icontains='Faturado').count(),
            'erp_expedidos': erp_qs.filter(expedido=True).count(),
        }
        cache.set(cache_met_key, metricas, 300)
    
    # =========================================================
    # 3. FILTRAGEM DE PEDIDOS (SEM CACHE - TEMPO REAL)
    # =========================================================
    if filtro == 'sincronizados':
        status_erp = ['SEPARACAO', 'EXPEDICAO', 'FINALIZADO']
        pedidos_qs = por_empresa(Pedido.objects.filter(
            status__in=status_erp,
            data_criacao__date=hoje
        ), request).order_by('-data_criacao')[:20]
    else:
        pedidos_qs = por_empresa(Pedido.objects.all(), request).order_by('-data_criacao')[:5]
    
    pedidos_com_total = []
    for pedido in pedidos_qs:
        total_pedido = float(pedido.get_total_geral() or 0)
        total_p_formatado = "{:,.2f}".format(total_pedido).replace(",", "X").replace(".", ",").replace("X", ".")
        
        pedidos_com_total.append({
            'id': pedido.id,
            'cliente': pedido.cliente,
            'data_criacao': pedido.data_criacao,
            'status': pedido.status,
            'status_display': pedido.get_status_display(),
            'total_str': total_p_formatado,
            'is_sincronizado': filtro == 'sincronizados'
        })

    # =========================================================
    # 4. OPORTUNIDADES DE RETORNO (WISHLIST) (CACHE: 15 MINUTOS)
    # =========================================================
    oportunidades_wishlist_cache = cache.get(f'dashboard_wishlist_{empresa_key}')
    
    # Usamos "is None" porque a lista de cache pode estar propositalmente vazia []
    if oportunidades_wishlist_cache is None:
        data_limite = hoje - timedelta(days=30)
        itens_pendentes = por_empresa(
            ItemPedidoIgnorado.objects.filter(
                notificado=False,
                descartado_pelo_cliente=False,
                motivo_erro__icontains="estoque",
                data_tentativa__gte=data_limite
            ), request, campo='cliente__empresa'
        ).select_related('cliente', 'cliente__client_state')

        oportunidades_wishlist = {}
        codigos_pendentes = itens_pendentes.values_list('codigo_produto', flat=True).distinct()
        produtos_dict = {p.product_code: p for p in por_empresa(Product.objects.filter(product_code__in=codigos_pendentes), request)}

        for item in itens_pendentes:
            produto = produtos_dict.get(item.codigo_produto)
            if not produto or not item.cliente:
                continue
                
            estado = item.cliente.client_state.uf_name
            preco_atual = getattr(produto, 'product_value_sp' if estado == 'SP' else 'product_value_es')
            
            if preco_atual and preco_atual > 0:
                c_id = item.cliente.client_id
                if c_id not in oportunidades_wishlist:
                    oportunidades_wishlist[c_id] = {
                        'cliente': item.cliente,
                        'produtos': []
                    }
                prods = oportunidades_wishlist[c_id]['produtos']
                existing = next((p for p in prods if p['codigo'] == produto.product_code), None)
                if existing:
                    existing['quantidade'] += item.quantidade_tentada or 0
                else:
                    prods.append({
                        'codigo': produto.product_code,
                        'descricao': produto.product_description,
                        'preco': float(preco_atual),
                        'quantidade': item.quantidade_tentada or 0,
                        'data': item.data_tentativa,
                    })
        
        # Converte o dicionário em lista para facilitar a renderização e o cache
        oportunidades_wishlist_cache = list(oportunidades_wishlist.values())
        cache.set(f'dashboard_wishlist_{empresa_key}', oportunidades_wishlist_cache, 900)

    # --- CONTEXTO ---
    contexto = {
        'titulo': 'Dashboard Administrativo',
        'total_clientes': metricas['total_clientes'],
        'total_pedidos': metricas['total_pedidos'],
        'total_vendas': faturamento_formatado,
        'pedidos_recentes': pedidos_com_total,
        
        # Enviando os novos números do ERP para a tela HTML
        'erp_total': metricas['erp_total'],
        'erp_credito': metricas['erp_credito'],
        'erp_preco': metricas['erp_preco'],
        'erp_separacao': metricas['erp_separacao'],
        'erp_faturados': metricas['erp_faturados'],
        'erp_expedidos': metricas['erp_expedidos'],
        
        'filtro_ativo': filtro,
        'oportunidades_wishlist': oportunidades_wishlist_cache, 
    }
    
    return render(request, 'dashboard.html', contexto)


# View para exportação de pedidos
@staff_member_required
def exportar_pedidos_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pedidos_recentes.xlsx"'
    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Pedidos Recentes"
        columns = ['ID do Pedido', 'Cliente', 'Data', 'Valor Total']
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            worksheet.cell(row=row_num, column=col_num, value=column_title)
        pedidos_recentes = por_empresa(Pedido.objects.all(), request).order_by('-data_criacao')[:10]
        for pedido in pedidos_recentes:
            row_num += 1
            data_sem_tz = timezone.localtime(pedido.data_criacao).replace(tzinfo=None)
            worksheet.cell(row=row_num, column=1, value=pedido.id)
            worksheet.cell(row=row_num, column=2, value=pedido.cliente.client_name)
            worksheet.cell(row=row_num, column=3, value=data_sem_tz)
            worksheet.cell(row=row_num, column=4, value=pedido.get_total_geral())
        workbook.save(response)
        return response
    except Exception as e:
        return redirect('dashboard_admin')

@login_required
def exportar_detalhes_pedido_excel(request, pedido_id):
    try:
        # Garante que o usuário logado só pode exportar o próprio pedido
        cliente_logado = request.user.wfclient
        pedido = get_object_or_404(Pedido, id=pedido_id, cliente=cliente_logado)
        uf_cliente = cliente_logado.client_state.uf_name
    except WfClient.DoesNotExist:
        # Se o usuário não tiver um cliente associado, redireciona
        return redirect('pedidos')
    except Pedido.DoesNotExist:
        # Se o pedido não existir ou não pertencer ao cliente, retorna um erro
        return redirect('pedidos')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="pedido_{pedido.id}.xlsx"'
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"Pedido #{pedido.id}"
    
    # Define as colunas dinamicamente com base no estado do cliente
    if uf_cliente == 'SP':
        columns = ['Código', 'Descrição', 'Quantidade', 'Valor Unitário (SP)', 'Subtotal']
        valor_key = 'product_value_sp'
    elif uf_cliente == 'ES':
        columns = ['Código', 'Descrição', 'Quantidade', 'Valor Unitário (ES)', 'Subtotal']
        valor_key = 'product_value_es'
    else:
        columns = ['Código', 'Descrição', 'Quantidade', 'Valor Unitário', 'Subtotal']
        valor_key = None
        
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=row_num, column=col_num, value=column_title)
        
    itens = ItemPedido.objects.filter(pedido=pedido)
    total_geral = 0
    
    for item in itens:
        row_num += 1
        
        # Acessa o valor do produto dinamicamente
        if valor_key:
            valor_unitario = getattr(item.produto, valor_key, 0)
        else:
            valor_unitario = 0

        # Calcula o subtotal diretamente
        subtotal = valor_unitario * item.quantidade
        total_geral += subtotal
        
        worksheet.cell(row=row_num, column=1, value=item.produto.product_code)
        worksheet.cell(row=row_num, column=2, value=item.produto.product_description)
        worksheet.cell(row=row_num, column=3, value=item.quantidade)
        worksheet.cell(row=row_num, column=4, value=valor_unitario)
        worksheet.cell(row=row_num, column=5, value=subtotal)
        
    row_num += 1
    worksheet.cell(row=row_num, column=4, value="Total Geral:")
    worksheet.cell(row=row_num, column=5, value=total_geral)
    
    workbook.save(response)
    return response


@login_required
@staff_member_required
def exportar_faltas_pedido_excel(request, pedido_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    pedido = get_empresa_or_404(Pedido, request, id=pedido_id)
    itens = ItemPedidoIgnorado.objects.filter(pedido=pedido).order_by('codigo_produto')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Faltas Pedido {pedido_id}'

    header_fill = PatternFill('solid', fgColor='6366F1')
    header_font = Font(bold=True, color='FFFFFF')
    headers = ['Cód. Produto', 'Descrição', 'Qtd. Tentada', 'Motivo']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_num, item in enumerate(itens, 2):
        ws.cell(row=row_num, column=1, value=item.codigo_produto)
        ws.cell(row=row_num, column=2, value=item.descricao_produto or '')
        ws.cell(row=row_num, column=3, value=item.quantidade_tentada or 0)
        ws.cell(row=row_num, column=4, value=item.motivo_erro)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="faltas_pedido_{pedido_id}.xlsx"'
    wb.save(response)
    return response


@staff_member_required
def exportar_detalhes_pedido_admin_excel(request, pedido_id):
    try:
        pedido = get_empresa_or_404(Pedido, request, id=pedido_id)
        cliente = pedido.cliente 
        uf_cliente = cliente.client_state.uf_name
    except Pedido.DoesNotExist:
        return redirect('dashboard_admin')

    # Formata a data para o padrão 'dd-mm-aaaa'
    data_formatada = pedido.data_criacao.strftime('%d-%m-%Y')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="pedido_{pedido.cliente.client_code}_{data_formatada}.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"Pedido #{pedido.id}"

    # Define as colunas dinamicamente com base no estado do cliente
    if uf_cliente == 'SP':
        columns = ['CÓDIGO', 'DESCRIÇÃO', 'GRUPO', 'MARCA', 'Valor Unitário (SP)', 'QUANTIDADE', 'SUBTOTAL']
        valor_key = 'valor_unitario_sp'
    elif uf_cliente == 'ES':
        columns = ['CÓDIGO', 'DESCRIÇÃO', 'GRUPO', 'MARCA', 'Valor Unitário (ES)', 'QUANTIDADE', 'SUBTOTAL']
        valor_key = 'valor_unitario_es'
    else:
        # Padrão caso o estado não seja SP ou ES
        columns = ['CÓDIGO', 'DESCRIÇÃO', 'GRUPO', 'MARCA', 'Valor Unitário', 'QUANTIDADE', 'SUBTOTAL']
        valor_key = 'valor_unitario_sp'

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=row_num, column=col_num, value=column_title)

    itens = ItemPedido.objects.filter(pedido=pedido)
    total_geral = 0

    for item in itens:
        row_num += 1

        valor_unitario = getattr(item, valor_key)

        if valor_unitario is None:
            valor_unitario = 0

        subtotal = valor_unitario * item.quantidade
        total_geral += subtotal

        worksheet.cell(row=row_num, column=1, value=item.produto.product_code)
        worksheet.cell(row=row_num, column=2, value=item.produto.product_description)
        worksheet.cell(row=row_num, column=3, value=item.produto.product_group)
        worksheet.cell(row=row_num, column=4, value=item.produto.product_brand)
        worksheet.cell(row=row_num, column=5, value=valor_unitario)
        worksheet.cell(row=row_num, column=6, value=item.quantidade)
        worksheet.cell(row=row_num, column=7, value=subtotal)

    row_num += 1
    worksheet.cell(row=row_num, column=6, value="Total Geral:")
    worksheet.cell(row=row_num, column=7, value=total_geral)

    workbook.save(response)
    return response

# Exportar para o cliente

@login_required
def exportar_detalhes_pedido_cliente_excel(request, pedido_id):
    try:
        # Garante que o usuário logado só possa exportar o próprio pedido
        cliente_logado = request.user.wfclient
        pedido = get_object_or_404(Pedido, id=pedido_id, cliente=cliente_logado)
        uf_cliente = cliente_logado.client_state.uf_name
    except WfClient.DoesNotExist:
        # Se o usuário não tiver um cliente associado, redireciona para a página de pedidos
        return redirect('pedidos')
    except Pedido.DoesNotExist:
        # Se o pedido não existir ou não pertencer ao cliente, retorna um erro 404
        return redirect('pedidos')

    data_formatada = pedido.data_criacao.strftime('%d-%m-%Y')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="pedido_{pedido.cliente.client_code}_{data_formatada}.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"Pedido #{pedido.id}"

    # Define as colunas dinamicamente com base no estado do cliente
    if uf_cliente == 'SP':
        columns = ['Código', 'Descrição', 'Quantidade', 'Valor Unitário (SP)', 'Subtotal']
        valor_key = 'valor_unitario_sp'  # ✅ Alterado para o campo do ItemPedido
    elif uf_cliente == 'ES':
        columns = ['Código', 'Descrição', 'Quantidade', 'Valor Unitário (ES)', 'Subtotal']
        valor_key = 'valor_unitario_es'  # ✅ Alterado para o campo do ItemPedido
    else:
        columns = ['Código', 'Descrição', 'Quantidade', 'Valor Unitário', 'Subtotal']
        valor_key = 'valor_unitario_sp'  # Padrão para SP

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=row_num, column=col_num, value=column_title)

    itens = ItemPedido.objects.filter(pedido=pedido)
    total_geral = 0

    for item in itens:
        row_num += 1

        # ✅ Acessa o valor do item de pedido, não do produto
        valor_unitario = getattr(item, valor_key)
        if valor_unitario is None:
            valor_unitario = 0

        subtotal = valor_unitario * item.quantidade
        total_geral += subtotal

        worksheet.cell(row=row_num, column=1, value=item.produto.product_code)
        worksheet.cell(row=row_num, column=2, value=item.produto.product_description)
        worksheet.cell(row=row_num, column=3, value=item.quantidade)
        worksheet.cell(row=row_num, column=4, value=valor_unitario)
        worksheet.cell(row=row_num, column=5, value=subtotal)

    row_num += 1
    worksheet.cell(row=row_num, column=4, value="Total Geral:")
    worksheet.cell(row=row_num, column=5, value=total_geral)

    workbook.save(response)
    return response


# GARANTA QUE APENAS ESTA VERSÃO DA FUNÇÃO EXISTA NO SEU views.py
# NOVO ARQUIVO views.py (versão limpa)

@staff_member_required
def todos_os_pedidos(request):
    if request.empresa:
        from django.db.models import Q
        pedidos_qs = Pedido.objects.filter(
            Q(empresa=request.empresa) | Q(empresa__isnull=True, cliente__empresa=request.empresa)
        ).order_by('-data_criacao')
    else:
        pedidos_qs = Pedido.objects.all().order_by('-data_criacao')

    paginator = Paginator(pedidos_qs, 20)
    page_number = request.GET.get('page', 1)
    
    try:
        pedidos_paginados = paginator.page(page_number)
    except PageNotAnInteger:
        pedidos_paginados = paginator.page(1)
    except EmptyPage:
        pedidos_paginados = paginator.page(paginator.num_pages)

    hoje = timezone.localdate()
    for pedido in pedidos_paginados:
        if pedido.data_envio_solicitada and pedido.data_envio_solicitada < hoje and pedido.status not in ['FINALIZADO', 'CANCELADO']:
            pedido.is_atrasado = True
        else:
            pedido.is_atrasado = False

    contexto = {
        'titulo': 'Todos os Pedidos',
        'pedidos': pedidos_paginados,
    }

    # --- ESTA É A PARTE QUE REATIVA O SCROLL INFINITO ---
    if request.htmx:
        # Quando o usuário rola a página, o HTMX faz uma requisição especial.
        # Esta linha responde a essa requisição enviando APENAS as novas linhas da tabela.
        return render(request, '_pedidos_rows.html', contexto)

    # Quando o usuário carrega a página pela primeira vez, esta linha é executada.
    return render(request, 'todos_os_pedidos.html', contexto)
# Deixe o resto do arquivo em branco por enquanto


'''
@staff_member_required
def todos_os_pedidos(request):
    pedidos_qs = Pedido.objects.all().order_by('-data_criacao')

    paginator = Paginator(pedidos_qs, 20)
    page_number = request.GET.get('page', 1)
    
    try:
        pedidos_paginados = paginator.page(page_number)
    except PageNotAnInteger:
        pedidos_paginados = paginator.page(1)
    except EmptyPage:
        pedidos_paginados = paginator.page(paginator.num_pages)

    hoje = timezone.localdate()
    for pedido in pedidos_paginados:
        if pedido.data_envio_solicitada and pedido.data_envio_solicitada < hoje and pedido.status not in ['FINALIZADO', 'CANCELADO']:
            pedido.is_atrasado = True
        else:
            pedido.is_atrasado = False

    contexto = {
        'titulo': 'Todos os Pedidos',
        'pedidos': pedidos_paginados,
    }

    # Verifica se a requisição é do HTMX (para o scroll infinito)
    if request.htmx:
        # Se for, renderiza apenas o template parcial com as novas linhas
        return render(request, '_pedidos_rows.html', contexto) # <-- CAMINHO CORRIGIDO

    # Se for uma requisição normal, renderiza a página completa
    return render(request, 'todos_os_pedidos.html', contexto) # <-- CAMINHO CORRIGIDO'''

# Modelo de Pedido
    
@staff_member_required
def atualizar_status_pedido(request, pedido_id):
    if request.method == 'POST':
        pedido = get_empresa_or_404(Pedido, request, id=pedido_id)
        novo_status = request.POST.get('status')
        # CORREÇÃO AQUI: Atualize a lista de status permitidos
        if novo_status in ['PENDENTE', 'ORCAMENTO', 'FINANCEIRO', 'SEPARACAO', 'EXPEDICAO', 'FINALIZADO', 'CANCELADO']:
            pedido.status = novo_status
            pedido.save()
            messages.success(request, f'Status do Pedido #{pedido.id} alterado para {novo_status} com sucesso!')
    return redirect('todos_os_pedidos')

@login_required
def upload_foto_perfil(request):
    if request.method == 'POST':
        foto = request.FILES.get('foto')
        if foto:
            # Membro (PerfilUsuario)
            if hasattr(request.user, 'perfil'):
                perfil = request.user.perfil
                if perfil.foto_perfil:
                    perfil.foto_perfil.delete(save=False)
                perfil.foto_perfil.save(foto.name, foto, save=True)
            # Cliente (WfClient)
            elif hasattr(request.user, 'wfclient'):
                cliente = request.user.wfclient
                if cliente.foto_perfil:
                    cliente.foto_perfil.delete(save=False)
                cliente.foto_perfil.save(foto.name, foto, save=True)
    if hasattr(request.user, 'perfil'):
        return redirect('perfil_representante')
    return redirect('editar_perfil')


@login_required
def editar_perfil(request):
    # Membros (is_staff) não têm WfClient — exibe só o card de foto
    if hasattr(request.user, 'perfil'):
        return render(request, 'editar_perfil.html', {
            'titulo': 'Meu Perfil',
            'perfil_membro': request.user.perfil,
        })

    try:
        cliente = request.user.wfclient
    except WfClient.DoesNotExist:
        return redirect('home')

    if request.method == 'POST':
        form = EnderecoForm(request.POST)
        if form.is_valid():
            novo_endereco = form.save(commit=False)
            novo_endereco.cliente = cliente
            novo_endereco.save()
            return redirect('editar_perfil')
    else:
        form = EnderecoForm()

    enderecos = Endereco.objects.filter(cliente=cliente)

    contexto = {
        'form': form,
        'enderecos': enderecos,
        'titulo': 'Gerenciar Endereços',
        'cliente': cliente,
    }
    return render(request, 'editar_perfil.html', contexto)


@login_required
def gerenciar_enderecos(request):
    try:
        cliente = request.user.wfclient
    except WfClient.DoesNotExist:
        return redirect('home')
        
    enderecos = Endereco.objects.filter(cliente=cliente)
    
    if request.method == 'POST':
        form = EnderecoForm(request.POST)
        if form.is_valid():
            novo_endereco = form.save(commit=False)
            novo_endereco.cliente = cliente
            novo_endereco.save()
            return redirect('gerenciar_enderecos')
    else:
        form = EnderecoForm()

    contexto = {
        'titulo': 'Gerenciar Endereços',
        'enderecos': enderecos,
        'form': form
    }
    return render(request, 'gerenciar_enderecos.html', contexto)


@login_required
def editar_endereco(request, endereco_id):
    try:
        endereco = get_object_or_404(Endereco, id=endereco_id, cliente=request.user.wfclient)
    except WfClient.DoesNotExist:
        return redirect('home')

    if request.method == 'POST':
        form = EnderecoForm(request.POST, instance=endereco)
        if form.is_valid():
            form.save()
            return redirect('editar_perfil')
    else:
        form = EnderecoForm(instance=endereco)

    contexto = {
        'form': form,
        'titulo': 'Editar Endereço',
    }
    return render(request, 'editar_endereco.html', contexto)


# A página para exibir o formulário de upload
@staff_member_required
def pagina_upload(request):
    return render(request, 'upload_planilha.html')

@staff_member_required
def processar_upload(request):
    if request.method == 'POST':
        planilha_es_file = request.FILES.get('planilha_es')
        planilha_sp_file = request.FILES.get('planilha_sp')

        if not planilha_es_file or not planilha_sp_file:
            messages.error(request, 'Por favor, selecione ambas as planilhas.')
            return redirect('pagina_upload')

        try:
            # Mapeamento de aliases de colunas
            aliases = {
                'codigo':   ['CÓDIGO', 'CODIGO', 'PRODUTO', 'COD'],
                'descricao': ['DESCRIÇÃO', 'DESCRICAO', 'NOME DO PRODUTO'],
                'grupo':    ['GRUPO', 'CATEGORIA'],
                'marca':    ['MARCA', 'FABRICANTE'],
                'tabela':   ['TABELA', 'PREÇO NOVO', 'PRECO', 'VALOR']
            }

            def encontrar_linha_cabecalho(file, max_rows=15):
                """
                Escaneia as primeiras linhas da planilha e retorna o índice (0-based)
                da linha que contém ao mesmo tempo um alias de 'codigo' e de 'tabela'.
                Faz reset do ponteiro do arquivo após a leitura.
                """
                df_scan = pd.read_excel(file, header=None, nrows=max_rows, dtype=str)
                for i, row in df_scan.iterrows():
                    vals = [str(v).upper().strip() for v in row if pd.notnull(v) and str(v).strip()]
                    tem_codigo = any(alias in v for alias in aliases['codigo'] for v in vals)
                    tem_tabela = any(alias in v for alias in aliases['tabela'] for v in vals)
                    if tem_codigo and tem_tabela:
                        file.seek(0)
                        return i
                file.seek(0)
                return 0  # fallback: primeira linha

            def encontrar_coluna(df, nomes_possiveis):
                for nome in nomes_possiveis:
                    if nome in df.columns:
                        return nome
                return None

            # 1. Descobre qual linha é o cabeçalho em cada planilha
            header_es = encontrar_linha_cabecalho(planilha_es_file)
            header_sp = encontrar_linha_cabecalho(planilha_sp_file)

            # 2. Leitura com o cabeçalho correto
            df_es_raw = pd.read_excel(planilha_es_file, header=header_es, dtype=str)
            df_sp_raw = pd.read_excel(planilha_sp_file, header=header_sp, dtype=str)

            # Normaliza nomes de colunas: remove espaços e converte para maiúsculo
            df_es_raw.columns = [str(c).upper().strip() for c in df_es_raw.columns]
            df_sp_raw.columns = [str(c).upper().strip() for c in df_sp_raw.columns]

            # Processa as colunas da Planilha ES
            col_cod_es = encontrar_coluna(df_es_raw, aliases['codigo'])
            col_tab_es = encontrar_coluna(df_es_raw, aliases['tabela'])
            
            # Processa as colunas da Planilha SP
            col_cod_sp = encontrar_coluna(df_sp_raw, aliases['codigo'])
            col_tab_sp = encontrar_coluna(df_sp_raw, aliases['tabela'])

            # Validação: Se não achar Código e Preço, avisa o usuário sem dar Erro 500
            if not col_cod_es or not col_tab_es:
                messages.error(request, 'Planilha ES inválida: É obrigatório ter as colunas de "Código/Produto" e "Tabela/Preço".')
                return redirect('pagina_upload')
            if not col_cod_sp or not col_tab_sp:
                messages.error(request, 'Planilha SP inválida: É obrigatório ter as colunas de "Código/Produto" e "Tabela/Preço".')
                return redirect('pagina_upload')

            # 3. Montagem dos DataFrames limpos
            # ES
            df_es = pd.DataFrame()
            df_es['product_code'] = df_es_raw[col_cod_es].str.strip()
            df_es['product_value_es'] = pd.to_numeric(df_es_raw[col_tab_es].str.replace(',', '.'), errors='coerce').fillna(0)
            
            col_desc = encontrar_coluna(df_es_raw, aliases['descricao'])
            col_grupo = encontrar_coluna(df_es_raw, aliases['grupo'])
            col_marca = encontrar_coluna(df_es_raw, aliases['marca'])
            
            if col_desc: df_es['product_description'] = df_es_raw[col_desc].str.strip()
            if col_grupo: df_es['product_group'] = df_es_raw[col_grupo].str.strip()
            if col_marca: df_es['product_brand'] = df_es_raw[col_marca].str.strip()

            # SP
            df_sp = pd.DataFrame()
            df_sp['product_code'] = df_sp_raw[col_cod_sp].str.strip()
            df_sp['product_value_sp'] = pd.to_numeric(df_sp_raw[col_tab_sp].str.replace(',', '.'), errors='coerce').fillna(0)

            # 4. Merge das planilhas (how='outer' garante que não perca itens que só estão em uma planilha)
            df_final = pd.merge(df_es, df_sp, on='product_code', how='outer')

            # Detecta status de estoque antes do replace(nan→None)
            # Produto com pelo menos um preço válido (>0) é DISPONIVEL, caso contrário SEM_ESTOQUE
            val_es_num = pd.to_numeric(df_final['product_value_es'], errors='coerce').fillna(0)
            val_sp_num = pd.to_numeric(df_final['product_value_sp'], errors='coerce').fillna(0)
            df_final['status_estoque'] = ((val_es_num > 0) | (val_sp_num > 0)).map({True: 'DISPONIVEL', False: 'SEM_ESTOQUE'})

            df_final = df_final.replace({np.nan: None})
            df_final = df_final.drop_duplicates(subset=['product_code'])
            
            # Remove linhas vazias/lixo da planilha
            df_final = df_final[df_final['product_code'].notnull()]
            df_final = df_final[df_final['product_code'] != ""]

            # --- 5. OTIMIZAÇÃO DE BANCO E PREVENÇÃO DE PERDA DE DADOS ---
            codigos_na_planilha = df_final['product_code'].tolist()
            hoje = date.today()
            
            produtos_existentes = {
                p.product_code: p for p in Product.objects.filter(
                    product_code__in=codigos_na_planilha,
                    empresa=request.empresa,
                )
            }

            produtos_para_criar = []
            produtos_para_atualizar = []
            historico_para_criar = []

            for _, row in df_final.iterrows():
                codigo = str(row['product_code']).strip()

                # Extração segura dos dados
                desc_planilha = row.get('product_description')
                grupo_planilha = row.get('product_group')
                marca_planilha = row.get('product_brand')

                val_es = row.get('product_value_es', 0)
                val_sp = row.get('product_value_sp', 0)
                estoque = row.get('status_estoque', 'DISPONIVEL')

                if codigo in produtos_existentes:
                    obj = produtos_existentes[codigo]

                    preco_sp_antigo = obj.product_value_sp
                    preco_es_antigo = obj.product_value_es

                    # ATUALIZAÇÃO CONDICIONAL: Só atualiza se a célula da planilha não for nula/vazia
                    if desc_planilha and str(desc_planilha).strip().lower() not in ['nan', 'none', '']:
                        obj.product_description = str(desc_planilha).strip()[:255]

                    if grupo_planilha and str(grupo_planilha).strip().lower() not in ['nan', 'none', '']:
                        obj.product_group = str(grupo_planilha).strip()[:100]

                    if marca_planilha and str(marca_planilha).strip().lower() not in ['nan', 'none', '']:
                        obj.product_brand = str(marca_planilha).strip()[:100]

                    # Atualização de preços (só atualiza se veio com valor real — evita zerar preço existente)
                    if val_sp: obj.product_value_sp = val_sp
                    if val_es: obj.product_value_es = val_es

                    obj.status = 'PENDENTE'
                    obj.status_estoque = estoque
                    obj.date_product = hoje

                    # Registra histórico apenas se algum preço mudou
                    preco_sp_novo = Decimal(str(val_sp)) if val_sp else preco_sp_antigo
                    preco_es_novo = Decimal(str(val_es)) if val_es else preco_es_antigo
                    if preco_sp_novo != preco_sp_antigo or preco_es_novo != preco_es_antigo or estoque != obj.status_estoque:
                        historico_para_criar.append(HistoricoPreco(
                            product_code=codigo,
                            product_description=obj.product_description,
                            product_value_sp=preco_sp_novo,
                            product_value_es=preco_es_novo,
                            status_estoque=estoque,
                            empresa=request.empresa,
                        ))

                    produtos_para_atualizar.append(obj)
                else:
                    # Criação de um produto novo
                    desc_nova = str(desc_planilha).strip()[:255] if desc_planilha and str(desc_planilha).strip().lower() not in ['nan', 'none', ''] else ""
                    produtos_para_criar.append(Product(
                        product_code=codigo,
                        product_description=desc_nova,
                        product_group=str(grupo_planilha).strip()[:100] if grupo_planilha and str(grupo_planilha).strip().lower() not in ['nan', 'none', ''] else "",
                        product_brand=str(marca_planilha).strip()[:100] if marca_planilha and str(marca_planilha).strip().lower() not in ['nan', 'none', ''] else "",
                        product_value_sp=val_sp if val_sp is not None else 0,
                        product_value_es=val_es if val_es is not None else 0,
                        status='PENDENTE',
                        status_estoque=estoque,
                        date_product=hoje,
                        empresa=request.empresa,
                    ))
                    # Registra histórico do primeiro preço
                    historico_para_criar.append(HistoricoPreco(
                        product_code=codigo,
                        product_description=desc_nova,
                        product_value_sp=Decimal(str(val_sp)) if val_sp else None,
                        product_value_es=Decimal(str(val_es)) if val_es else None,
                        status_estoque=estoque,
                        empresa=request.empresa,
                    ))

            # 6. Gravação Atômica e em Massa
            with transaction.atomic():
                if produtos_para_criar:
                    Product.objects.bulk_create(produtos_para_criar, batch_size=500)

                if produtos_para_atualizar:
                    Product.objects.bulk_update(
                        produtos_para_atualizar,
                        fields=['product_description', 'product_group', 'product_brand', 'product_value_sp', 'product_value_es', 'status', 'status_estoque', 'date_product'],
                        batch_size=500
                    )

                if historico_para_criar:
                    HistoricoPreco.objects.bulk_create(historico_para_criar, batch_size=500)
            
            # Invalida cache da home para esta empresa
            empresa_key = request.empresa.slug if request.empresa else 'global'
            cache.delete(f'home_products_{empresa_key}')

            messages.success(request, f'Processamento concluído: {len(produtos_para_criar)} novos e {len(produtos_para_atualizar)} atualizados.')
            return redirect('pagina_upload')

        except Exception as e:
            messages.error(request, f'Erro crítico no processamento: {e}')
            return redirect('pagina_upload')

    return render(request, 'upload_planilha.html')

# Início da função gerar_pedido
@login_required
def gerar_pedido(request):
    if request.method == 'POST':
        cart_data_json = request.POST.get('cart_data', '{}')
        
        if not cart_data_json:
            messages.error(request, 'Erro ao processar os dados do carrinho.')
            return redirect('home')

        try:
            carrinho_da_sessao = json.loads(cart_data_json)
            carrinho_filtrado = {}
            for product_id_str, quantidade_str in carrinho_da_sessao.items():
                try:
                    quantidade = int(quantidade_str)
                    if quantidade > 0:
                        carrinho_filtrado[product_id_str] = quantidade
                except (ValueError, TypeError):
                    continue
            
            if not carrinho_filtrado:
                messages.error(request, 'Seu carrinho está vazio ou contém itens inválidos.')
                return redirect('home')

            request.session['carrinho'] = carrinho_filtrado
            
            # --- AQUI ESTÁ A CORREÇÃO ---
            # Verifica se há um ID de pedido rascunho na sessão
            pedido_id_rascunho = request.session.get('pedido_id_rascunho')

            if pedido_id_rascunho:
                # Se houver, redireciona para o checkout com o ID do rascunho
                return redirect('checkout_rascunho', pedido_id_rascunho=pedido_id_rascunho)
            else:
                # Se não houver, redireciona para o checkout padrão
                return redirect('checkout')

        except json.JSONDecodeError:
            messages.error(request, 'Erro ao processar os dados do carrinho.')
            return redirect('home')
    
    return redirect('home')
# Fim da função gerar_pedido


# views.py
# ... (seus imports) ...

# views.py

@login_required
def atualizar_rascunho(request):
    if request.method == 'POST':
        carrinho_da_sessao = request.session.get('carrinho', {})
        pedido_id_rascunho = request.session.get('pedido_id_rascunho')

        if not pedido_id_rascunho:
            messages.error(request, 'Não há pedido rascunho na sessão para atualizar.')
            return redirect('carrinho')

        try:
            pedido = get_empresa_or_404(Pedido, request, id=pedido_id_rascunho)
        except Exception:
            messages.error(request, 'Pedido rascunho não encontrado.')
            return redirect('home')

        with transaction.atomic():
            # Apaga todos os itens do pedido antigo
            pedido.itens.all().delete()

            # Batch fetch: 1 query para todos os produtos do carrinho
            product_ids = list(carrinho_da_sessao.keys())
            produtos_dict = {str(p.product_id): p for p in Product.objects.filter(product_id__in=product_ids)}

            # Adiciona os itens atualizados do carrinho na sessão
            for product_id, quantidade in carrinho_da_sessao.items():
                product = produtos_dict.get(str(product_id))
                if not product:
                    continue
                
                # ... (sua lógica de valor unitário) ...

                ItemPedido.objects.create(
                    pedido=pedido,
                    produto=product,
                    quantidade=quantidade,
                    valor_unitario_sp=product.product_value_sp,
                    valor_unitario_es=product.product_value_es,
                )

        messages.success(request, 'Pedido rascunho atualizado com sucesso!')
        
        # AQUI ESTÁ A LINHA DE REDIRECIONAMENTO QUE PRECISA SER VERIFICADA
        return redirect('checkout_rascunho', pedido_id_rascunho=pedido_id_rascunho)
    
    return redirect('carrinho')

# Em views.py
# GARANTA QUE APENAS ESTA VERSÃO DA FUNÇÃO EXISTA NO ARQUIVO

@staff_member_required
def detalhes_pedido_admin(request, pedido_id):
    try:
        pedido = get_empresa_or_404(Pedido, request, id=pedido_id)
        
        estado_cliente = pedido.cliente.client_state.uf_name
        preco_exibido = 'sp' if estado_cliente == 'SP' else 'es'
        
        itens_detalhes = []
        total_geral = Decimal('0.00')
        itens = ItemPedido.objects.filter(pedido=pedido)

        for item in itens:
            valor_unitario = item.valor_unitario_sp if estado_cliente == 'SP' else item.valor_unitario_es
            
            if valor_unitario is None:
                valor_unitario = Decimal('0.00')

            # Garante que a quantidade não seja None para o cálculo
            quantidade = item.quantidade if item.quantidade is not None else 0
            valor_total_item = valor_unitario * quantidade
            total_geral += valor_total_item

            itens_detalhes.append({
                'item': item,
                'valor_unitario': valor_unitario,
                'valor_total': valor_total_item,
            })
        
        # Lógica para verificar se o pedido está atrasado
        hoje = timezone.localdate()
        is_atrasado = False
        if pedido.data_envio_solicitada and pedido.data_envio_solicitada < hoje and pedido.status not in ['FINALIZADO', 'CANCELADO']:
            is_atrasado = True
        
        # Dicionário de contexto que é enviado para o template
        comentarios = ComentarioPedido.objects.filter(
            pedido=pedido
        ).select_related('autor')

        itens_ignorados = ItemPedidoIgnorado.objects.filter(pedido=pedido).order_by('codigo_produto')

        contexto = {
            'titulo': f"Detalhes do Pedido #{pedido.id}",
            'pedido': pedido,
            'itens_detalhes': itens_detalhes,
            'total_geral': total_geral,
            'preco_exibido': preco_exibido,
            'is_atrasado': is_atrasado,
            'comentarios': comentarios,
            'itens_ignorados': itens_ignorados,
        }

        return render(request, 'detalhes_pedido.html', contexto)

    except Pedido.DoesNotExist:
        messages.error(request, "Erro: Pedido não encontrado.")
        return redirect('todos_os_pedidos')
    except WfClient.DoesNotExist:
        messages.error(request, "Erro: Cliente associado ao pedido não encontrado.")
        return redirect('todos_os_pedidos')


@login_required
def comentar_pedido(request, pedido_id):
    if request.method != 'POST':
        return redirect('home')

    texto = request.POST.get('texto', '').strip()
    if not texto:
        return redirect(request.POST.get('next', 'home'))

    # Staff acessa via empresa; cliente só acessa o próprio pedido
    if request.user.is_staff:
        pedido = get_empresa_or_404(Pedido, request, id=pedido_id)
        interno = request.POST.get('interno') == '1'
        next_url = reverse('detalhes_pedido_admin', args=[pedido_id])
    else:
        try:
            cliente = request.user.wfclient
        except WfClient.DoesNotExist:
            return redirect('home')
        pedido = get_object_or_404(Pedido, id=pedido_id, cliente=cliente)
        interno = False
        next_url = reverse('detalhes_pedido', args=[pedido_id])

    ComentarioPedido.objects.create(
        pedido=pedido,
        autor=request.user,
        texto=texto,
        interno=interno,
    )
    return redirect(next_url)


@staff_member_required
def pedidos_para_hoje(request):
    """
    View que filtra e exibe os pedidos agendados para a data atual
    e também os pedidos atrasados que ainda estão pendentes.
    """
    # Pega a data de hoje, respeitando o timezone do seu projeto
    hoje = timezone.localdate()

    # --- Consulta 1: Pedidos agendados para HOJE que ainda precisam de ação ---
    pedidos_hoje = por_empresa(Pedido.objects.filter(
        data_envio_solicitada=hoje
    ), request).exclude(status__in=['FINALIZADO', 'CANCELADO'])

    # --- Consulta 2: Pedidos ATRASADOS que ainda precisam de ação ---
    pedidos_atrasados = por_empresa(Pedido.objects.filter(
        data_envio_solicitada__lt=hoje
    ), request).exclude(
        status__in=['FINALIZADO', 'CANCELADO']
    ).order_by('data_envio_solicitada')

    # Monta o contexto com as duas listas de pedidos
    context = {
        'pedidos_hoje': pedidos_hoje,
        'pedidos_atrasados': pedidos_atrasados,
        'data_hoje': hoje,
        'titulo': 'Pedidos para Saída' # Adicionado para o título da página
    }
    
    return render(request, 'pedidos/pedidos_hoje.html', context)


@staff_member_required
def gerar_pedido_manual(request):
  
    # 1. Lógica de Limpeza (DEVE vir antes de qualquer outra coisa)
    if 'limpar' in request.GET:
        cliente_id = request.GET.get('cliente')
        # Redireciona para a mesma página apenas com o ID do cliente, limpando os filtros
        return redirect(f"{reverse('gerar_pedido_manual')}?cliente={cliente_id}")

    cliente_selecionado = None
    # Usamos request.GET para manter o estado do formulário de seleção de cliente
    form_cliente = SelectClientForm(request.GET or None)
    form_cliente.fields['cliente'].queryset = por_empresa(WfClient.objects.all(), request).order_by('client_name')
    product_list = []
    query_params = request.GET.copy()
    preco_exibido = 'todos'
    initial_data = {}

    # Remove o parâmetro de página para não quebrar os filtros de busca
    if 'page' in query_params:
        query_params.pop('page')

    # 2. Identificação do Cliente
    # Tentamos pegar o cliente tanto pelo form quanto direto pelo GET (para o redirecionamento do 'limpar')
    cliente_id_get = request.GET.get('cliente')
    
    if form_cliente.is_valid() and form_cliente.cleaned_data.get('cliente'):
        cliente_selecionado = form_cliente.cleaned_data['cliente']
    elif cliente_id_get:
        cliente_selecionado = get_empresa_or_404(WfClient, request, pk=cliente_id_get)
        # Atualiza o form para mostrar o cliente correto no select
        form_cliente = SelectClientForm(initial={'cliente': cliente_selecionado})

    context = {
        'form_cliente': form_cliente,
        'cliente_selecionado': cliente_selecionado,
        'initial_data': initial_data,
        'query_params': query_params,
        'preco_exibido': preco_exibido,
    }
    
    # 3. Processamento de Produtos se houver cliente
    if cliente_selecionado:
        # Preferências do Cliente
        initial_data['frete_preferencia'] = cliente_selecionado.frete_preferencia
        initial_data['nota_fiscal_preferencia'] = cliente_selecionado.nota_fiscal_preferencia
        
        endereco_padrao = Endereco.objects.filter(cliente=cliente_selecionado, is_default=True).first()
        if endereco_padrao:
            initial_data['endereco_padrao_id'] = endereco_padrao.id

        enderecos_do_cliente = Endereco.objects.filter(cliente=cliente_selecionado)
        
        # Filtros de Produtos
        hoje = timezone.localdate()
        products = por_empresa(Product.objects.filter(date_product=hoje), request).order_by('product_code')
        if not products.exists():
            from django.db.models import Max
            ultima_data = Product.objects.aggregate(d=Max('date_product'))['d']
            if ultima_data:
                products = por_empresa(Product.objects.filter(date_product=ultima_data), request).order_by('product_code')

        # Captura dos filtros do GET
        codigo = request.GET.get('codigo')
        descricao = request.GET.get('descricao')
        grupo = request.GET.get('grupo')
        marca = request.GET.get('marca')

        if codigo:
            products = products.filter(product_code__icontains=codigo)
        if descricao:
            products = products.filter(product_description__icontains=descricao)
        if grupo:
            products = products.filter(product_group__icontains=grupo)
        if marca:
            products = products.filter(product_brand__icontains=marca)
        
        # Lógica de Preço baseada no Estado
        if not cliente_selecionado.client_state:
            messages.error(request, f'Cliente {cliente_selecionado.client_name} não possui estado (SP/ES) cadastrado.')
            return redirect('gerar_pedido_manual')
        estado_cliente = cliente_selecionado.client_state.uf_name.upper()
        preco_exibido = estado_cliente.lower()

        # Formatação de valores
        for product in products:
            if estado_cliente == 'SP':
                if product.product_value_sp:
                    product.valor_sp_formatado = f"R$ {product.product_value_sp:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                else:
                    product.valor_sp_formatado = "N/A"
            elif estado_cliente == 'ES':
                if product.product_value_es:
                    product.valor_es_formatado = f"R$ {product.product_value_es:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                else:
                    product.valor_es_formatado = "N/A"
            
        # Paginação
        paginator = Paginator(products, 30)
        page_number = request.GET.get('page')
        product_list = paginator.get_page(page_number)

        context.update({
            'enderecos': enderecos_do_cliente,
            'preco_exibido': preco_exibido,
            'product_list': product_list,
            'initial_data': initial_data,
        })
    
    return render(request, 'gerar_pedido_manual.html', context)

# --- SUBSTITUA A VIEW PROCESSAR_PEDIDO_MANUAL EXISTENTE POR ESTA VERSÃO OTIMIZADA ---

@staff_member_required
def processar_pedido_manual(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        cart_data_json = request.POST.get('cart_data')
        endereco_id = request.POST.get('endereco_selecionado')
        data_envio = request.POST.get('data_envio')
        frete_option = request.POST.get('frete_option')
        nota_fiscal = request.POST.get('nota_fiscal')
        usuario_logado = request.user
        observacao = request.POST.get('observacao')
        
        fretes_sem_endereco = ['ONIBUS', 'RETIRADA']
        endereco_selecionado = None
        
        if frete_option not in fretes_sem_endereco:
            if not endereco_id:
                messages.error(request, 'Por favor, selecione um endereço válido.')
                return redirect('gerar_pedido_manual')
            try:
                endereco_selecionado = get_empresa_or_404(Endereco, request, id=endereco_id, cliente=cliente_selecionado)
            except Exception:
                messages.error(request, 'Endereço inválido.')
                return redirect('gerar_pedido_manual')

        try:
            cliente_selecionado = get_empresa_or_404(WfClient, request, client_id=cliente_id)
            data_envio = datetime.strptime(data_envio, '%Y-%m-%d').date()
            cart_data = json.loads(cart_data_json)

            if not cart_data:
                messages.error(request, 'Não há itens para gerar o pedido.')
                return redirect('gerar_pedido_manual')

            with transaction.atomic():
                pedido_criado = Pedido.objects.create(
                    cliente=cliente_selecionado,
                    empresa=cliente_selecionado.empresa,
                    endereco=endereco_selecionado,
                    data_envio_solicitada=data_envio,
                    frete_option=frete_option,
                    nota_fiscal=nota_fiscal,
                    status='PENDENTE',
                    criado_por=usuario_logado,
                    valor_total=Decimal('0.00'),
                    observacao=observacao
                )
                
                # --- OTIMIZAÇÃO APLICADA: 1 SELECT IN (...) ---
                product_ids = [int(pid) for pid in cart_data.keys()]
                produtos_dict = Product.objects.in_bulk(product_ids, field_name='product_id')
                
                itens_para_criar = []
                if not cliente_selecionado.client_state:
                    raise ValueError(f'Cliente {cliente_selecionado.client_name} não possui estado (SP/ES) cadastrado.')
                estado_cliente = cliente_selecionado.client_state.uf_name

                for product_id_str, quantidade in cart_data.items():
                    produto = produtos_dict.get(int(product_id_str))
                    
                    if not produto:
                        messages.warning(request, f'Produto com ID {product_id_str} não encontrado e foi ignorado.')
                        continue
                    
                    # Lógica para salvar o valor no campo correto e manter o outro nulo
                    if estado_cliente == 'SP':
                        val_sp = produto.product_value_sp
                        val_es = None
                    elif estado_cliente == 'ES':
                        val_sp = None
                        val_es = produto.product_value_es
                    else:
                        messages.warning(request, f'Produto {produto.product_code} não pôde ser adicionado. Estado inválido.')
                        continue

                    itens_para_criar.append(ItemPedido(
                        pedido=pedido_criado,
                        produto=produto,
                        quantidade=quantidade,
                        valor_unitario_sp=val_sp,
                        valor_unitario_es=val_es
                    ))

                # --- OTIMIZAÇÃO APLICADA: 1 INSERT MÚLTIPLO ---
                if itens_para_criar:
                    ItemPedido.objects.bulk_create(itens_para_criar)

                # Cálculo e atualização do Valor Total
                total_pedido = ItemPedido.objects.filter(pedido=pedido_criado).aggregate(
                    total_sp=Sum(F('quantidade') * F('valor_unitario_sp')),
                    total_es=Sum(F('quantidade') * F('valor_unitario_es'))
                )
                
                if estado_cliente == 'SP':
                    valor_final_do_pedido = total_pedido['total_sp']
                else:
                    valor_final_do_pedido = total_pedido['total_es']
                
                pedido_criado.valor_total = valor_final_do_pedido if valor_final_do_pedido is not None else Decimal('0.00')
                pedido_criado.save()

            _notificar_novo_pedido(pedido_criado)
            messages.success(request, f'Pedido #{pedido_criado.id} criado com sucesso para o cliente {cliente_selecionado.client_name}!')
            return redirect(reverse('gerar_pedido_manual') + '?pedido_gerado=sucesso')

        except (WfClient.DoesNotExist, Endereco.DoesNotExist, ValueError) as e:
            messages.error(request, f'Dados inválidos. Erro: {e}')
            return redirect('gerar_pedido_manual')
        except json.JSONDecodeError:
            messages.error(request, 'Erro nos dados do pedido. Tente novamente.')
            return redirect('gerar_pedido_manual')

    return redirect('gerar_pedido_manual')


def normalize_text(text):
    """Normaliza o texto, remove acentos, e converte para minúsculas."""
    if not isinstance(text, str):
        return ""
    text = text.lower().strip()
    return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
# --- FIM DA CORREÇÃO ---


@staff_member_required
def upload_pedido(request):
    """
    Processa o upload de uma planilha de pedido, cria um Pedido (Rascunho),
    salva os itens válidos em ItemPedido e os itens com falha em ItemPedidoIgnorado.
    Ignora linhas de somatório (TOTAL, SUBTOTAL) silenciosamente.
    """
    cliente_selecionado = None
    form_cliente = SelectClientForm(request.GET or None)
    form_cliente.fields['cliente'].queryset = por_empresa(WfClient.objects.all(), request).order_by('client_code')
    initial_data = {}
    upload_form = None

    # Lógica de seleção prévia do cliente (via GET)
    if form_cliente.is_valid():
        cliente_selecionado = form_cliente.cleaned_data.get('cliente')
        if cliente_selecionado:
            initial_data['frete_preferencia'] = cliente_selecionado.frete_preferencia
            initial_data['nota_fiscal_preferencia'] = cliente_selecionado.nota_fiscal_preferencia
            initial_data['observacao_preferencia'] = cliente_selecionado.observacao_preferencia
            endereco_padrao = Endereco.objects.filter(cliente=cliente_selecionado, is_default=True).first()
            if endereco_padrao:
                initial_data['endereco_selecionado'] = endereco_padrao.id
    
    if request.method == 'POST':
        cliente_id_post = request.POST.get('cliente_id')
        if not cliente_id_post:
            messages.error(request, 'Por favor, selecione um cliente.')
            return redirect('upload_pedido')

        cliente_para_validacao = get_empresa_or_404(WfClient, request, pk=cliente_id_post)
        
        form = UploadPedidoForm(request.POST, request.FILES)
        enderecos_do_cliente = Endereco.objects.filter(cliente=cliente_para_validacao)
        form.fields['endereco_selecionado'].queryset = enderecos_do_cliente

        if form.is_valid():
            try:
                planilha_pedido = request.FILES.get('planilha_pedido')
                if not planilha_pedido:
                    messages.error(request, 'Nenhum arquivo de planilha foi selecionado.')
                    return redirect('upload_pedido')
                    
                # 1. Leitura do arquivo (Excel ou CSV)
                try:
                    if planilha_pedido.name.endswith('.csv'):
                        df_list = [pd.read_csv(planilha_pedido)]
                    else:
                        # Detecta linha de cabeçalho dinamicamente (planilhas com linha de data no topo)
                        header_row = 0
                        _aliases_cod = ['CODIGO', 'COD', 'CÓDIGO', 'CÓD']
                        _aliases_qtd = ['QUANTIDADE', 'QTD', 'QTDE', 'QNT', 'QUANT']
                        df_scan = pd.read_excel(planilha_pedido, header=None, nrows=15, dtype=str)
                        for _i, _row in df_scan.iterrows():
                            _vals_norm = [normalize_text(str(v)).upper().strip() for v in _row if pd.notnull(v) and str(v).strip()]
                            _vals_raw  = [str(v).upper().strip() for v in _row if pd.notnull(v) and str(v).strip()]
                            _all_vals  = _vals_norm + _vals_raw
                            _tem_cod = any(alias in v for alias in _aliases_cod for v in _all_vals)
                            _tem_qtd = any(alias in v for alias in _aliases_qtd for v in _all_vals)
                            if _tem_cod and _tem_qtd:
                                header_row = _i
                                break
                        planilha_pedido.seek(0)
                        xls_data = pd.read_excel(planilha_pedido, sheet_name=None, header=header_row)
                        df_list = list(xls_data.values())
                except Exception:
                    messages.error(request, 'Arquivo inválido ou corrompido. Envie um arquivo Excel (.xlsx) ou CSV válido.')
                    return redirect('upload_pedido')

                if not df_list:
                     messages.error(request, 'A planilha de upload está vazia.')
                     return redirect('upload_pedido')

                df = pd.concat(df_list, ignore_index=True).dropna(how='all')

                if df.empty:
                    messages.error(request, 'A planilha de upload não contém dados após a leitura.')
                    return redirect('upload_pedido')

                # 2. Normalização e Mapeamento de Colunas
                df.columns = [normalize_text(col) for col in df.columns]

                expected_cols = {
                    'codigo': ['codigo', 'código', 'cod'],
                    'quantidade': ['quantidade', 'qtd', 'qtde'],
                    'descricao': ['descricao', 'descrição', 'produto', 'nome', 'description']
                }

                col_mapping = {
                    key: next((c for c in values if c in df.columns), None)
                    for key, values in expected_cols.items()
                }

                if not col_mapping['codigo'] or not col_mapping['quantidade']:
                    messages.error(request, "A planilha deve ter colunas para 'código' e 'quantidade'.")
                    return redirect('upload_pedido')

                # Filtra apenas linhas com quantidade preenchida e positiva antes do loop
                # (evita iterar 16k linhas quando só ~50 têm qty > 0)
                col_qtd_raw = col_mapping['quantidade']
                df = df[pd.to_numeric(df[col_qtd_raw], errors='coerce').fillna(0) > 0]

                if df.empty:
                    messages.error(request, 'Nenhuma linha com quantidade positiva encontrada na planilha.')
                    return redirect('upload_pedido')

                # Agrupa por código somando quantidades (mesmo produto em múltiplas linhas/abas)
                col_cod_raw = col_mapping['codigo']
                df = df.dropna(subset=[col_cod_raw])
                df[col_cod_raw] = df[col_cod_raw].astype(str).str.strip()
                _other_cols = {c: 'first' for c in df.columns if c not in [col_cod_raw, col_qtd_raw]}
                df = df.groupby(col_cod_raw, as_index=False).agg({col_qtd_raw: 'sum', **_other_cols})
                
                with transaction.atomic():
                    # 3. Criação do Pedido Rascunho
                    novo_pedido = Pedido.objects.create(
                        cliente=cliente_para_validacao,
                        empresa=cliente_para_validacao.empresa,
                        endereco=form.cleaned_data.get('endereco_selecionado'),
                        data_criacao=timezone.now(),
                        data_envio_solicitada=form.cleaned_data['data_expedicao'],
                        frete_option=form.cleaned_data['frete_option'],
                        nota_fiscal=form.cleaned_data['nota_fiscal'],
                        status='RASCUNHO',
                        criado_por=request.user,
                        observacao=form.cleaned_data['observacao_preferencia'],
                    )
                    
                    erros_texto = []
                    itens_pedido_para_criar = []
                    itens_ignorados_db = []
                    total_valor_pedido = Decimal('0.0')

                    # Busca produtos em lote (1 query)
                    from django.db.models import Max as _Max
                    _base = por_empresa(Product.objects, request)
                    _latest = _base.aggregate(d=_Max('date_product'))['d']
                    produtos_atuais = {p.product_code: p for p in _base.filter(date_product=_latest)}

                    # Constantes fora do loop
                    col_cod  = col_mapping['codigo']
                    col_qtd  = col_mapping['quantidade']
                    col_desc = col_mapping.get('descricao')
                    termos_ignorar = {'TOTAL', 'SUBTOTAL', 'GERAL', 'VALOR TOTAL'}
                    if not cliente_para_validacao.client_state:
                        raise ValueError(f'Cliente {cliente_para_validacao.client_name} não possui estado (SP/ES) cadastrado.')
                    regiao      = cliente_para_validacao.client_state.uf_name
                    valor_field = 'product_value_sp' if regiao == 'SP' else 'product_value_es'

                    # 4. Processamento linha a linha (to_dict é ~10x mais rápido que iterrows)
                    for index, row in enumerate(df.to_dict('records'), start=2):

                        codigo_produto_raw = row.get(col_cod)
                        if pd.isna(codigo_produto_raw) if not isinstance(codigo_produto_raw, str) else not str(codigo_produto_raw).strip():
                            continue

                        codigo_produto = str(codigo_produto_raw).strip()

                        if any(termo in codigo_produto.upper() for termo in termos_ignorar):
                            continue

                        quantidade_raw = row.get(col_qtd)
                        descricao_excel = row.get(col_desc, 'Descrição não informada na planilha') if col_desc else 'Descrição não informada na planilha'

                        # --- Validação A: Quantidade Nula ---
                        if quantidade_raw is None or (not isinstance(quantidade_raw, str) and pd.isnull(quantidade_raw)):
                            continue

                        # --- Validação B: Quantidade Numérica ---
                        try:
                            quantidade = int(quantidade_raw)
                        except (ValueError, TypeError):
                            msg = "Quantidade inválida (não-numérica)"
                            erros_texto.append(f"Produto '{codigo_produto}' na linha {index}: {msg}.")
                            itens_ignorados_db.append(ItemPedidoIgnorado(
                                pedido=novo_pedido,
                                cliente=cliente_para_validacao,
                                codigo_produto=codigo_produto,
                                descricao_produto=str(descricao_excel),
                                quantidade_tentada=0,
                                motivo_erro=msg
                            ))
                            continue

                        # --- Validação C: Quantidade Zero ou Negativa ---
                        if quantidade <= 0:
                            continue

                        # --- Validação D: Produto Existe no Catálogo? ---
                        produto = produtos_atuais.get(codigo_produto)
                        if not produto:
                            msg = "Não encontrado no catálogo"
                            erros_texto.append(f"Produto '{codigo_produto}' na linha {index}: {msg}.")
                            itens_ignorados_db.append(ItemPedidoIgnorado(
                                pedido=novo_pedido,
                                cliente=cliente_para_validacao,
                                codigo_produto=codigo_produto,
                                descricao_produto=str(descricao_excel),
                                quantidade_tentada=quantidade,
                                motivo_erro=msg
                            ))
                            continue

                        # --- Validação E: Estoque e Preço ---
                        valor_unitario = getattr(produto, valor_field)
                        
                        if valor_unitario is None or valor_unitario <= 0:
                            msg = "Produto indisponível no estoque/tabela"
                            erros_texto.append(f"Produto '{codigo_produto}' na linha {index}: {msg}.")
                            itens_ignorados_db.append(ItemPedidoIgnorado(
                                pedido=novo_pedido,
                                cliente=cliente_para_validacao,
                                codigo_produto=codigo_produto,
                                descricao_produto=produto.product_description, # Usa descrição oficial
                                quantidade_tentada=quantidade,
                                motivo_erro=msg
                            ))
                            continue

                        # Se passou por tudo, adiciona aos itens válidos
                        total_valor_pedido += valor_unitario * Decimal(quantidade)
                        itens_pedido_para_criar.append(ItemPedido(
                            pedido=novo_pedido,
                            produto=produto,
                            quantidade=quantidade,
                            valor_unitario_sp=produto.product_value_sp,
                            valor_unitario_es=produto.product_value_es,
                        ))
                    
                    # 5. Finalização
                    
                    # Se a planilha não gerou NADA (nem válido, nem erro - vazia de dados úteis)
                    if not itens_pedido_para_criar and not itens_ignorados_db:
                        messages.error(request, "Nenhum dado processável encontrado na planilha.")
                        novo_pedido.delete()
                        return redirect('upload_pedido')
                    
                    # Salva Itens Válidos em lote
                    if itens_pedido_para_criar:
                        ItemPedido.objects.bulk_create(itens_pedido_para_criar)
                    
                    # Salva Itens Ignorados em lote
                    if itens_ignorados_db:
                        ItemPedidoIgnorado.objects.bulk_create(itens_ignorados_db)
                    
                    # Mensagens e Logs
                    if erros_texto:
                        novo_pedido.erros_upload = '\n'.join(erros_texto)
                        
                        erros_msg = 'Alguns itens foram ignorados:\n' + '\n'.join(erros_texto[:5])
                        if len(erros_texto) > 5:
                            erros_msg += f'\n...e mais {len(erros_texto) - 5} erros.'
                        messages.warning(request, f"Pedido criado parcialmente. {erros_msg}")
                    else:
                        messages.success(request, f"Itens carregados com sucesso. Por favor, confira os dados e finalize o pedido.")
                    
                    novo_pedido.valor_total = total_valor_pedido
                    novo_pedido.save()
                    
                    return redirect('checkout_rascunho', pedido_id_rascunho=novo_pedido.id)
            
            except Exception as e:
                messages.error(request, f"Erro crítico ao processar a planilha: {e}")
                # Garante que não fica lixo no banco se der erro fatal
                if 'novo_pedido' in locals():
                    novo_pedido.delete()
                upload_form = form
        else:
            upload_form = form
    else:
        # GET request
        upload_form = UploadPedidoForm(initial=initial_data)
        if cliente_selecionado:
            enderecos_do_cliente = Endereco.objects.filter(cliente=cliente_selecionado)
            upload_form.fields['endereco_selecionado'].queryset = enderecos_do_cliente

    context = {
        'form_cliente': form_cliente,
        'cliente_selecionado': cliente_selecionado,
        'upload_form': upload_form,
    }
    
    return render(request, 'upload_pedido.html', context)


@staff_member_required
def pedidos_em_andamento(request):
    """
    Exibe uma lista de pedidos com o status 'ORCAMENTO' (rascunhos)
    para que o administrador possa visualizá-los e continuar a edição.
    """
    pedidos_rascunho = por_empresa(Pedido.objects.filter(
        status='RASCUNHO'
    ), request).order_by('-data_criacao')
    
    context = {
        'titulo': 'Pedidos em Andamento (Rascunhos)',
        'pedidos': pedidos_rascunho,
    }
    
    # AQUI ESTÁ O AJUSTE. O caminho aponta direto para o arquivo.
    return render(request, 'pedidos_em_andamento.html', context)


# No seu arquivo views.py

# views.py

# ... (restante do código)

@staff_member_required
def continuar_pedido(request, pedido_id):
    try:
        pedido = get_empresa_or_404(Pedido, request, id=pedido_id, status='RASCUNHO')
    except Pedido.DoesNotExist:
        messages.error(request, 'O pedido especificado não é um rascunho válido.')
        return redirect('pedidos_em_andamento')

    # ... (restante da sua lógica de sessão) ...

    # 1. Popula o carrinho da sessão com os itens do pedido rascunho
    carrinho_da_sessao = {}
    for item in pedido.itens.select_related('produto').all():
        carrinho_da_sessao[str(item.produto.product_id)] = item.quantidade
    request.session['carrinho'] = carrinho_da_sessao
    
    # 2. Salva o ID do pedido na sessão
    request.session['pedido_id_rascunho'] = pedido.id

    messages.info(request, f'Você está continuando a edição do Pedido #{pedido.id}.')

    # 3. Redireciona para o checkout usando a URL específica com o ID
    return redirect('checkout_rascunho', pedido_id_rascunho=pedido.id)

# ... (restante do código)




@staff_member_required
def upload_orcamento_pdf(request, pedido_id):
    if request.method == 'POST':
        pedido = get_empresa_or_404(Pedido, request, id=pedido_id)
        orcamento_file = request.FILES.get('orcamento_pdf_file')

        if orcamento_file:
            client_code = pedido.cliente.client_code
            hoje = timezone.localdate().strftime('%d-%m-%Y')
            novo_nome = f'orcamento_{client_code}_{pedido.id}_{hoje}.pdf'

            # Apaga o arquivo anterior se existir
            if pedido.orcamento_pdf:
                pedido.orcamento_pdf.delete(save=False)

            # Salva via storage backend (local em dev, R2 em produção)
            pedido.orcamento_pdf.save(novo_nome, orcamento_file, save=True)

            messages.success(request, f'Orçamento PDF "{novo_nome}" enviado com sucesso!')
        else:
            messages.error(request, 'Nenhum arquivo foi selecionado.')
            
        return redirect(reverse('detalhes_pedido_admin', args=[pedido_id]))

    return redirect('dashboard_admin')


def exportar_detalhes_pedido_publico_excel(request, pedido_id):
    """
    Exporta os detalhes de um pedido para uma planilha Excel.
    A planilha é personalizada para o estado do cliente.
    """
    pedido = get_empresa_or_404(Pedido, request, id=pedido_id)
    itens_pedido = ItemPedido.objects.filter(pedido=pedido).select_related('produto')
    
    # Obtém o estado do cliente para definir a lógica de exportação
    uf_cliente = pedido.cliente.client_state.uf_name

    # Define as colunas e a chave de valor dinamicamente
    if uf_cliente == 'SP':
        columns = ['CÓDIGO', 'DESCRIÇÃO', 'GRUPO', 'MARCA', 'Valor Unitário (SP)', 'QUANTIDADE', 'SUBTOTAL']
        valor_key = 'valor_unitario_sp'
    elif uf_cliente == 'ES':
        columns = ['CÓDIGO', 'DESCRIÇÃO', 'GRUPO', 'MARCA', 'Valor Unitário (ES)', 'QUANTIDADE', 'SUBTOTAL']
        valor_key = 'valor_unitario_es'
    else:
        # Padrão caso o estado não seja SP ou ES
        columns = ['CÓDIGO', 'DESCRIÇÃO', 'GRUPO', 'MARCA', 'Valor Unitário', 'QUANTIDADE', 'SUBTOTAL']
        valor_key = 'valor_unitario_sp'

    # Criação do DataFrame com os dados dos itens
    data = []
    total_geral = 0

    for item in itens_pedido:
        # Acessa o valor do item de pedido usando a chave definida
        valor_unitario = getattr(item, valor_key)
        if valor_unitario is None:
            valor_unitario = 0

        subtotal = float(item.get_total())

        data.append({
            'CÓDIGO': item.produto.product_code,
            'DESCRIÇÃO': item.produto.product_description,
            'GRUPO': item.produto.product_group,
            'MARCA': item.produto.product_brand,
            columns[4]: float(valor_unitario),
            'QUANTIDADE': item.quantidade,
            'SUBTOTAL': subtotal
        })

        total_geral += subtotal

    df = pd.DataFrame(data)
    df = df[columns]  # Garante a ordem das colunas

    df.loc[len(df)] = ['', '', '', '', '', 'Total Geral:', total_geral]

    # Criação da resposta HTTP
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Itens do Pedido')
        
        # Opcional: Ajuste automático da largura das colunas (melhoria visual)
        worksheet = writer.sheets['Itens do Pedido']
        worksheet.set_column('A:A', 15) # Código
        worksheet.set_column('B:B', 40) # Descrição
        worksheet.set_column('C:D', 20) # Grupo e Marca
        worksheet.set_column('E:G', 15) # Qtd e Valores

    output.seek(0)

    # Renomeia o arquivo com as novas informações
    data_hoje = date.today().strftime('%d-%m-%Y')
    filename = f"pedido_{pedido.cliente.client_code}_{data_hoje}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


'''
def encurtar_url(long_url):
    """
    Encurta uma URL usando a API da TinyURL.
    """
    api_url = f"http://tinyurl.com/api-create.php?url={long_url}"
    try:
        response = requests.get(api_url, timeout=10)
        if response.status_code == 200:
            return response.text
    except requests.RequestException:
        pass
    return long_url


@staff_member_required
def enviar_whatsapp(request, pedido_id):
    pedido = get_empresa_or_404(Pedido, request, id=pedido_id)

    # Construir o link de download público da planilha
    link_download_excel = request.build_absolute_uri(
        reverse('exportar_detalhes_pedido_publico_excel', args=[pedido.id])
    )

    # Encurtar a URL
    link_encurtado = encurtar_url(link_download_excel)

    # Informações básicas do pedido
    mensagem_base = (
        f"*Dados do Pedido*\n\n"
        f"*Codigo Interno:* {pedido.id}\n"
        f"*Código do Cliente:* {pedido.cliente.client_code}\n"
        f"*Razão Social:* {pedido.cliente.client_name}\n"
        f"*Data da Expedição:* {pedido.data_envio_solicitada.strftime('%d/%m/%Y')}\n"
        f"*Opção de Frete:* {pedido.get_frete_option_display()}\n"
        f"*OBS:* {pedido.observacao}\n"
    )

    # Adiciona o endereço de entrega apenas se não for ÔNIBUS ou RETIRADA
    fretes_com_endereco = ['SEDEX', 'CORREIOS', 'TRANSPORTADORA']
    endereco_texto = ""
    if pedido.frete_option in fretes_com_endereco and pedido.endereco:
        endereco = pedido.endereco
        endereco_texto = (
            f"*Endereço de Entrega:* "
            f"{endereco.logradouro}, {endereco.bairro}, {endereco.numero} "
            f"{endereco.cidade} - {endereco.estado} (CEP: {endereco.cep})\n"
        )
    else:
        endereco_texto = ""

    # Conclui a mensagem com a opção de nota fiscal e adiciona o link de download encurtado
    mensagem_final = (
        f"{mensagem_base}"
        f"{endereco_texto}"
        f"*Opção de Nota Fiscal:* {pedido.get_nota_fiscal_display()}\n\n"
        f"*Download da Planilha de Itens:*\n"
        f"{link_encurtado}"
    )

    # ✅ Use a função 'quote' para codificar a URL
    link_whatsapp = f"https://wa.me/5516991273974?text={quote(mensagem_final)}"

    return redirect(link_whatsapp)
'''


''' Função original
@staff_member_required
def enviar_whatsapp(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    # 1. Construir o link de download público da planilha (isso permanece igual)
    link_download_excel = request.build_absolute_uri(
        reverse('exportar_detalhes_pedido_publico_excel', args=[pedido.id])
    )

    # 2. A linha do encurtador foi REMOVIDA
    # link_encurtado = encurtar_url(link_download_excel)

    # Informações básicas do pedido
    mensagem_base = (
        f"*Dados do Pedido*\n\n"
        f"*Codigo Interno:* {pedido.id}\n"
        f"*Código do Cliente:* {pedido.cliente.client_code}\n"
        f"*Razão Social:* {pedido.cliente.client_name}\n"
        f"*Data da Expedição:* {pedido.data_envio_solicitada.strftime('%d/%m/%Y')}\n"
        f"*Opção de Frete:* {pedido.get_frete_option_display()}\n"
        
        
    )

    # Adiciona o endereço de entrega (lógica inalterada)
    fretes_com_endereco = ['SEDEX', 'CORREIOS', 'TRANSPORTADORA']
    endereco_texto = ""
    if pedido.frete_option in fretes_com_endereco and pedido.endereco:
        endereco = pedido.endereco
        endereco_texto = (
            f"*Endereço de Entrega:* "
            f"{endereco.logradouro}, {endereco.bairro}, {endereco.numero} "
            f"{endereco.cidade} - {endereco.estado} (CEP: {endereco.cep})\n"
        )
    else:
        endereco_texto = ""

    # 3. Conclui a mensagem usando o link original e completo
    mensagem_final = (
        f"{mensagem_base}"
        f"{endereco_texto}"
        f"*Opção de Nota Fiscal:* {pedido.get_nota_fiscal_display()}\n"
        f"*Valor total:* R${pedido.valor_total:,.2f}\n"
        f"*OBS:* {pedido.observacao}\n\n"
        f"*Download da Planilha de Itens:*\n" 
        f"{link_download_excel}"  # <- MUDANÇA AQUI: usando o link direto
    )

    # Codifica a mensagem para a URL do WhatsApp (lógica inalterada)
    link_whatsapp = f"https://wa.me/5516991273974?text={quote(mensagem_final)}"

    return redirect(link_whatsapp)
'''


@staff_member_required
def enviar_whatsapp(request, pedido_id):
    pedido = get_empresa_or_404(Pedido, request, id=pedido_id)

    # 1. Link para a nova planilha simplificada (Cód, Qtd, Preço)
    link_download_excel = request.build_absolute_uri(
        reverse('exportar_detalhes_pedido_whatsapp_excel', args=[pedido.id])
    )

    # 2. Construção da mensagem formatada conforme seu exemplo
    mensagem_corpo = (
        f"*Dados do Pedido*\n\n"
        f"*Codigo Interno:* {pedido.id}\n"
        f"*Código do Cliente:* {pedido.cliente.client_code}\n"
        f"*Razão Social:* {pedido.cliente.client_name}\n"
        f"*Data da Expedição:* {pedido.data_envio_solicitada.strftime('%d/%m/%Y')}\n"
        f"*Opção de Frete:* {pedido.get_frete_option_display()}\n"
    )

    # 3. Lógica de Endereço (apenas se houver frete que exija entrega)
    fretes_com_endereco = ['SEDEX', 'CORREIOS', 'TRANSPORTADORA']
    if pedido.frete_option in fretes_com_endereco and pedido.endereco:
        end = pedido.endereco
        mensagem_corpo += (
            f"*Endereço de Entrega:* {end.logradouro}, {end.bairro}, {end.numero} "
            f"{end.cidade} - {end.estado} (CEP: {end.cep})\n"
        )

    # 4. Adicionando Nota Fiscal, Valor Total e OBS
    mensagem_final = (
        f"{mensagem_corpo}"
        f"*Opção de Nota Fiscal:* {pedido.get_nota_fiscal_display()}\n" # <-- AQUI A NOTA FISCAL
        f"*Valor total:* R$ {pedido.valor_total:,.2f}\n".replace(",", "X").replace(".", ",").replace("X", ".") +
        f"*OBS:* {pedido.observacao}\n\n"
        f"*Link para Digitação (Planilha):*\n"
        f"{link_download_excel}"
    )

    # 5. Codifica e Redireciona
    link_whatsapp = f"https://wa.me/5516991273974?text={quote(mensagem_final)}"
    return redirect(link_whatsapp)




@staff_member_required
def pedidos_atrasados_view(request):
    # Pega a data de hoje para comparação
    hoje = date.today()

    # Esta é a consulta principal:
    pedidos_atrasados = por_empresa(Pedido.objects.filter(
        data_envio_solicitada__lt=hoje
    ).exclude(
        status__in=['FINALIZADO', 'CANCELADO']
    ), request).order_by('data_envio_solicitada')

    contexto = {
        'titulo': 'Pedidos Atrasados',
        'pedidos_atrasados': pedidos_atrasados
    }

    return render(request, 'pedidos_atrasados.html', contexto)

# Em seu arquivo views.py

@staff_member_required
def marcar_pedido_finalizado(request, pedido_id):
    # Apenas aceita requisições POST por segurança
    if request.method == 'POST':
        # Pega o pedido ou retorna um erro 404 se não existir
        pedido = get_empresa_or_404(Pedido, request, id=pedido_id)
        
        # Altera o status para 'FINALIZADO'
        pedido.status = 'FINALIZADO'
        pedido.save()
        registrar_log(request, 'PEDIDO_FINALIZADO', f'Pedido #{pedido.id} marcado como FINALIZADO', 'Pedido', pedido.id)
        messages.success(request, f'O Pedido #{pedido.id} foi marcado como FINALIZADO.')
    
    # Redireciona o usuário de volta para a página de detalhes de onde ele veio
    return redirect('detalhes_pedido_admin', pedido_id=pedido_id)

# views.py (ADICIONE AO SEU ARQUIVO)


@staff_member_required
def analise_dados_dashboard(request):
    # --- 1. Definição de Período ---
    periodo_geral_solicitado = 'periodo_geral' in request.GET
    data_fim_str = request.GET.get('data_fim')
    data_inicio_str = request.GET.get('data_inicio')
    
    def parse_date(date_str, default):
        if not date_str: return default
        for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
            try: return datetime.strptime(date_str, fmt).date()
            except ValueError: continue
        return default

    data_fim = timezone.localdate() if not data_fim_str or periodo_geral_solicitado else parse_date(data_fim_str, timezone.localdate())
    data_inicio = datetime(2000, 1, 1).date() if periodo_geral_solicitado else parse_date(data_inicio_str, data_fim - timedelta(days=90))

    # --- 2. FATURAMENTO REAL ERP (VendaReal) ---
    # Aqui corrigimos o erro dos 915k aplicando o exclude de segurança
    vendas_erp_qs = por_empresa(VendaReal.objects.exclude(Produto_Codigo__icontains='TOTAL'), request)
    
    if not periodo_geral_solicitado:
        vendas_erp_qs = vendas_erp_qs.filter(Emissao__range=(data_inicio, data_fim))
    
    total_faturamento_erp = vendas_erp_qs.aggregate(total=Sum('Total'))['total'] or Decimal('0.00')

    # --- 3. Sugestões ERP para Cliente Específico ---
    cliente_id_analise = request.GET.get('cliente')
    sugestoes_erp = []
    lista_clientes_filtro = por_empresa(WfClient.objects.only('client_id', 'client_name'), request).order_by('client_name')

    if cliente_id_analise:
        try:
            cliente_obj = por_empresa(WfClient.objects.all(), request).get(pk=cliente_id_analise)
            sugestoes_erp = por_empresa(VendaReal.objects.filter(
                Codigo_Cliente=cliente_obj.client_code
            ).exclude(Produto_Codigo__icontains='TOTAL'), request).values(
                'Produto_Codigo', 'Produto_Descricao'
            ).annotate(
                total_qtd=Sum('Quantidade'),
                vezes_faturado=Count('id')
            ).order_by('-total_qtd')[:10]
        except WfClient.DoesNotExist:
            pass

    # --- 4. QuerySet Base do SITE (ItemPedido) ---
    CAMPO_DATA_FILTRO = 'pedido__data_envio_solicitada'
    base_queryset = ItemPedido.objects.select_related(
        'pedido', 'pedido__cliente', 'produto', 'pedido__cliente__client_state'
    ).exclude(pedido__status='CANCELADO')
    if request.empresa:
        base_queryset = base_queryset.filter(
            Q(pedido__empresa=request.empresa) |
            Q(pedido__empresa__isnull=True, pedido__cliente__empresa=request.empresa)
        )

    if not periodo_geral_solicitado:
        itens_filtrados = base_queryset.filter(
            **{f"{CAMPO_DATA_FILTRO}__gte": data_inicio, f"{CAMPO_DATA_FILTRO}__lte": data_fim}
        ).exclude(**{f"{CAMPO_DATA_FILTRO}__isnull": True})
    else:
        itens_filtrados = base_queryset

    # Anotação de valores do site
    valor_unit_pref = Coalesce(F('valor_unitario_sp'), F('valor_unitario_es'), Value(0, output_field=DecimalField()))
    itens_filtrados = itens_filtrados.annotate(
        valor_total_item=ExpressionWrapper(F('quantidade') * valor_unit_pref, output_field=DecimalField())
    )

    # --- 5. Rankings e Agregações ---
    # Totais do Site
    totais_site = itens_filtrados.aggregate(
        total_periodo=Sum('valor_total_item'),
        total_sp=Sum('valor_total_item', filter=Q(pedido__cliente__client_state__uf_name='SP')),
        total_es=Sum('valor_total_item', filter=Q(pedido__cliente__client_state__uf_name='ES'))
    )

    # NOVO: Produtos Mais Vendidos (Recuperado)
    produtos_top = itens_filtrados.values(
        'produto__product_code', 'produto__product_description'
    ).annotate(total_vendido=Sum('quantidade')).order_by('-total_vendido')[:10]

    def get_top_clients(uf):
        return list(itens_filtrados.filter(pedido__cliente__client_state__uf_name=uf)
            .values('pedido__cliente__client_id', 'pedido__cliente__client_name', 'pedido__cliente__client_code')
            .annotate(total_gasto=Sum('valor_total_item'), num_pedidos=Count('pedido__id', distinct=True))
            .order_by('-total_gasto')[:5])

    # --- 6. Vendas por Mês (Recuperado) ---
    vendas_por_mes = itens_filtrados.annotate(
        mes_ano=TruncMonth(CAMPO_DATA_FILTRO)
    ).values('mes_ano').annotate(total_vendas=Sum('valor_total_item')).order_by('mes_ano')

    contexto = {
        'titulo': 'Dashboard de Análise de Dados',
        'data_inicio': data_inicio.strftime('%Y-%m-%d'),
        'data_fim': data_fim.strftime('%Y-%m-%d'),
        'total_faturamento_erp': total_faturamento_erp,  # VALOR REAL ERP (914k corrigido)
        'total_vendas_periodo_calculado': totais_site['total_periodo'] or 0, # VALOR SITE
        'total_vendas_sp_clientes': totais_site['total_sp'] or 0,
        'total_vendas_es_clientes': totais_site['total_es'] or 0,
        'clientes_top_sp': get_top_clients('SP'),
        'clientes_top_es': get_top_clients('ES'),
        'produtos_top': produtos_top,  # Ranking de produtos recuperado
        'vendas_por_mes': vendas_por_mes,
        'sugestoes_erp': sugestoes_erp,
        'lista_clientes_filtro': lista_clientes_filtro,
        'cliente_selecionado_id': cliente_id_analise,
        'periodo_geral_ativo': periodo_geral_solicitado,
    }
    return render(request, 'analise/analise_dashboard.html', contexto)



@staff_member_required
def pedido_por_texto_admin(request):
    """
    Permite que o admin cole um pedido em texto livre para qualquer cliente.
    O cliente é selecionado via GET (?cliente=ID), igual ao upload_pedido.
    """
    cliente_selecionado = None
    form_cliente = SelectClientForm(request.GET or None)
    form_cliente.fields['cliente'].queryset = por_empresa(WfClient.objects.all(), request).order_by('client_code')

    cliente_id_get = request.GET.get('cliente')
    if form_cliente.is_valid() and form_cliente.cleaned_data.get('cliente'):
        cliente_selecionado = form_cliente.cleaned_data['cliente']
    elif cliente_id_get:
        cliente_selecionado = get_empresa_or_404(WfClient, request, pk=cliente_id_get)

    if request.method == 'POST' and cliente_selecionado:
        texto = request.POST.get('texto_pedido', '').strip()
        endereco_id = request.POST.get('endereco_selecionado')
        data_envio_str = request.POST.get('data_expedicao')
        frete_option = request.POST.get('frete_option')
        nota_fiscal = request.POST.get('nota_fiscal')
        observacao = request.POST.get('observacao', '')

        _frete_validos = [c[0] for c in Pedido.FRETE_CHOICES]
        _nota_validos = [c[0] for c in Pedido.NOTA_FISCAL_CHOICES]
        if frete_option not in _frete_validos or nota_fiscal not in _nota_validos:
            messages.error(request, 'Opção de frete ou nota fiscal inválida.')
            return redirect(request.get_full_path())

        if not texto:
            messages.error(request, 'Cole o texto do pedido antes de enviar.')
            return redirect(request.get_full_path())

        fretes_sem_endereco = ['ONIBUS', 'RETIRADA']
        endereco_selecionado = None
        if frete_option not in fretes_sem_endereco:
            if not endereco_id:
                messages.error(request, 'Selecione um endereço de entrega.')
                return redirect(request.get_full_path())
            endereco_selecionado = get_object_or_404(Endereco, id=endereco_id, cliente=cliente_selecionado)

        try:
            data_envio_obj = datetime.strptime(data_envio_str, '%Y-%m-%d').date() if data_envio_str else None
        except ValueError:
            messages.error(request, 'Data de expedição inválida.')
            return redirect(request.get_full_path())

        itens_parseados = _parse_texto_pedido(texto)
        if not itens_parseados:
            messages.error(request, 'Nenhum item reconhecido. Formato: CÓDIGO —- QUANTIDADE.')
            return redirect(request.get_full_path())

        if not cliente_selecionado.client_state:
            messages.error(request, f'Cliente {cliente_selecionado.client_name} sem estado (SP/ES) cadastrado.')
            return redirect(request.get_full_path())

        regiao = cliente_selecionado.client_state.uf_name
        valor_field = 'product_value_sp' if regiao == 'SP' else 'product_value_es'

        from django.db.models import Max as _Max
        _base = por_empresa(Product.objects, request)
        _latest = _base.aggregate(d=_Max('date_product'))['d']
        produtos_atuais = {p.product_code: p for p in _base.filter(date_product=_latest)}

        erros = []
        itens_pedido = []
        itens_ignorados = []
        total = Decimal('0.00')

        for item in itens_parseados:
            codigo = item['codigo']
            quantidade = item['quantidade']
            produto = produtos_atuais.get(codigo)
            if not produto:
                erros.append(f"'{codigo}': não encontrado no catálogo.")
                itens_ignorados.append({'codigo': codigo, 'motivo': 'Não encontrado no catálogo'})
                continue
            valor_unitario = getattr(produto, valor_field)
            if not valor_unitario or valor_unitario <= 0:
                erros.append(f"'{codigo}': indisponível no estoque.")
                itens_ignorados.append({'codigo': codigo, 'motivo': 'Indisponível no estoque'})
                continue
            total += valor_unitario * Decimal(quantidade)
            itens_pedido.append(ItemPedido(
                produto=produto,
                quantidade=quantidade,
                valor_unitario_sp=produto.product_value_sp,
                valor_unitario_es=produto.product_value_es,
            ))

        if not itens_pedido:
            messages.error(request, 'Nenhum item válido. Erros: ' + ' | '.join(erros))
            return redirect(request.get_full_path())

        with transaction.atomic():
            novo_pedido = Pedido.objects.create(
                cliente=cliente_selecionado,
                empresa=cliente_selecionado.empresa,
                endereco=endereco_selecionado,
                data_envio_solicitada=data_envio_obj,
                frete_option=frete_option,
                nota_fiscal=nota_fiscal,
                observacao=observacao,
                status='RASCUNHO',
                criado_por=request.user,
                valor_total=total,
            )
            for item in itens_pedido:
                item.pedido = novo_pedido
            ItemPedido.objects.bulk_create(itens_pedido)
            if itens_ignorados:
                ItemPedidoIgnorado.objects.bulk_create([
                    ItemPedidoIgnorado(
                        pedido=novo_pedido, cliente=cliente_selecionado,
                        codigo_produto=i['codigo'], descricao_produto='',
                        quantidade_tentada=0, motivo_erro=i['motivo'],
                    ) for i in itens_ignorados
                ])

        if erros:
            messages.warning(request, f'{len(itens_pedido)} iten(s) adicionados. Ignorados: ' + ' | '.join(erros))
        else:
            messages.success(request, f'Pedido interpretado! {len(itens_pedido)} iten(s) adicionados.')

        return redirect('checkout_rascunho', pedido_id_rascunho=novo_pedido.id)

    enderecos = Endereco.objects.filter(cliente=cliente_selecionado) if cliente_selecionado else []
    contexto = {
        'titulo': 'Pedido por Texto',
        'form_cliente': form_cliente,
        'cliente_selecionado': cliente_selecionado,
        'enderecos': enderecos,
        'frete_choices': Pedido.FRETE_CHOICES,
        'nota_choices': Pedido.NOTA_FISCAL_CHOICES,
    }
    return render(request, 'pedido_por_texto_admin.html', contexto)


@login_required
def upload_pedido_cliente(request):
    try:
        # Puxa o cliente vinculado ao usuário logado
        cliente = request.user.wfclient
    except WfClient.DoesNotExist:
        messages.error(request, 'Seu usuário não possui um perfil de cliente vinculado.')
        return redirect('home')
    
    codigo_cliente_analise = request.GET.get('cliente_codigo') # Exemplo

    sugestoes_erp = []
    if codigo_cliente_analise:
        sugestoes_erp = VendaReal.objects.filter(
            Codigo_Cliente=codigo_cliente_analise
        ).values(
            'Produto_Codigo', 'Produto_Descricao'
        ).annotate(
            total_qtd=Sum('Quantidade'),
            vezes_faturado=Count('id')
        ).order_by('-total_qtd')[:10]

    # Dados iniciais baseados nas preferências do cliente
    initial_data = {
        'frete_option': cliente.frete_preferencia,
        'nota_fiscal': cliente.nota_fiscal_preferencia,
        'observacao_preferencia': cliente.observacao_preferencia,
    }

    if request.method == 'POST':
        form = UploadPedidoForm(request.POST, request.FILES)
        # Filtra os endereços apenas para o cliente logado
        form.fields['endereco_selecionado'].queryset = Endereco.objects.filter(cliente=cliente)

        if form.is_valid():
            try:
                planilha_pedido = request.FILES.get('planilha_pedido')
                if not planilha_pedido:
                    messages.error(request, 'Nenhum arquivo foi selecionado.')
                    return redirect('upload_pedido_cliente')
                
                # 1. Leitura do arquivo (Excel ou CSV)
                if planilha_pedido.name.endswith('.csv'):
                    df_list = [pd.read_csv(planilha_pedido)]
                else:
                    # Detecta linha de cabeçalho dinamicamente (planilhas com linha de data no topo)
                    header_row = 0
                    df_scan = pd.read_excel(planilha_pedido, header=None, nrows=15, dtype=str)
                    for _i, _row in df_scan.iterrows():
                        _vals = [str(v).upper().strip() for v in _row if pd.notnull(v) and str(v).strip()]
                        if any('COD' in v for v in _vals) and any('QTD' in v or 'QUANT' in v for v in _vals):
                            header_row = _i
                            break
                    planilha_pedido.seek(0)
                    xls_data = pd.read_excel(planilha_pedido, sheet_name=None, header=header_row)
                    df_list = list(xls_data.values())

                df = pd.concat(df_list, ignore_index=True).dropna(how='all')

                if df.empty:
                    messages.error(request, 'A planilha está vazia.')
                    return redirect('upload_pedido_cliente')

                # 2. Normalização e Mapeamento de Colunas
                df.columns = [normalize_text(col) for col in df.columns]

                expected_cols = {
                    'codigo': ['codigo', 'código', 'cod'],
                    'quantidade': ['quantidade', 'qtd', 'qtde'],
                    'descricao': ['descricao', 'descrição', 'produto', 'nome', 'description']
                }

                col_mapping = {
                    key: next((c for c in values if c in df.columns), None)
                    for key, values in expected_cols.items()
                }

                if not col_mapping['codigo'] or not col_mapping['quantidade']:
                    messages.error(request, "A planilha deve ter colunas para 'código' e 'quantidade'.")
                    return render(request, 'upload_pedido_cliente.html', {'upload_form': form, 'cliente': cliente})

                # Filtra apenas linhas com quantidade preenchida e positiva antes do loop
                col_qtd_raw = col_mapping['quantidade']
                df = df[pd.to_numeric(df[col_qtd_raw], errors='coerce').fillna(0) > 0]
                # Deduplica por código — mantém a linha com maior quantidade se aparecer em múltiplas abas
                col_cod_raw = col_mapping['codigo']
                df = df.dropna(subset=[col_cod_raw])
                df[col_cod_raw] = df[col_cod_raw].astype(str).str.strip()
                df = df.sort_values(col_qtd_raw, ascending=False).drop_duplicates(subset=[col_cod_raw])

                with transaction.atomic():
                    # 3. Criação do Pedido Rascunho
                    novo_pedido = Pedido.objects.create(
                        cliente=cliente,
                        empresa=cliente.empresa,
                        endereco=form.cleaned_data.get('endereco_selecionado'),
                        data_criacao=timezone.now(),
                        data_envio_solicitada=form.cleaned_data['data_expedicao'],
                        frete_option=form.cleaned_data['frete_option'],
                        nota_fiscal=form.cleaned_data['nota_fiscal'],
                        status='RASCUNHO',
                        criado_por=request.user,
                        observacao=form.cleaned_data['observacao_preferencia'],
                    )
                    
                    erros_texto = []
                    itens_pedido_para_criar = []
                    itens_ignorados_db = []
                    total_valor_pedido = Decimal('0.0')

                    # Busca produtos em lote (1 query)
                    from django.db.models import Max as _Max
                    _base = por_empresa(Product.objects, request)
                    _latest = _base.aggregate(d=_Max('date_product'))['d']
                    produtos_atuais = {p.product_code: p for p in _base.filter(date_product=_latest)}

                    # Constantes fora do loop
                    col_cod  = col_mapping['codigo']
                    col_qtd  = col_mapping['quantidade']
                    col_desc = col_mapping.get('descricao')
                    termos_ignorar = {'TOTAL', 'SUBTOTAL', 'GERAL', 'VALOR TOTAL'}
                    regiao      = cliente.client_state.uf_name
                    valor_field = 'product_value_sp' if regiao == 'SP' else 'product_value_es'

                    # 4. Processamento linha a linha (to_dict é ~10x mais rápido que iterrows)
                    for index, row in enumerate(df.to_dict('records'), start=2):
                        codigo_raw = row.get(col_cod)
                        if codigo_raw is None or (not isinstance(codigo_raw, str) and pd.isna(codigo_raw)):
                            continue

                        codigo_produto = str(codigo_raw).strip()

                        if any(termo in codigo_produto.upper() for termo in termos_ignorar):
                            continue

                        quantidade_raw = row.get(col_qtd)
                        descricao_excel = row.get(col_desc, 'Descrição não informada') if col_desc else 'Descrição não informada'

                        if quantidade_raw is None or (not isinstance(quantidade_raw, str) and pd.isnull(quantidade_raw)):
                            continue

                        # Validação: Quantidade Numérica
                        try:
                            quantidade = int(quantidade_raw)
                        except (ValueError, TypeError):
                            msg = "Quantidade inválida (não-numérica)"
                            erros_texto.append(f"Produto '{codigo_produto}' na linha {index}: {msg}.")
                            itens_ignorados_db.append(ItemPedidoIgnorado(
                                pedido=novo_pedido, cliente=cliente, codigo_produto=codigo_produto,
                                descricao_produto=str(descricao_excel), quantidade_tentada=0, motivo_erro=msg
                            ))
                            continue

                        # Validação: Quantidade Positiva
                        if quantidade <= 0:
                            continue

                        # Validação: Catálogo
                        produto = produtos_atuais.get(codigo_produto)
                        if not produto:
                            msg = "Não encontrado no catálogo"
                            erros_texto.append(f"Produto '{codigo_produto}' na linha {index}: {msg}.")
                            itens_ignorados_db.append(ItemPedidoIgnorado(
                                pedido=novo_pedido, cliente=cliente, codigo_produto=codigo_produto,
                                descricao_produto=str(descricao_excel), quantidade_tentada=quantidade, motivo_erro=msg
                            ))
                            continue

                        # Validação: Preço por Região (SP ou ES)
                        valor_unitario = getattr(produto, valor_field)
                        
                        if valor_unitario is None or valor_unitario <= 0:
                            msg = f"Indisponível no estoque"
                            erros_texto.append(f"Produto '{codigo_produto}' na linha {index + 2}: {msg}.")
                            itens_ignorados_db.append(ItemPedidoIgnorado(
                                pedido=novo_pedido, cliente=cliente, codigo_produto=codigo_produto,
                                descricao_produto=produto.product_description, quantidade_tentada=quantidade, motivo_erro=msg
                            ))
                            continue

                        # Sucesso: Prepara para criar
                        total_valor_pedido += valor_unitario * Decimal(quantidade)
                        itens_pedido_para_criar.append(ItemPedido(
                            pedido=novo_pedido,
                            produto=produto,
                            quantidade=quantidade,
                            valor_unitario_sp=produto.product_value_sp,
                            valor_unitario_es=produto.product_value_es,
                        ))
                    
                    # 5. Finalização
                    if not itens_pedido_para_criar and not itens_ignorados_db:
                        messages.error(request, "Nenhum dado processável encontrado.")
                        novo_pedido.delete()
                        return redirect('upload_pedido_cliente')
                    
                    if itens_pedido_para_criar:
                        ItemPedido.objects.bulk_create(itens_pedido_para_criar)
                    
                    if itens_ignorados_db:
                        ItemPedidoIgnorado.objects.bulk_create(itens_ignorados_db)
                    
                    if erros_texto:
                        novo_pedido.erros_upload = '\n'.join(erros_texto)
                        erros_msg = 'Alguns itens foram ignorados:\n' + '\n'.join(erros_texto[:3])
                        if len(erros_texto) > 3: erros_msg += f'\n... e mais {len(erros_texto) - 3} erros.'
                        messages.warning(request, erros_msg)
                    else:
                        messages.success(request, "Itens carregados com sucesso!")
                    
                    novo_pedido.valor_total = total_valor_pedido
                    novo_pedido.save()
                    _notificar_novo_pedido(novo_pedido)

                    return redirect('checkout_rascunho', pedido_id_rascunho=novo_pedido.id)

            except Exception as e:
                messages.error(request, f"Erro crítico no processamento: {e}")
                if 'novo_pedido' in locals(): novo_pedido.delete()
        
    else:
        # GET: Prepara formulário com endereços do cliente e sugestão de endereço padrão
        form = UploadPedidoForm(initial=initial_data)
        form.fields['endereco_selecionado'].queryset = Endereco.objects.filter(cliente=cliente)
        end_padrao = Endereco.objects.filter(cliente=cliente, is_default=True).first()
        if end_padrao:
            form.initial['endereco_selecionado'] = end_padrao.id

    return render(request, 'upload_pedido_cliente.html', {'upload_form': form, 'cliente': cliente})



def exportar_detalhes_pedido_whatsapp_excel(request, pedido_id):
    """
    Exporta uma planilha simplificada (Código, Quantidade, Preço)
    para facilitar a digitação via WhatsApp.
    """
    pedido = get_empresa_or_404(Pedido, request, id=pedido_id)
    itens_pedido = ItemPedido.objects.filter(pedido=pedido).select_related('produto')
    uf_cliente = pedido.cliente.client_state.uf_name

    # Define qual valor unitário usar com base no estado
    valor_key = 'valor_unitario_sp' if uf_cliente == 'SP' else 'valor_unitario_es'

    data = []
    for item in itens_pedido:
        valor_unitario = getattr(item, valor_key) or 0
        data.append({
            'Código': item.produto.product_code,
            'Quantidade': item.quantidade,
            'Preço Unitário': float(valor_unitario),
        })

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Resumo para Digitação')
        
        # Ajuste de largura básico
        worksheet = writer.sheets['Resumo para Digitação']
        worksheet.set_column('A:A', 15) # Código
        worksheet.set_column('B:C', 12) # Qtd e Preço

    output.seek(0)
    data_hoje = date.today().strftime('%d-%m-%Y')
    filename = f"pedido_{pedido.cliente.client_code}_{data_hoje}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

def _parse_texto_pedido(texto):
    """
    Interpreta um texto livre de pedido enviado via WhatsApp/mensagem.
    Suporta formatos:
        K6020 —-10
        K1071 ——3
        C2011    PLACA CARGA SAMSUNG A03    -7
    Retorna lista de dicts: [{'codigo': 'K6020', 'quantidade': 10}, ...]
    """
    itens = []
    vistos = {}  # deduplica somando quantidades

    for linha in texto.splitlines():
        linha = linha.strip()
        if not linha:
            continue
        # Código: 1-2 letras maiúsculas + dígitos no início da linha
        # Quantidade: último número após traços (-, —, combinações)
        match = re.match(r'^([A-Z]{1,2}\d+).*[-—]+\s*(\d+)\s*$', linha, re.IGNORECASE)
        if match:
            codigo = match.group(1).upper()
            quantidade = int(match.group(2))
            if quantidade > 0:
                vistos[codigo] = vistos.get(codigo, 0) + quantidade

    for codigo, quantidade in vistos.items():
        itens.append({'codigo': codigo, 'quantidade': quantidade})

    return itens


@login_required
def pedido_por_texto(request):
    """
    Permite que clientes colem um pedido em texto livre (formato WhatsApp)
    e o sistema interpreta, valida e cria o rascunho automaticamente.
    """
    try:
        cliente = request.user.wfclient
    except WfClient.DoesNotExist:
        messages.error(request, 'Usuário sem cliente associado.')
        return redirect('home')

    enderecos = Endereco.objects.filter(cliente=cliente)

    if request.method == 'POST':
        texto = request.POST.get('texto_pedido', '').strip()
        endereco_id = request.POST.get('endereco_selecionado')
        data_envio_str = request.POST.get('data_expedicao')
        frete_option = request.POST.get('frete_option')
        nota_fiscal = request.POST.get('nota_fiscal')
        observacao = request.POST.get('observacao', '')

        # Validações básicas
        _frete_validos = [c[0] for c in Pedido.FRETE_CHOICES]
        _nota_validos = [c[0] for c in Pedido.NOTA_FISCAL_CHOICES]
        if frete_option not in _frete_validos or nota_fiscal not in _nota_validos:
            messages.error(request, 'Opção de frete ou nota fiscal inválida.')
            return redirect('pedido_por_texto')

        if not texto:
            messages.error(request, 'Cole o texto do pedido antes de enviar.')
            return redirect('pedido_por_texto')

        fretes_sem_endereco = ['ONIBUS', 'RETIRADA']
        endereco_selecionado = None
        if frete_option not in fretes_sem_endereco:
            if not endereco_id:
                messages.error(request, 'Selecione um endereço de entrega.')
                return redirect('pedido_por_texto')
            endereco_selecionado = get_object_or_404(Endereco, id=endereco_id, cliente=cliente)

        try:
            data_envio_obj = datetime.strptime(data_envio_str, '%Y-%m-%d').date() if data_envio_str else None
        except ValueError:
            messages.error(request, 'Data de expedição inválida.')
            return redirect('pedido_por_texto')

        itens_parseados = _parse_texto_pedido(texto)
        if not itens_parseados:
            messages.error(request, 'Nenhum item reconhecido no texto. Verifique o formato: CÓDIGO —- QUANTIDADE.')
            return redirect('pedido_por_texto')

        # Busca produtos em lote
        if not cliente.client_state:
            messages.error(request, 'Seu cadastro não possui estado (SP/ES) definido. Contate o suporte.')
            return redirect('pedido_por_texto')
        regiao = cliente.client_state.uf_name
        valor_field = 'product_value_sp' if regiao == 'SP' else 'product_value_es'

        from django.db.models import Max as _Max
        _base = por_empresa(Product.objects, request)
        _latest = _base.aggregate(d=_Max('date_product'))['d']
        produtos_atuais = {p.product_code: p for p in _base.filter(date_product=_latest)}

        erros = []
        itens_pedido = []
        itens_ignorados = []
        total = Decimal('0.00')

        for item in itens_parseados:
            codigo = item['codigo']
            quantidade = item['quantidade']
            produto = produtos_atuais.get(codigo)

            if not produto:
                erros.append(f"'{codigo}': não encontrado no catálogo.")
                itens_ignorados.append({'codigo': codigo, 'motivo': 'Não encontrado no catálogo'})
                continue

            valor_unitario = getattr(produto, valor_field)
            if not valor_unitario or valor_unitario <= 0:
                erros.append(f"'{codigo}': indisponível no estoque.")
                itens_ignorados.append({'codigo': codigo, 'motivo': 'Indisponível no estoque'})
                continue

            total += valor_unitario * Decimal(quantidade)
            itens_pedido.append(ItemPedido(
                produto=produto,
                quantidade=quantidade,
                valor_unitario_sp=produto.product_value_sp,
                valor_unitario_es=produto.product_value_es,
            ))

        if not itens_pedido:
            messages.error(request, 'Nenhum item válido encontrado. Erros: ' + ' | '.join(erros))
            return redirect('pedido_por_texto')

        with transaction.atomic():
            novo_pedido = Pedido.objects.create(
                cliente=cliente,
                empresa=cliente.empresa,
                endereco=endereco_selecionado,
                data_envio_solicitada=data_envio_obj,
                frete_option=frete_option,
                nota_fiscal=nota_fiscal,
                observacao=observacao,
                status='RASCUNHO',
                criado_por=request.user,
                valor_total=total,
            )
            for item in itens_pedido:
                item.pedido = novo_pedido
            ItemPedido.objects.bulk_create(itens_pedido)

            if itens_ignorados:
                ItemPedidoIgnorado.objects.bulk_create([
                    ItemPedidoIgnorado(
                        pedido=novo_pedido,
                        cliente=cliente,
                        codigo_produto=i['codigo'],
                        descricao_produto='',
                        quantidade_tentada=0,
                        motivo_erro=i['motivo'],
                    ) for i in itens_ignorados
                ])

        if erros:
            messages.warning(request, f'Pedido criado com {len(itens_pedido)} iten(s). Ignorados: ' + ' | '.join(erros))
        else:
            messages.success(request, f'Pedido interpretado com sucesso! {len(itens_pedido)} iten(s) adicionados.')

        return redirect('checkout_rascunho', pedido_id_rascunho=novo_pedido.id)

    contexto = {
        'titulo': 'Pedido por Texto',
        'enderecos': enderecos,
        'cliente': cliente,
        'frete_choices': Pedido.FRETE_CHOICES,
        'nota_choices': Pedido.NOTA_FISCAL_CHOICES,
        'frete_preferencia': cliente.frete_preferencia,
        'nota_preferencia': cliente.nota_fiscal_preferencia,
    }
    return render(request, 'pedido_por_texto.html', contexto)


@login_required
def detalhes_produto(request, product_id):
    product = get_object_or_404(Product, product_id=product_id)
    
    # Busca recomendações
    recomendacoes_raw = product.get_recommendations(limit=4)
    
    # Transforma em objetos Product para facilitar o uso no template (preços, etc)
    recomendacoes_ids = [item['produto_id'] for item in recomendacoes_raw]
    produtos_recomendados = Product.objects.filter(product_id__in=recomendacoes_ids)

    return render(request, 'detalhes_produto.html', {
        'product': product,
        'recomendacoes': produtos_recomendados
    })


@login_required
def sugestoes_compra(request):
    try:
        cliente_logado = request.user.wfclient
        # Busca os itens frequentes usando o método que criamos no Model
        itens_frequentes = cliente_logado.get_frequent_items(limit=24) # Aumentei o limite para uma página cheia
    except WfClient.DoesNotExist:
        return redirect('home')

    contexto = {
        'titulo': 'Sugestões de Reabastecimento',
        'itens_frequentes': itens_frequentes,
        'cliente_logado': cliente_logado,
    }
    return render(request, 'sugestoes.html', contexto)

# upload planilha do trs
from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect, render
import pandas as pd
from decimal import Decimal

def _extrair_df_do_pdf_vendas(file_obj):
    """
    Extrai um DataFrame de vendas a partir de um PDF do ERP.
    Retorna DataFrame com colunas: Código_Cliente, Emissão, Pedido,
    Produto_Código, Produto_Descrição, Quantidade, Unitário, Total.
    """
    import fitz

    conteudo_bytes = file_obj.read()
    doc = fitz.open(stream=conteudo_bytes, filetype="pdf")
    texto_completo = "".join(page.get_text() for page in doc)
    doc.close()

    if not texto_completo.strip():
        return pd.DataFrame()

    dados_combinados = []

    padrao_cliente = re.compile(r"(Cliente:\s*(\d+)\s+-\s*.*?)", re.IGNORECASE | re.DOTALL)
    padrao_linhas = re.compile(
        r"(R\$\s*[\d\.\,]+(?:[0-9]*))\s*(R\$\s*[\d\.\,]+)\s*"
        r"(.+?)\s*([A-Z0-9]+)\s*"
        r"([\d\.\,]+)\s*(\d{2}/\d{2}/\d{4})\s*(\d+)",
        re.IGNORECASE | re.DOTALL
    )
    padrao_bloco_tabela = re.compile(
        r"The following table:[\s\n]*\"Emissão\s*\n\s*\",\"Pedido\s*\n\s*\",\"Produto\s*\n\s*\",,?"
        r"(?:\"Quantidade\s*\n\s*\",\"Unitário\s*\n\s*\",\"Total\s*\n\s*\")?[\s\n]*"
        r"((?:\"[^\"]*\"(?:,\s*\"[^\"]*\")*,\s*\"[^\"]*\"\s*\n)+)",
        re.IGNORECASE
    )

    secoes = re.split(padrao_cliente, texto_completo)
    secoes = [s.strip() for s in secoes if s.strip()]
    codigo_cliente_atual = "N/A"

    for secao in secoes:
        match_cliente = padrao_cliente.search(secao)
        if match_cliente:
            codigo_cliente_atual = match_cliente.group(2)
            continue

        for linha in padrao_linhas.findall(secao):
            dados_combinados.append([
                codigo_cliente_atual,
                linha[5],                              # Data
                linha[6],                              # Pedido
                linha[3],                              # Código Produto
                linha[2].strip().replace('\n', ' '),   # Descrição
                linha[4],                              # Quantidade
                linha[0],                              # Unitário
                linha[1],                              # Total
            ])

        for bloco in padrao_bloco_tabela.findall(secao):
            for linha in bloco.strip().split('\n'):
                partes = [p.strip().replace('"', '').replace('\n', '') for p in linha.split(',')]
                partes = [p for p in partes if p and not p.isspace()]
                if len(partes) >= 7:
                    dados_combinados.append([codigo_cliente_atual] + partes[:7])
                elif len(partes) >= 6:
                    dados_combinados.append([codigo_cliente_atual, partes[0], partes[1], partes[2], "", partes[3], partes[4], partes[5]])

    colunas = ["Código_Cliente", "Emissão", "Pedido", "Produto_Código", "Produto_Descrição", "Quantidade", "Unitário", "Total"]
    if not dados_combinados:
        return pd.DataFrame(columns=colunas)

    df = pd.DataFrame(dados_combinados, columns=colunas)

    for col in ["Quantidade", "Unitário", "Total"]:
        df[col] = (df[col].astype(str)
                   .str.replace(r'R\$\s*', '', regex=True)
                   .str.replace(r'\.', '', regex=True)
                   .str.replace(r',', '.', regex=False)
                   .str.strip())
        df[col] = pd.to_numeric(df[col], errors='coerce')

    for col in ["Emissão", "Pedido", "Produto_Código", "Produto_Descrição"]:
        df[col] = df[col].astype(str).str.strip().str.replace('\n', ' ', regex=False)

    df = df.dropna(subset=['Total']).drop_duplicates().reset_index(drop=True)
    return df


@staff_member_required
def upload_vendas_reais(request):
    if request.method == 'POST' and request.FILES.get('planilha_vendas'):
        file = request.FILES['planilha_vendas']
        try:
            nome = file.name.lower()
            if nome.endswith('.pdf'):
                df = _extrair_df_do_pdf_vendas(file)
                if df.empty:
                    messages.error(request, 'Não foi possível extrair dados do PDF. Verifique se o arquivo é um relatório de vendas válido.')
                    return redirect('dashboard_admin')
                # Renomeia colunas para bater com o formato esperado
                df = df.rename(columns={
                    'Código_Cliente': 'Código_Cliente',
                    'Emissão': 'Emissão',
                    'Pedido': 'Pedido',
                    'Produto_Código': 'Produto_Código',
                    'Produto_Descrição': 'Produto_Descrição',
                    'Quantidade': 'Quantidade',
                    'Unitário': 'Unitário',
                    'Total': 'Total',
                })
            elif nome.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file)
            else:
                messages.error(request, 'Formato inválido. Envie um arquivo PDF ou Excel (.xlsx).')
                return redirect('dashboard_admin')
            df = df.dropna(how='all')

            # 1. OTIMIZAÇÃO: Conversão de data
            df['Emissão_dt'] = pd.to_datetime(df['Emissão'], dayfirst=True)
            dias_na_planilha = df['Emissão_dt'].dt.date.unique()

            with transaction.atomic():
                VendaReal.objects.filter(Emissao__in=dias_na_planilha).delete()

            # 2. LIMPEZA E SANITIZAÇÃO FINANCEIRA
            df['Código_Cliente'] = df['Código_Cliente'].fillna(0)
            
            # Garante que o Pandas consiga somar valores caso o Excel venha com vírgulas e formato texto
            df['Total'] = pd.to_numeric(df['Total'].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)
            df['Unitário'] = pd.to_numeric(df['Unitário'].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)
            df['Quantidade'] = pd.to_numeric(df['Quantidade'], errors='coerce').fillna(0)

            # 3. OTIMIZAÇÃO DE AGRUPAMENTO (A correção do gargalo)
            # Sem a 'Produto_Descrição', todos os códigos genéricos (ex: 'FL') do mesmo pedido 
            # se fundem em uma linha única, somando os valores e blindando contra perda de dados.
            df_grouped = df.groupby(
                ['Emissão_dt', 'Código_Cliente', 'Pedido', 'Produto_Código'], 
                dropna=False, as_index=False
            ).agg({
                'Produto_Descrição': 'first', # Guarda apenas a primeira descrição encontrada
                'Quantidade': 'sum',
                'Unitário': 'first',
                'Total': 'sum'
            })

            # 4. PROCESSAMENTO PARA O BANCO — versão vetorizada (sem iterrows)
            # Cria colunas auxiliares com nomes ASCII para compatibilidade com itertuples
            cod_str = df_grouped['Código_Cliente'].astype(str).str.replace('.0', '', regex=False)

            codigos_na_planilha = cod_str.unique().tolist()
            clientes_dict = {
                str(c.client_code): c.client_name
                for c in WfClient.objects.filter(client_code__in=codigos_na_planilha).only('client_code', 'client_name')
            }

            df_grouped['col_emissao'] = df_grouped['Emissão_dt'].dt.date
            df_grouped['col_cod'] = cod_str
            df_grouped['col_nome'] = cod_str.apply(
                lambda cod: "Consumidor Final / Não Identificado" if cod == "0"
                else clientes_dict.get(cod, f"Cod: {cod}")
            )
            df_grouped['col_pedido'] = df_grouped['Pedido'].astype(str).str.replace('.0', '', regex=False)
            df_grouped['col_prod'] = df_grouped['Produto_Código'].astype(str).str.replace('.0', '', regex=False)
            df_grouped['col_descricao'] = df_grouped['Produto_Descrição'].astype(str)
            df_grouped['col_qtd'] = df_grouped['Quantidade']
            df_grouped['col_unit'] = df_grouped['Unitário'].astype(str)
            df_grouped['col_total'] = df_grouped['Total'].astype(str)

            empresa = request.empresa
            novas_vendas = [
                VendaReal(
                    Emissao=row.col_emissao,
                    Codigo_Cliente=int(float(row.col_cod)) if row.col_cod not in ('', 'nan') else 0,
                    Pedido=row.col_pedido,
                    Produto_Codigo=row.col_prod,
                    cliente_nome=row.col_nome,
                    Produto_Descricao=row.col_descricao,
                    Quantidade=int(row.col_qtd),
                    Unitario=Decimal(row.col_unit),
                    Total=Decimal(row.col_total),
                    empresa=empresa,
                )
                for row in df_grouped.itertuples(index=False)
            ]

            if novas_vendas:
                with transaction.atomic():
                    VendaReal.objects.bulk_create(novas_vendas, batch_size=1000, ignore_conflicts=True)
                messages.success(request, f'Sucesso! {len(novas_vendas)} itens processados. Histórico atualizado para {len(dias_na_planilha)} dias.')
            else:
                messages.warning(request, 'Nenhum dado válido encontrado na planilha.')
            
        except Exception as e:
            messages.error(request, f'Erro crítico ao processar: {e}')
        
        return redirect('dashboard_admin')

    return render(request, 'analise/upload_vendas_reais.html')

from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q
from .models import VendaReal
from django.contrib.admin.views.decorators import staff_member_required

from datetime import date # Certifique-se de que isso está nos seus imports no topo do arquivo

@staff_member_required
def listar_vendas_reais(request):
    # 1. Captura e limpa filtros
    filtro_pedido = request.GET.get('pedido', '').strip()
    filtro_produto = request.GET.get('produto', '').strip()
    filtro_cliente = request.GET.get('cliente', '').strip()
    
    # NOVOS FILTROS: Mês e Ano
    filtro_mes = request.GET.get('mes', '').strip()
    filtro_ano = request.GET.get('ano', '').strip()

    vendas_qs = por_empresa(VendaReal.objects.all(), request).order_by('-Emissao', '-Pedido')

    # 2. Aplica Filtros Dinâmicos
    if filtro_pedido:
        vendas_qs = vendas_qs.filter(Pedido__icontains=filtro_pedido)
    if filtro_produto:
        vendas_qs = vendas_qs.filter(
            Q(Produto_Codigo__icontains=filtro_produto) | 
            Q(Produto_Descricao__icontains=filtro_produto)
        )
    if filtro_cliente:
        vendas_qs = vendas_qs.filter(
            Q(cliente_nome__icontains=filtro_cliente) | 
            Q(Codigo_Cliente__icontains=filtro_cliente)
        )
        
    # APLICAÇÃO DOS NOVOS FILTROS (com validação para evitar erros na URL)
    if filtro_mes and filtro_mes.isdigit():
        vendas_qs = vendas_qs.filter(Emissao__month=int(filtro_mes))
    if filtro_ano and filtro_ano.isdigit():
        vendas_qs = vendas_qs.filter(Emissao__year=int(filtro_ano))

    # 3. Paginação Robusta (50 itens)
    paginator = Paginator(vendas_qs, 50)
    page_number = request.GET.get('page', 1)
    vendas_paginadas = paginator.get_page(page_number)

    # 4. Formatação Blindada para o Template
    for v in vendas_paginadas:
        v.unit_str = "{:,.2f}".format(float(v.Unitario)).replace(",", "X").replace(".", ",").replace("X", ".")
        v.total_str = "{:,.2f}".format(float(v.Total)).replace(",", "X").replace(".", ",").replace("X", ".")

    # Otimização para o Template: Enviamos a lista de anos para o Select HTML
    hoje = date.today()
    lista_anos = range(hoje.year - 2, hoje.year + 1)

    contexto = {
        'titulo': 'Histórico Detalhado ERP',
        'vendas': vendas_paginadas,
        # Enviamos os filtros de volta para o template para manter os inputs preenchidos
        'filtro_pedido': filtro_pedido,
        'filtro_produto': filtro_produto,
        'filtro_cliente': filtro_cliente,
        'filtro_mes': filtro_mes,
        'filtro_ano': filtro_ano,
        'lista_anos': lista_anos,
    }
    return render(request, 'analise/listar_vendas_reais.html', contexto)

@staff_member_required
def exportar_vendas_reais_excel(request):
    # 1. Captura os mesmos filtros da tela de listagem
    filtro_pedido = request.GET.get('pedido', '').strip()
    filtro_produto = request.GET.get('produto', '').strip()
    filtro_cliente = request.GET.get('cliente', '').strip()
    
    # NOVOS FILTROS: Mês e Ano
    filtro_mes = request.GET.get('mes', '').strip()
    filtro_ano = request.GET.get('ano', '').strip()

    vendas_qs = por_empresa(VendaReal.objects.all(), request).order_by('-Emissao')

    # 2. Aplica Filtros Dinâmicos
    if filtro_pedido:
        vendas_qs = vendas_qs.filter(Pedido__icontains=filtro_pedido)
    if filtro_produto:
        vendas_qs = vendas_qs.filter(
            Q(Produto_Codigo__icontains=filtro_produto) | 
            Q(Produto_Descricao__icontains=filtro_produto)
        )
    if filtro_cliente:
        vendas_qs = vendas_qs.filter(
            Q(cliente_nome__icontains=filtro_cliente) | 
            Q(Codigo_Cliente__icontains=filtro_cliente)
        )
        
    # APLICAÇÃO DOS NOVOS FILTROS
    if filtro_mes and filtro_mes.isdigit():
        vendas_qs = vendas_qs.filter(Emissao__month=int(filtro_mes))
    if filtro_ano and filtro_ano.isdigit():
        vendas_qs = vendas_qs.filter(Emissao__year=int(filtro_ano))

    # 3. Transforma o QuerySet em uma lista de dicionários para o Pandas
    data = []
    for v in vendas_qs:
        data.append({
            'Emissão': v.Emissao.strftime('%d/%m/%Y'),
            'Pedido': v.Pedido,
            'Cód. Cliente': v.Codigo_Cliente,
            'Cliente': v.cliente_nome,
            'Cód. Produto': v.Produto_Codigo,
            'Descrição': v.Produto_Descricao,
            'Quantidade': v.Quantidade,
            'Unitário': float(v.Unitario),
            'Total': float(v.Total),
        })

    # 4. Criação do Excel em memória
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Vendas Reais')
        
    output.seek(0)

    # 5. Resposta HTTP para download (Nome do arquivo dinâmico)
    # Se filtrou por ano e mês, coloca no nome do arquivo (ex: vendas_reais_2026_02.xlsx)
    periodo_str = f"_{filtro_ano}_{filtro_mes}" if filtro_ano and filtro_mes else ""
    filename = f"vendas_reais{periodo_str}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response




@login_required
def meus_itens_comprados(request):
    try:
        cliente = request.user.wfclient
    except WfClient.DoesNotExist:
        return redirect('home')

    # --- NOVO: CÁLCULO DO EFETIVO (MÊS VIGENTE) ---
    hoje = timezone.localdate()
    total_efetivo_mes = VendaReal.objects.filter(
        Codigo_Cliente=cliente.client_code,
        Emissao__year=hoje.year,
        Emissao__month=hoje.month
    ).aggregate(total=Sum('Total'))['total'] or Decimal('0.00')

    # Formatação do Efetivo para o padrão R$ 0.000,00
    efetivo_formatado = "{:,.2f}".format(float(total_efetivo_mes)).replace(",", "X").replace(".", ",").replace("X", ".")
    # ----------------------------------------------

    filtro_produto = request.GET.get('produto', '').strip()

    # Aplica filtro de produto ANTES do agrupamento para evitar query pesada
    base_qs = VendaReal.objects.filter(Codigo_Cliente=cliente.client_code)
    if filtro_produto:
        base_qs = base_qs.filter(
            Q(Produto_Codigo__icontains=filtro_produto) |
            Q(Produto_Descricao__icontains=filtro_produto)
        )

    # Lógica de agrupamento compatível com MySQL
    vendas_ids = base_qs.values('Produto_Codigo').annotate(
        ultima_venda=Max('id')
    ).values_list('ultima_venda', flat=True)

    vendas_qs = VendaReal.objects.filter(id__in=vendas_ids).order_by('-id')

    # Paginação
    paginator = Paginator(vendas_qs, 50)
    page_number = request.GET.get('page', 1)
    vendas_paginadas = paginator.get_page(page_number)

    # Formatação blindada de valores da tabela
    for v in vendas_paginadas:
        v.unit_str = "{:,.2f}".format(float(v.Unitario)).replace(",", "X").replace(".", ",").replace("X", ".")

    contexto = {
        'titulo': 'Meus Itens Comprados',
        'vendas': vendas_paginadas,
        'cliente_logado': cliente,
        'efetivo_mes': efetivo_formatado, # Novo dado no contexto
        'mes_nome': hoje.strftime('%B').capitalize() # Opcional: nome do mês
    }
    return render(request, 'meus_itens_comprados.html', contexto)

@login_required
def exportar_meus_itens_excel(request):
    try:
        cliente = request.user.wfclient
    except WfClient.DoesNotExist:
        return redirect('home')

    # 1. Filtra todas as vendas do cliente logado no ERP
    vendas_qs = VendaReal.objects.filter(
        Codigo_Cliente=cliente.client_code
    ).order_by('-Emissao', '-Pedido')

    # 2. Prepara os dados para o DataFrame (Pandas)
    data = []
    for v in vendas_qs:
        data.append({
            'Emissão': v.Emissao.strftime('%d/%m/%Y'),
            'Pedido ERP': v.Pedido,
            'Cód. Produto': v.Produto_Codigo,
            'Descrição': v.Produto_Descricao,
            'Quantidade': v.Quantidade,
            'Unitário': float(v.Unitario),
            'Total Item': float(v.Total),
        })

    # 3. Criação do Excel em memória (usando o padrão BytesIO que você já utiliza)
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Meu Histórico de Compras')
        
    output.seek(0)

    # 4. Resposta HTTP para download
    filename = f"historico_compras_{cliente.client_code}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

# wefixhub/views.py

@staff_member_required
def upload_status_pdf(request):
    if request.method == 'POST' and request.FILES.get('pdf_file'):
        pdf_file = request.FILES['pdf_file']
        
        try:
            # Chama o serviço que faz todo o processamento de texto e banco
            qtd_processados = processar_status_pdf(pdf_file, empresa=request.empresa)
            
            messages.success(request, f"Sucesso! {qtd_processados} pedidos processados.")
            return redirect('dashboard_admin')

        except Exception as e:
            messages.error(request, f"Erro crítico ao ler o PDF: {str(e)}")
            return redirect('dashboard_admin')

    return render(request, 'analise/upload_status_pdf.html')


@staff_member_required
def listar_status_erp(request):
    # 1. Busca todos os registros
    status_qs = por_empresa(StatusPedidoERP.objects.all(), request).order_by('-emissao', '-id')
    
    # 2. Filtro de busca por número do pedido
    numero_pedido = request.GET.get('numero_pedido')
    if numero_pedido:
        status_qs = status_qs.filter(numero_pedido__icontains=numero_pedido)

    # 3. CÁLCULO PARA OS CARDS DO TOPO
    # Conta quantos registros únicos estão marcados como expedidos
    total_expedidos = status_qs.filter(expedido=True).count()

    # 4. Paginação (50 por página)
    paginator = Paginator(status_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    contexto = {
        'titulo': 'Monitoramento de Status ERP',
        'page_obj': page_obj,
        'numero_pedido': numero_pedido,
        'total_expedidos': total_expedidos, # Agora o template consegue ler este valor
    }
    return render(request, 'analise/listar_status_erp.html', contexto)

@staff_member_required
def pedidos_nao_expedidos(request):
    qs = por_empresa(StatusPedidoERP.objects.filter(expedido=False), request).order_by('cod_cliente', 'numero_pedido', '-emissao')

    # Agrupa por cliente → pedido
    clientes = {}
    for item in qs:
        cod = item.cod_cliente
        if cod not in clientes:
            clientes[cod] = {'nome': item.nome_cliente, 'cod': cod, 'pedidos': {}}
        num = item.numero_pedido
        if num not in clientes[cod]['pedidos']:
            clientes[cod]['pedidos'][num] = {
                'numero': num,
                'emissao': item.emissao,
                'situacao': item.situacao,
            }

    # Monta lista ordenada para o template, com link WhatsApp por cliente
    lista_clientes = []
    for cod, dados in sorted(clientes.items()):
        pedidos = sorted(dados['pedidos'].values(), key=lambda x: x['numero'])
        emissao = pedidos[0]['emissao'] if pedidos else None
        emissao_fmt = emissao.strftime('%d/%m/%Y') if emissao else '-'

        linhas = [
            '*Pedidos para Expedir*',
            f"*Cliente: {dados['cod']} - {dados['nome']}*",
            f"Emissão: {emissao_fmt}",
            '',
        ]
        for pedido in pedidos:
            linhas.append(f"• Pedido {pedido['numero']} — {pedido['situacao']}")

        mensagem = '\n'.join(linhas)
        lista_clientes.append({
            'cod': dados['cod'],
            'nome': dados['nome'],
            'emissao': emissao_fmt,
            'pedidos': pedidos,
            'link_whatsapp': f"https://wa.me/5516991273974?text={quote(mensagem)}",
        })

    return render(request, 'analise/pedidos_nao_expedidos.html', {
        'titulo': 'Pedidos para Expedir',
        'lista_clientes': lista_clientes,
        'total': qs.values('numero_pedido').distinct().count(),
    })

@staff_member_required
def exportar_status_erp_excel(request):
    # 1. Pega os dados base (respeitando filtros de busca se houver)
    status_qs = por_empresa(StatusPedidoERP.objects.all(), request).order_by('-emissao', '-id')

    numero_pedido = request.GET.get('numero_pedido')
    if numero_pedido:
        status_qs = status_qs.filter(numero_pedido__icontains=numero_pedido)

    # 2. Prepara a lista de dicionários para o Pandas
    dados = []
    for item in status_qs:
        dados.append({
            'Emissão ERP': item.emissao,
            'Pedido': item.numero_pedido,
            'Cód. Cliente': item.cod_cliente,
            'Cliente': item.nome_cliente,
            'Situação': item.situacao,
            'Expedido': 'SIM' if item.expedido else 'NÃO'
        })

    # 3. Cria o DataFrame e o arquivo Excel
    df = pd.DataFrame(dados)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Relatorio_Status_ERP.xlsx'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Status')
    
    return response

@staff_member_required
def notificar_wishlist_whatsapp(request, cliente_id):
    cliente = get_empresa_or_404(WfClient, request, client_id=cliente_id)
    estado = cliente.client_state.uf_name
    
    itens_pendentes = ItemPedidoIgnorado.objects.filter(
        cliente=cliente, notificado=False, descartado_pelo_cliente=False,
        motivo_erro__icontains="estoque"
    )
    
    produtos_recuperados = []
    ids_para_atualizar = []

    codigos = [item.codigo_produto for item in itens_pendentes]
    produtos_map = {p.product_code: p for p in Product.objects.filter(product_code__in=codigos, empresa=cliente.empresa)}

    for item in itens_pendentes:
        produto = produtos_map.get(item.codigo_produto)
        if produto:
            preco_atual = getattr(produto, 'product_value_sp' if estado == 'SP' else 'product_value_es')
            if preco_atual and preco_atual > 0:
                qtd = item.quantidade_tentada if item.quantidade_tentada else 0
                produto_texto = f"- {produto.product_code}: {qtd} un. ({produto.product_description}) | R$ {preco_atual:.2f}".replace(".", ",")
                
                if produto_texto not in produtos_recuperados:
                    produtos_recuperados.append(produto_texto)
                ids_para_atualizar.append(item.id)
    
    if not produtos_recuperados:
        messages.warning(request, "Os produtos deste cliente ficaram sem estoque novamente.")
        return redirect('dashboard_admin')
        
    # ---------------------------------------------------------------------
    # NOVO: GERA O ID DO LOTE (Ex: REP-7X9P2K) E O LINK
    # ---------------------------------------------------------------------
    lote_id = f"REP-{get_random_string(6).upper()}"
    base_url = reverse('exportar_itens_recuperados_excel', args=[cliente.client_id])
    link_download = request.build_absolute_uri(f"{base_url}?lote={lote_id}") # <-- Link limpo e seguro!

    # Atualiza o banco com o lote gerado
    ItemPedidoIgnorado.objects.filter(id__in=ids_para_atualizar).update(
        notificado=True,
        data_notificacao=timezone.now(),
        lote_notificacao=lote_id # <-- Salva o Lote!
    )
    empresa_key = request.empresa.slug if request.empresa else 'superuser'
    cache.delete(f'dashboard_wishlist_{empresa_key}')
    
    produtos_texto = "\n".join(produtos_recuperados)
    mensagem = (
        f"Olá, {cliente.client_name}! Tudo bem?\n\n"
        f"Temos uma ótima notícia! Aqueles itens que você tentou pedir recentemente e estavam em falta, *acabaram de chegar no nosso estoque*:\n\n"
        f"{produtos_texto}\n\n"
        f"*Baixe a planilha de reposição aqui:* \n{link_download}\n\n"
        f"Gostaria de aproveitar e incluir no seu próximo pedido?"
    )
    
    link_whatsapp = f"https://api.whatsapp.com/send?text={quote(mensagem)}"
    messages.success(request, f"Cliente {cliente.client_name} notificado (Lote {lote_id})!")
    return redirect(link_whatsapp)

@login_required
@require_POST
def avisar_quando_disponivel(request):
    product_id = request.POST.get('product_id')
    produto = get_object_or_404(Product, product_id=product_id)

    try:
        cliente = request.user.wfclient
    except WfClient.DoesNotExist:
        return JsonResponse({'erro': 'Cliente não encontrado.'}, status=400)

    # Evita duplicatas: só registra se ainda não há aviso pendente para este produto/cliente
    ja_registrado = ItemPedidoIgnorado.objects.filter(
        cliente=cliente,
        codigo_produto=produto.product_code,
        notificado=False,
        motivo_erro__icontains='estoque'
    ).exists()

    if not ja_registrado:
        ItemPedidoIgnorado.objects.create(
            pedido=None,
            cliente=cliente,
            codigo_produto=produto.product_code,
            descricao_produto=produto.product_description,
            quantidade_tentada=0,
            motivo_erro='Sem estoque — cliente solicitou aviso via catálogo'
        )

    return JsonResponse({'sucesso': True, 'mensagem': 'Você será avisado quando o produto estiver disponível!'})

@login_required
def novidades(request):
    limite = timezone.now() - timedelta(days=15)

    produtos = por_empresa(Product.objects.filter(
        criado_em__gte=limite,
        status_estoque='DISPONIVEL'
    ), request).order_by('-criado_em')

    # Filtra pelo estado do cliente se não for staff
    preco_exibido = 'todos'
    cliente_logado = None

    if not request.user.is_staff:
        try:
            cliente_logado = request.user.wfclient
            estado = cliente_logado.client_state.uf_name
            if estado == 'SP':
                preco_exibido = 'sp'
                produtos = produtos.exclude(product_value_sp=0, status_estoque='DISPONIVEL')
            elif estado == 'ES':
                preco_exibido = 'es'
                produtos = produtos.exclude(product_value_es=0, status_estoque='DISPONIVEL')
        except WfClient.DoesNotExist:
            return redirect('home')

    for produto in produtos:
        produto.valor_sp_formatado = f"{produto.product_value_sp:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if produto.product_value_sp else "0,00"
        produto.valor_es_formatado = f"{produto.product_value_es:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if produto.product_value_es else "0,00"

    paginator = Paginator(produtos, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'novidades.html', {
        'titulo': 'Novidades',
        'page_obj': page_obj,
        'preco_exibido': preco_exibido,
        'cliente_logado': cliente_logado,
        'total_novidades': produtos.count(),
    })

@login_required
def meus_avisos(request):
    try:
        cliente = request.user.wfclient
    except WfClient.DoesNotExist:
        return redirect('home')

    pendentes = ItemPedidoIgnorado.objects.filter(
        cliente=cliente,
        notificado=False,
        descartado_pelo_cliente=False,
        motivo_erro__icontains='estoque'
    ).order_by('-data_tentativa')

    notificados = ItemPedidoIgnorado.objects.filter(
        cliente=cliente,
        notificado=True,
        motivo_erro__icontains='estoque'
    ).order_by('-data_notificacao')[:20]

    return render(request, 'meus_avisos.html', {
        'titulo': 'Meus Avisos',
        'pendentes': pendentes,
        'notificados': notificados,
    })

@login_required
@require_POST
def cancelar_aviso(request, item_id):
    try:
        cliente = request.user.wfclient
    except WfClient.DoesNotExist:
        return redirect('home')

    ItemPedidoIgnorado.objects.filter(
        id=item_id,
        cliente=cliente,
    ).update(descartado_pelo_cliente=True)

    messages.success(request, 'Aviso cancelado com sucesso.')
    return redirect('meus_avisos')

@login_required
@require_POST
def descartar_wishlist_home(request):
    product_code = request.POST.get('product_code', '').strip().upper()
    if not product_code:
        return JsonResponse({'erro': 'Código inválido.'}, status=400)
    try:
        cliente = request.user.wfclient
    except WfClient.DoesNotExist:
        return JsonResponse({'erro': 'Cliente não encontrado.'}, status=400)
    ItemPedidoIgnorado.objects.filter(
        cliente=cliente,
        codigo_produto=product_code,
        descartado_pelo_cliente=False,
    ).update(descartado_pelo_cliente=True)
    return JsonResponse({'sucesso': True})

@login_required
@require_POST
def adicionar_ao_carrinho_bd(request):
    product_id = request.POST.get('product_id')
    try:
        quantidade = int(request.POST.get('quantidade', 1))
    except (ValueError, TypeError):
        return JsonResponse({'erro': 'Quantidade inválida.'}, status=400)

    try:
        cliente_logado = request.user.wfclient
    except WfClient.DoesNotExist:
        return JsonResponse({'erro': 'Cliente não encontrado.'}, status=400)

    if quantidade <= 0:
        return JsonResponse({'erro': 'Quantidade inválida.'}, status=400)

    produto = get_object_or_404(Product, product_id=product_id)

    # 1. Pega o carrinho do cliente (ou cria um novo, vazio)
    carrinho, _ = Carrinho.objects.get_or_create(cliente=cliente_logado)

    # 2. Adiciona o item ao carrinho ou ATUALIZA a quantidade
    item, created = ItemCarrinho.objects.get_or_create(
        carrinho=carrinho,
        produto=produto,
        defaults={'quantidade': quantidade}
    )
    
    if not created:
        # CORREÇÃO APLICADA: Substitui o valor em vez de somar, 
        # garantindo a consistência com o input preenchido no frontend.
        item.quantidade = quantidade
        item.save()

    # =========================================================
    # --- NOVO: AUTO-LIMPEZA DA WISHLIST ---
    # Se o cliente adicionou ao carrinho, marcamos o erro como resolvido/notificado.
    # O produto nunca mais aparecerá no banner de "voltou ao estoque" para este cliente.
    # =========================================================
    from .models import ItemPedidoIgnorado # Importação local de segurança
    ItemPedidoIgnorado.objects.filter(
        cliente=cliente_logado,
        codigo_produto=produto.product_code,
        descartado_pelo_cliente=False
    ).update(descartado_pelo_cliente=True)

    # 3. Retorna os valores em tempo real para o Frontend atualizar a tela
    return JsonResponse({
        'sucesso': True, 
        'item_quantidade': item.quantidade,
        'item_total': float(item.get_subtotal()),
        'pedido_total': float(carrinho.get_total_carrinho()),
        'codigo_produto': produto.product_code # INJETADO: Enviamos o código de volta para animar a tela
    })

# =========================================================
# EXPORTAÇÃO DE ITENS RECUPERADOS (WISHLIST)
# =========================================================

from django.db.models import Max

def exportar_itens_recuperados_excel(request, cliente_id):
    cliente = get_object_or_404(WfClient, client_id=cliente_id)
    estado = cliente.client_state.uf_name
    
    # Captura o parâmetro ?lote=REP-XXXXXX
    lote_param = request.GET.get('lote')
    
    queryset = ItemPedidoIgnorado.objects.filter(
        cliente=cliente,
        motivo_erro__icontains="estoque"
    )

    if lote_param:
        # Busca EXATAMENTE os itens daquele lote
        queryset = queryset.filter(lote_notificacao=lote_param)
    else:
        # Fallback de segurança (mantido por precaução)
        ultima_notificacao = queryset.filter(notificado=True).aggregate(Max('data_notificacao'))['data_notificacao__max']
        if ultima_notificacao:
            queryset = queryset.filter(data_notificacao=ultima_notificacao)
        else:
            queryset = queryset.filter(notificado=False)

    itens_agrupados = {}

    for item in queryset:
        produto = Product.objects.filter(product_code=item.codigo_produto, empresa=cliente.empresa).first()
        if produto:
            preco_atual = getattr(produto, 'product_value_sp' if estado == 'SP' else 'product_value_es')
            
            if preco_atual and preco_atual > 0:
                codigo = item.codigo_produto
                qtd = item.quantidade_tentada or 0
                
                if codigo in itens_agrupados:
                    itens_agrupados[codigo]['Quantidade Solicitada'] += qtd
                    itens_agrupados[codigo]['Subtotal'] = float(preco_atual) * itens_agrupados[codigo]['Quantidade Solicitada']
                else:
                    itens_agrupados[codigo] = {
                        'Código': codigo,
                        'Descrição': produto.product_description,
                        'Quantidade Solicitada': qtd,
                        'Preço Unitário': float(preco_atual),
                        'Subtotal': float(preco_atual * qtd)
                    }

    data = list(itens_agrupados.values())

    if not data:
        return HttpResponse("Nenhum item disponível para exportação no momento.", status=404)

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Itens Recuperados')
    
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Coloquei o ID do lote no nome do arquivo para ficar ainda mais profissional!
    lote_nome = f"_{lote_param}" if lote_param else ""
    response['Content-Disposition'] = f'attachment; filename=recuperacao_{cliente.client_code}{lote_nome}.xlsx'
    return response

@staff_member_required
def historico_precos(request):
    filtro_codigo = request.GET.get('codigo', '').strip()
    filtro_descricao = request.GET.get('descricao', '').strip()

    historico = por_empresa(HistoricoPreco.objects.all(), request)

    if filtro_codigo:
        historico = historico.filter(product_code__icontains=filtro_codigo)
    if filtro_descricao:
        historico = historico.filter(product_description__icontains=filtro_descricao)

    paginator = Paginator(historico, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'analise/historico_precos.html', {
        'titulo': 'Histórico de Preços',
        'page_obj': page_obj,
        'filtro_codigo': filtro_codigo,
        'filtro_descricao': filtro_descricao,
    })

@staff_member_required
def historico_wishlist(request):
    """
    Exibe o histórico de lotes de reposição notificados aos clientes.
    """
    # Agrupa os itens arquivados pelo Lote e pelo Cliente
    lotes_arquivados = por_empresa(
        ItemPedidoIgnorado.objects.filter(notificado=True, motivo_erro__icontains="estoque"),
        request, campo='cliente__empresa'
    ).values(
        'lote_notificacao',
        'cliente__client_id',
        'cliente__client_code',
        'cliente__client_name'
    ).annotate(
        # Conta quantos produtos diferentes foram no lote
        qtd_produtos=Count('codigo_produto', distinct=True),
        # Soma a quantidade total de unidades pedidas
        total_unidades=Sum('quantidade_tentada'),
        # Pega a data exata do envio deste lote
        data_envio=Max('data_notificacao')
    ).order_by('-data_envio')

    # Paginação (agora podemos por menos por página, ex: 20 lotes)
    paginator = Paginator(lotes_arquivados, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    contexto = {
        'titulo': 'Histórico de Lotes Enviados',
        'page_obj': page_obj,
    }
    
    return render(request, 'analise/historico_wishlist.html', contexto)

@staff_member_required
def reenviar_notificacao_whatsapp(request, cliente_id):
    cliente = get_empresa_or_404(WfClient, request, client_id=cliente_id)
    estado = cliente.client_state.uf_name
    
    # 1. Pega o parâmetro do lote da URL (enviado pelo botão do painel)
    lote_param = request.GET.get('lote')
    
    itens_arquivados = ItemPedidoIgnorado.objects.filter(
        cliente=cliente,
        notificado=True,
        motivo_erro__icontains="estoque"
    )
    
    # 2. Filtra estritamente os itens do lote selecionado
    if lote_param:
        itens_arquivados = itens_arquivados.filter(lote_notificacao=lote_param)
    
    produtos_recuperados_dict = {} # <-- Dicionário para agrupar o texto
    
    for item in itens_arquivados:
        produto = Product.objects.filter(product_code=item.codigo_produto, empresa=cliente.empresa).first()
        if produto:
            preco_atual = getattr(produto, 'product_value_sp' if estado == 'SP' else 'product_value_es')
            if preco_atual and preco_atual > 0:
                codigo = produto.product_code
                qtd = item.quantidade_tentada or 0
                
                # Agrupa a quantidade para a mensagem de texto
                if codigo in produtos_recuperados_dict:
                    produtos_recuperados_dict[codigo]['qtd'] += qtd
                else:
                    produtos_recuperados_dict[codigo] = {
                        'descricao': produto.product_description,
                        'preco': preco_atual,
                        'qtd': qtd
                    }
    
    if not produtos_recuperados_dict:
        messages.warning(request, "Nenhum produto deste lote está disponível no momento.")
        return redirect('historico_wishlist')

    # Monta a lista de textos já agrupada
    produtos_recuperados_texto = []
    for cod, dados in produtos_recuperados_dict.items():
        texto = f"- {cod}: {dados['qtd']} un. ({dados['descricao']}) | R$ {dados['preco']:.2f}".replace(".", ",")
        produtos_recuperados_texto.append(texto)

    # 3. Gerar link limpo usando o Lote (e não mais IDs gigantes)
    base_url = reverse('exportar_itens_recuperados_excel', args=[cliente.client_id])
    if lote_param:
        link_download = request.build_absolute_uri(f"{base_url}?lote={lote_param}")
    else:
        # Fallback para envios antigos sem lote
        link_download = request.build_absolute_uri(base_url)
    
    produtos_texto = "\n".join(produtos_recuperados_texto) 
    
    mensagem = (
        f"Olá, {cliente.client_name}!\n\n"
        f"Temos uma ótima notícia! Aqueles itens que você tentou pedir recentemente e estavam em falta, *acabaram de chegar no nosso estoque*:\n\n"
        f"{produtos_texto}\n\n"
        f"*Baixe a planilha de reposição aqui:* \n{link_download}\n\n"
        f"Gostaria de aproveitar e incluir no seu próximo pedido?"
    )
    
    return redirect(f"https://api.whatsapp.com/send?text={quote(mensagem)}")


@staff_member_required
def sugestoes_admin(request):
    filtro_cliente = request.GET.get('cliente', '').strip()

    # Recalcular sugestões de um cliente específico via POST
    if request.method == 'POST':
        cliente_code = request.POST.get('recalcular_cliente')
        if cliente_code:
            processar_giro_cliente(cliente_code)
        return redirect(request.get_full_path())

    # Sem filtro: só mostra a barra de busca
    if not filtro_cliente:
        return render(request, 'analise/sugestoes_admin.html', {
            'clientes': None,
            'filtro_cliente': '',
        })

    sugestoes_qs = por_empresa(SugestaoCompraERP.objects.select_related(
        'cliente', 'cliente__client_state'
    ).filter(
        Q(cliente__client_name__icontains=filtro_cliente) |
        Q(cliente__client_code__icontains=filtro_cliente)
    ), request, campo='cliente__empresa').order_by('cliente__client_name', '-giro_diario')

    # Se não há sugestões calculadas, tenta calcular para os clientes encontrados
    if not sugestoes_qs.exists():
        clientes_encontrados = por_empresa(WfClient.objects.filter(
            Q(client_name__icontains=filtro_cliente) |
            Q(client_code__icontains=filtro_cliente)
        ), request)
        for cliente in clientes_encontrados:
            processar_giro_cliente(cliente.client_code)
        # Re-executa a query após calcular
        sugestoes_qs = SugestaoCompraERP.objects.select_related(
            'cliente', 'cliente__client_state'
        ).filter(
            Q(cliente__client_name__icontains=filtro_cliente) |
            Q(cliente__client_code__icontains=filtro_cliente)
        ).order_by('cliente__client_name', '-giro_diario')

    clientes_map = {}
    for s in sugestoes_qs:
        cid = s.cliente.client_id
        if cid not in clientes_map:
            clientes_map[cid] = {'cliente': s.cliente, 'sugestoes': [], 'total_itens': 0}
        clientes_map[cid]['sugestoes'].append(s)
        clientes_map[cid]['total_itens'] += 1

    return render(request, 'analise/sugestoes_admin.html', {
        'clientes': clientes_map.values(),
        'filtro_cliente': filtro_cliente,
        'total_clientes': len(clientes_map),
    })


@staff_member_required
def dashboard_analise(request):
    hoje = date.today()
    mes_selecionado = int(request.GET.get('mes', hoje.month))
    ano_selecionado = int(request.GET.get('ano', hoje.year))

    contexto = gerar_dados_dashboard_analise(mes_selecionado, ano_selecionado, empresa=request.empresa)

    return render(request, 'analise/dashboard_analise.html', contexto)


@staff_member_required
def evolucao_clientes(request):
    ano = request.GET.get('ano')
    try:
        ano = int(ano)
    except (TypeError, ValueError):
        ano = date.today().year

    dados = calcular_evolucao_clientes(empresa=request.empresa, ano=ano)

    # Anos disponíveis para o filtro (baseado nos dados existentes)
    from django.db.models.functions import ExtractYear
    qs_anos = VendaReal.objects.all()
    if request.empresa:
        qs_anos = qs_anos.filter(empresa=request.empresa)
    anos_disponiveis = sorted(
        qs_anos.annotate(y=ExtractYear('Emissao')).values_list('y', flat=True).distinct(),
        reverse=True
    )

    dados['anos_disponiveis'] = anos_disponiveis
    dados['ano_selecionado'] = ano

    # Nomes dos meses em PT
    MESES_PT = ['', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    dados['meses_nomes'] = {m: MESES_PT[m] for m in range(1, 13)}

    return render(request, 'analise/evolucao_clientes.html', dados)


@staff_member_required
def exportar_evolucao_clientes_excel(request):
    ano = request.GET.get('ano')
    try:
        ano = int(ano)
    except (TypeError, ValueError):
        ano = date.today().year

    dados = calcular_evolucao_clientes(empresa=request.empresa, ano=ano)

    MESES_PT = ['', 'Janeiro', 'Fevereiro', 'Marco', 'Abril', 'Maio', 'Junho',
                'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Evolucao {ano}'

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    total_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
    green_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    red_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')

    meses = dados['meses_com_dados']
    headers = ['Cod.', 'Cliente'] + [MESES_PT[m] for m in meses] + ['Total Ano', f'Total {dados["ano_anterior"]}', 'Var %']
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    fmt_brl = '#,##0.00'
    for c in dados['clientes']:
        row = [c['codigo'], c['nome']]
        for m in meses:
            row.append(float(c['meses'][m]))
        row.append(float(c['total_ano']))
        row.append(float(c['total_ant']))
        row.append(round(c['variacao_pct'], 1))
        ws.append(row)

        row_num = ws.max_row
        # Formata valores monetários
        for col_idx in range(3, 3 + len(meses) + 2):
            ws.cell(row=row_num, column=col_idx).number_format = fmt_brl

        # Colorir variação
        var_col = 3 + len(meses) + 2
        var_cell = ws.cell(row=row_num, column=var_col)
        var_cell.number_format = '0.0"%"'
        if c['variacao_pct'] >= 10:
            var_cell.fill = green_fill
        elif c['variacao_pct'] <= -10:
            var_cell.fill = red_fill

    # Linha de totais
    total_row = ['', 'TOTAL GERAL']
    for m in meses:
        total_row.append(float(dados['totais_mensais'][m]))
    total_row.append(float(dados['total_geral']))
    total_row.append('')
    total_row.append('')
    ws.append(total_row)
    for col_idx in range(3, 3 + len(meses) + 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.number_format = fmt_brl

    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    wb.save(output)
    output.seek(0)
    filename = f'evolucao_clientes_{ano}.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@staff_member_required
def exportar_rfm_pdf(request):
    from fpdf import FPDF
    from wefixhub.utils import calcular_rfm

    rfm = calcular_rfm(empresa=request.empresa)
    hoje = date.today()
    empresa_nome = request.empresa.nome if request.empresa else 'WFHUB'

    # Paleta de cores por segmento (R, G, B)
    CORES = {
        'Campeão':    (16,  185, 129),
        'Fiel':       (59,  130, 246),
        'Potencial':  (139, 92,  246),
        'Em Risco':   (245, 158, 11),
        'Adormecido': (107, 114, 128),
    }
    BG = {
        'Campeão':    (209, 250, 229),
        'Fiel':       (219, 234, 254),
        'Potencial':  (237, 233, 254),
        'Em Risco':   (254, 243, 199),
        'Adormecido': (243, 244, 246),
    }
    ACOES = {
        'Campeão':    'Fidelize e recompense',
        'Fiel':       'Cross-sell, venda mais',
        'Potencial':  'Converta em recorrente',
        'Em Risco':   'Ligar hoje - risco real de perda',
        'Adormecido': 'Ultima tentativa de reativacao',
    }

    def _pdf_str(s):
        """Remove chars fora do latin-1 para compatibilidade com Helvetica."""
        return s.encode('latin-1', errors='replace').decode('latin-1')

    class RFMPdf(FPDF):
        def header(self):
            self.set_fill_color(30, 41, 59)
            self.rect(0, 0, 210, 22, 'F')
            self.set_font('Helvetica', 'B', 14)
            self.set_text_color(255, 255, 255)
            self.set_y(6)
            self.cell(0, 10, f'Analise RFM - {empresa_nome}', align='L', new_x='LMARGIN', new_y='NEXT')
            self.set_font('Helvetica', '', 8)
            self.set_text_color(180, 190, 210)
            self.set_y(14)
            self.cell(0, 6, f'Gerado em {hoje.strftime("%d/%m/%Y")} | Ultimos 12 meses | {rfm["total"]} clientes analisados', align='L')
            self.ln(14)

        def footer(self):
            self.set_y(-12)
            self.set_font('Helvetica', '', 7)
            self.set_text_color(160, 160, 160)
            self.cell(0, 5, f'Pagina {self.page_no()} - RFM gerado por WFHUB', align='C')

    pdf = RFMPdf(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.set_margins(12, 12, 12)
    pdf.add_page()

    # ── Resumo por segmento ──────────────────────────────────────────
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 6, 'RESUMO POR SEGMENTO', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(1)

    card_w = (pdf.w - 24 - 4 * 4) / 5  # 5 cards com gap
    segs = rfm['segmentos']
    x0 = pdf.l_margin
    for nome, seg in segs.items():
        cor = CORES.get(nome, (100, 100, 100))
        bg  = BG.get(nome,  (240, 240, 240))
        acao = ACOES.get(nome, '')
        # Card background
        pdf.set_fill_color(*bg)
        pdf.set_draw_color(*cor)
        pdf.rect(x0, pdf.get_y(), card_w, 22, 'FD')
        # Barra colorida no topo do card
        pdf.set_fill_color(*cor)
        pdf.rect(x0, pdf.get_y(), card_w, 2, 'F')
        # Conteúdo
        pdf.set_text_color(*cor)
        pdf.set_font('Helvetica', 'B', 7)
        pdf.set_xy(x0 + 2, pdf.get_y() + 4)
        pdf.cell(card_w - 4, 4, _pdf_str(nome.upper()))
        pdf.set_text_color(30, 41, 59)
        pdf.set_font('Helvetica', 'B', 18)
        pdf.set_xy(x0 + 2, pdf.get_y() + 4)
        pdf.cell(card_w - 4, 8, str(seg['count']))
        pdf.set_text_color(100, 116, 139)
        pdf.set_font('Helvetica', '', 6)
        pdf.set_xy(x0 + 2, pdf.get_y() + 6)
        pdf.multi_cell(card_w - 4, 3, _pdf_str(acao))
        x0 += card_w + 4

    pdf.ln(28)

    # ── Tabela de clientes ───────────────────────────────────────────
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 6, 'CLIENTES - DETALHAMENTO RFM', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(1)

    # Cabeçalho da tabela
    cols = [
        ('Cod.',         14),
        ('Cliente',      72),
        ('Segmento',     28),
        ('Recencia',     22),
        ('Frequencia',   22),
        ('Valor (12m)',  38),
        ('Score',        18),
        ('Ult. Compra',  26),
    ]
    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 7.5)
    for label, w in cols:
        pdf.cell(w, 7, label, border=0, fill=True, align='C')
    pdf.ln()

    # Linhas
    pdf.set_font('Helvetica', '', 7)
    for i, c in enumerate(rfm['clientes']):
        seg = c['segmento']
        cor = CORES.get(seg, (100, 100, 100))
        bg  = BG.get(seg, (250, 250, 250))

        # Zebra leve alternando branco / bg do segmento
        if i % 2 == 0:
            pdf.set_fill_color(*bg)
        else:
            pdf.set_fill_color(255, 255, 255)

        pdf.set_text_color(30, 41, 59)
        row_h = 6
        pdf.cell(cols[0][1], row_h, str(c['codigo']),              fill=True, align='C')
        pdf.cell(cols[1][1], row_h, _pdf_str(c['nome'][:38]),      fill=True)
        # Segmento com cor
        pdf.set_text_color(*cor)
        pdf.set_font('Helvetica', 'B', 7)
        pdf.cell(cols[2][1], row_h, _pdf_str(seg),                 fill=True, align='C')
        pdf.set_text_color(30, 41, 59)
        pdf.set_font('Helvetica', '', 7)
        pdf.cell(cols[3][1], row_h, f"{c['recencia']}d",    fill=True, align='C')
        pdf.cell(cols[4][1], row_h, f"{c['frequencia']}x",  fill=True, align='C')
        pdf.cell(cols[5][1], row_h, f"R$ {c['monetario_fmt']}", fill=True, align='R')
        pdf.cell(cols[6][1], row_h, f"{c['rfm_score']}/15", fill=True, align='C')
        pdf.cell(cols[7][1], row_h, c['ultima_compra'],     fill=True, align='C')
        pdf.ln()

    buf = bytes(pdf.output())
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="RFM_{hoje.strftime("%Y%m%d")}.pdf"'
    return resp


@login_required
def sugestoes_inteligentes_erp(request):
    try:
        cliente_logado = request.user.wfclient
    except WfClient.DoesNotExist:
        return redirect('home')

    # Para evitar gargalos, você pode rodar isso assincronamente (ex: Celery) no futuro.
    # Por enquanto, rodamos on-the-fly. O processamento dura milissegundos graças ao 'annotate'.
    processar_giro_cliente(cliente_logado.client_code)

    # Busca os resultados do backup ordenando pelos itens de maior giro
    sugestoes_qs = SugestaoCompraERP.objects.filter(
        cliente=cliente_logado
    ).order_by('-giro_diario')

    # Paginação para não sobrecarregar a tela
    paginator = Paginator(sugestoes_qs, 20)
    page_number = request.GET.get('page')
    sugestoes_paginadas = paginator.get_page(page_number)

    contexto = {
        'titulo': 'Reposição de Estoque Inteligente',
        'sugestoes': sugestoes_paginadas,
        'cliente_logado': cliente_logado,
    }
    
    return render(request, 'sugestoes_inteligentes.html', contexto)


@login_required
def adicionar_sugestoes_ao_carrinho(request):
    if request.method == 'POST':
        try:
            cliente = request.user.wfclient
        except WfClient.DoesNotExist:
            return redirect('home')

        # CAPTURA OS CHECKBOXES: Pega a lista de códigos que o cliente deixou marcado
        produtos_selecionados = request.POST.getlist('produtos_selecionados')
        
        if not produtos_selecionados:
            messages.warning(request, "Nenhum produto foi selecionado para adicionar ao carrinho.")
            return redirect('sugestoes_inteligentes_erp')

        # FILTRA AS SUGESTÕES: Agora só processa o que o cliente escolheu
        sugestoes = SugestaoCompraERP.objects.filter(
            cliente=cliente, 
            produto_codigo__in=produtos_selecionados
        )
        
        if not sugestoes.exists():
            messages.warning(request, "As sugestões selecionadas não são mais válidas.")
            return redirect('sugestoes_inteligentes_erp')

        with transaction.atomic():
            carrinho, _ = Carrinho.objects.get_or_create(cliente=cliente)

            codigos_sugeridos = sugestoes.values_list('produto_codigo', flat=True)
            produtos_catalogo = {
                p.product_code: p
                for p in por_empresa(Product.objects, request).filter(product_code__in=codigos_sugeridos)
            }

            itens_carrinho_existentes = {
                item.produto.product_code: item 
                for item in ItemCarrinho.objects.filter(carrinho=carrinho).select_related('produto')
            }

            itens_para_criar = []
            itens_para_atualizar = []
            codigos_processados_agora = set() 

            for sug in sugestoes:
                codigo = sug.produto_codigo
                
                if codigo in codigos_processados_agora:
                    continue
                codigos_processados_agora.add(codigo)

                produto_real = produtos_catalogo.get(codigo)
                
                if produto_real:
                    if codigo in itens_carrinho_existentes:
                        item_existente = itens_carrinho_existentes[codigo]
                        item_existente.quantidade = sug.quantidade_sugerida
                        itens_para_atualizar.append(item_existente)
                    else:
                        itens_para_criar.append(ItemCarrinho(
                            carrinho=carrinho,
                            produto=produto_real,
                            quantidade=sug.quantidade_sugerida
                        ))

            if itens_para_criar:
                ItemCarrinho.objects.bulk_create(itens_para_criar)
            if itens_para_atualizar:
                ItemCarrinho.objects.bulk_update(itens_para_atualizar, ['quantidade'])

        messages.success(request, f"{len(codigos_processados_agora)} sugestões adicionadas ao carrinho!")
        return redirect('carrinho') 

    return redirect('sugestoes_inteligentes_erp')

# ==========================================
# SAAS — CADASTRO DE EMPRESA
# ==========================================

@login_required
def cadastrar_empresa(request):
    if not request.user.is_superuser:
        return redirect('home')

    if request.method == 'POST':
        form = CadastroEmpresaForm(request.POST)
        if form.is_valid():
            d = form.cleaned_data

            # Cria a empresa
            empresa = Empresa.objects.create(
                nome=d['nome'],
                slug=d['slug'],
                plano=d['plano'],
                email_contato=d['email_contato'],
                telefone=d['telefone'],
            )

            # Cria o usuário administrador da empresa
            usuario = User.objects.create_user(
                username=d['username'],
                email=d['email_usuario'],
                password=d['senha'],
                is_staff=True,
            )

            # Liga o usuário à empresa
            PerfilUsuario.objects.create(
                user=usuario,
                empresa=empresa,
                papel='ADMIN',
            )

            messages.success(request, f'Empresa "{empresa.nome}" criada com sucesso!')
            return redirect('listar_empresas')
    else:
        form = CadastroEmpresaForm()

    return render(request, 'saas/cadastrar_empresa.html', {'form': form, 'titulo': 'Nova Empresa'})


@login_required
def listar_empresas(request):
    if not request.user.is_superuser:
        return redirect('home')

    empresas = Empresa.objects.prefetch_related('membros__user').order_by('nome')
    return render(request, 'saas/listar_empresas.html', {'empresas': empresas, 'titulo': 'Empresas Cadastradas'})


@login_required
def detalhe_empresa(request, empresa_id):
    if not request.user.is_superuser:
        return redirect('home')

    empresa = get_object_or_404(Empresa, id=empresa_id)

    if request.method == 'POST' and request.POST.get('action') == 'editar':
        nome = request.POST.get('nome', '').strip()
        plano = request.POST.get('plano', 'FREE')
        email_contato = request.POST.get('email_contato', '').strip() or None
        telefone = request.POST.get('telefone', '').strip() or None
        expira_em_raw = request.POST.get('expira_em', '').strip() or None

        if nome:
            empresa.nome = nome
            empresa.plano = plano
            empresa.email_contato = email_contato
            empresa.telefone = telefone
            empresa.expira_em = expira_em_raw
            empresa.save(update_fields=['nome', 'plano', 'email_contato', 'telefone', 'expira_em'])
            messages.success(request, 'Empresa atualizada.')

        return redirect('detalhe_empresa', empresa_id=empresa.id)

    membros = empresa.membros.select_related('user').order_by('papel', 'user__username')
    clientes = WfClient.objects.filter(empresa=empresa).select_related('user', 'client_state').order_by('client_name')
    clientes_count = clientes.count()
    produtos_count = Product.objects.filter(empresa=empresa).count()
    pedidos_count = Pedido.objects.filter(empresa=empresa).count()
    vendas_count = VendaReal.objects.filter(empresa=empresa).count()

    checklist = [
        {'label': 'Empresa criada', 'ok': True},
        {'label': 'E-mail de contato configurado', 'ok': bool(empresa.email_contato)},
        {'label': 'Pelo menos 1 membro adicionado', 'ok': membros.exists()},
        {'label': 'Pelo menos 1 cliente cadastrado', 'ok': clientes_count > 0},
        {'label': 'Produtos carregados', 'ok': produtos_count > 0},
    ]
    checklist_completo = all(item['ok'] for item in checklist)
    checklist_done = sum(1 for item in checklist if item['ok'])

    ctx = {
        'empresa': empresa,
        'membros': membros,
        'clientes': clientes,
        'clientes_count': clientes_count,
        'produtos_count': produtos_count,
        'pedidos_count': pedidos_count,
        'vendas_count': vendas_count,
        'plano_choices': Empresa.PLANO_CHOICES,
        'checklist': checklist,
        'checklist_completo': checklist_completo,
        'checklist_done': checklist_done,
    }
    return render(request, 'saas/detalhe_empresa.html', ctx)


@login_required
def toggle_empresa_ativo(request, empresa_id):
    if not request.user.is_superuser:
        return redirect('home')
    if request.method == 'POST':
        empresa = get_object_or_404(Empresa, id=empresa_id)
        empresa.ativo = not empresa.ativo
        empresa.save(update_fields=['ativo'])
        status = 'ativada' if empresa.ativo else 'desativada'
        messages.success(request, f'Empresa "{empresa.nome}" {status}.')
    return redirect('detalhe_empresa', empresa_id=empresa_id)


@login_required
def adicionar_membro_empresa(request, empresa_id):
    if not request.user.is_superuser:
        return redirect('home')
    if request.method != 'POST':
        return redirect('detalhe_empresa', empresa_id=empresa_id)

    empresa = get_object_or_404(Empresa, id=empresa_id)
    username = request.POST.get('username', '').strip()
    email = request.POST.get('email', '').strip()
    senha = request.POST.get('senha', '').strip()
    papel = request.POST.get('papel', 'VENDEDOR')

    if not username:
        messages.error(request, 'Username obrigatório.')
        return redirect('detalhe_empresa', empresa_id=empresa_id)

    # Tenta encontrar usuário existente; se não, cria
    usuario = User.objects.filter(username=username).first()
    if usuario:
        # Verifica se já tem perfil em outra empresa
        if hasattr(usuario, 'perfil') and usuario.perfil.empresa != empresa:
            messages.error(request, f'O username "{username}" já está em uso. Cada usuário só pode pertencer a uma empresa. Escolha um username diferente, por exemplo: "{username}_{empresa.slug}".')
            return redirect('detalhe_empresa', empresa_id=empresa_id)
        if hasattr(usuario, 'perfil'):
            messages.warning(request, f'O usuário "{username}" já é membro desta empresa.')
            return redirect('detalhe_empresa', empresa_id=empresa_id)
    else:
        if not senha or not email:
            messages.error(request, 'E-mail e senha obrigatórios para criar novo usuário.')
            return redirect('detalhe_empresa', empresa_id=empresa_id)
        usuario = User.objects.create_user(
            username=username,
            email=email,
            password=senha,
            is_staff=True,
        )

    PerfilUsuario.objects.create(user=usuario, empresa=empresa, papel=papel)
    messages.success(request, f'Usuário "{username}" adicionado como {papel}.')
    return redirect('detalhe_empresa', empresa_id=empresa_id)


@login_required
def remover_membro_empresa(request, empresa_id, membro_id):
    if not request.user.is_superuser:
        return redirect('home')
    if request.method == 'POST':
        membro = get_object_or_404(PerfilUsuario, id=membro_id, empresa_id=empresa_id)
        username = membro.user.username
        membro.delete()
        messages.success(request, f'Membro "{username}" removido.')
    return redirect('detalhe_empresa', empresa_id=empresa_id)


# ==============================================================================
# GESTÃO DE CLIENTES
# ==============================================================================

@staff_member_required
def listar_clientes(request):
    busca = request.GET.get('q', '').strip()
    clientes_qs = por_empresa(WfClient.objects.select_related('client_state'), request).order_by('client_name')

    if busca:
        clientes_qs = clientes_qs.filter(
            Q(client_name__icontains=busca) |
            Q(client_code__icontains=busca) |
            Q(client_cnpj__icontains=busca) |
            Q(client_city__icontains=busca)
        )

    paginator = Paginator(clientes_qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'clientes/listar_clientes.html', {
        'page_obj': page_obj,
        'busca': busca,
    })


@staff_member_required
def cadastrar_cliente(request):
    from .models import wefixhub_uf
    ufs = wefixhub_uf.objects.all().order_by('uf_name')

    if request.method == 'POST':
        erros = {}

        client_code = request.POST.get('client_code', '').strip()
        client_name = request.POST.get('client_name', '').strip()
        client_cnpj = request.POST.get('client_cnpj', '').strip()
        client_adress = request.POST.get('client_adress', '').strip()
        client_city = request.POST.get('client_city', '').strip()
        client_cep = request.POST.get('client_cep', '').strip().replace('-', '')
        client_state_id = request.POST.get('client_state', '')
        client_state_subscription = request.POST.get('client_state_subscription', '').strip()
        client_is_active = request.POST.get('client_is_active') == 'on'
        frete_preferencia = request.POST.get('frete_preferencia', 'CORREIOS')
        nota_fiscal_preferencia = request.POST.get('nota_fiscal_preferencia', 'SEM')
        observacao_preferencia = request.POST.get('observacao_preferencia', '').strip()

        # Criar login de acesso ao portal (opcional)
        criar_login = request.POST.get('criar_login') == 'on'
        username = request.POST.get('username', '').strip()
        email_login = request.POST.get('email_login', '').strip()
        senha_login = request.POST.get('senha_login', '').strip()

        if not client_code:
            erros['client_code'] = 'Código obrigatório.'
        elif WfClient.objects.filter(empresa=request.empresa, client_code=client_code).exists():
            erros['client_code'] = 'Código já cadastrado nesta empresa.'

        if not client_name:
            erros['client_name'] = 'Nome obrigatório.'

        if not client_cnpj:
            erros['client_cnpj'] = 'CNPJ obrigatório.'
        elif WfClient.objects.filter(empresa=request.empresa, client_cnpj=client_cnpj).exists():
            erros['client_cnpj'] = 'CNPJ já cadastrado nesta empresa.'

        if not client_state_id:
            erros['client_state'] = 'Estado obrigatório.'

        if criar_login:
            if not username:
                erros['username'] = 'Username obrigatório para login.'
            elif User.objects.filter(username=username).exists():
                erros['username'] = 'Username já existe.'
            if not senha_login:
                erros['senha_login'] = 'Senha obrigatória.'

        if erros:
            return render(request, 'clientes/cadastrar_cliente.html', {
                'ufs': ufs, 'erros': erros, 'post': request.POST,
                'frete_choices': WfClient.FRETE_CHOICES,
                'nf_choices': WfClient.NOTA_FISCAL_CHOICES,
            })

        uf = get_object_or_404(wefixhub_uf, pk=client_state_id)

        usuario = None
        if criar_login:
            usuario = User.objects.create_user(username=username, email=email_login, password=senha_login)

        WfClient.objects.create(
            empresa=request.empresa,
            user=usuario,
            client_code=int(client_code),
            client_name=client_name,
            client_cnpj=client_cnpj,
            client_adress=client_adress,
            client_city=client_city,
            client_cep=client_cep,
            client_state=uf,
            client_state_subscription=client_state_subscription or None,
            client_is_active=client_is_active,
            frete_preferencia=frete_preferencia,
            nota_fiscal_preferencia=nota_fiscal_preferencia,
            observacao_preferencia=observacao_preferencia or None,
        )

        messages.success(request, f'Cliente "{client_name}" cadastrado com sucesso!')
        return redirect('listar_clientes')

    return render(request, 'clientes/cadastrar_cliente.html', {
        'ufs': ufs,
        'erros': {},
        'post': {},
        'frete_choices': WfClient.FRETE_CHOICES,
        'nf_choices': WfClient.NOTA_FISCAL_CHOICES,
    })


@staff_member_required
def editar_cliente(request, client_id):
    from .models import wefixhub_uf
    cliente = get_empresa_or_404(WfClient, request, client_id=client_id)
    ufs = wefixhub_uf.objects.all().order_by('uf_name')

    if request.method == 'POST':
        erros = {}

        client_name = request.POST.get('client_name', '').strip()
        client_cnpj = request.POST.get('client_cnpj', '').strip()
        client_adress = request.POST.get('client_adress', '').strip()
        client_city = request.POST.get('client_city', '').strip()
        client_cep = request.POST.get('client_cep', '').strip().replace('-', '')
        client_state_id = request.POST.get('client_state', '')
        client_state_subscription = request.POST.get('client_state_subscription', '').strip()
        client_is_active = request.POST.get('client_is_active') == 'on'
        frete_preferencia = request.POST.get('frete_preferencia', 'CORREIOS')
        nota_fiscal_preferencia = request.POST.get('nota_fiscal_preferencia', 'SEM')
        observacao_preferencia = request.POST.get('observacao_preferencia', '').strip()

        if not client_name:
            erros['client_name'] = 'Nome obrigatório.'

        if not client_cnpj:
            erros['client_cnpj'] = 'CNPJ obrigatório.'
        elif WfClient.objects.filter(empresa=request.empresa, client_cnpj=client_cnpj).exclude(client_id=cliente.client_id).exists():
            erros['client_cnpj'] = 'CNPJ já cadastrado em outro cliente desta empresa.'

        if not client_state_id:
            erros['client_state'] = 'Estado obrigatório.'

        if erros:
            return render(request, 'clientes/editar_cliente.html', {
                'cliente': cliente, 'ufs': ufs, 'erros': erros,
                'frete_choices': WfClient.FRETE_CHOICES,
                'nf_choices': WfClient.NOTA_FISCAL_CHOICES,
            })

        uf = get_object_or_404(wefixhub_uf, pk=client_state_id)

        cliente.client_name = client_name
        cliente.client_cnpj = client_cnpj
        cliente.client_adress = client_adress
        cliente.client_city = client_city
        cliente.client_cep = client_cep
        cliente.client_state = uf
        cliente.client_state_subscription = client_state_subscription or None
        cliente.client_is_active = client_is_active
        cliente.frete_preferencia = frete_preferencia
        cliente.nota_fiscal_preferencia = nota_fiscal_preferencia
        cliente.observacao_preferencia = observacao_preferencia or None
        cliente.save()

        messages.success(request, f'Cliente "{cliente.client_name}" atualizado.')
        return redirect('listar_clientes')

    enderecos = Endereco.objects.filter(cliente=cliente).order_by('-is_default', 'id')
    form_endereco = EnderecoForm()
    return render(request, 'clientes/editar_cliente.html', {
        'cliente': cliente,
        'ufs': ufs,
        'erros': {},
        'frete_choices': WfClient.FRETE_CHOICES,
        'nf_choices': WfClient.NOTA_FISCAL_CHOICES,
        'enderecos': enderecos,
        'form_endereco': form_endereco,
    })


@login_required
@staff_member_required
def adicionar_endereco_cliente_staff(request, client_id):
    cliente = get_empresa_or_404(WfClient, request, client_id=client_id)
    if request.method == 'POST':
        form = EnderecoForm(request.POST)
        if form.is_valid():
            novo = form.save(commit=False)
            novo.cliente = cliente
            novo.save()
            messages.success(request, 'Endereço adicionado.')
        else:
            messages.error(request, 'Verifique os campos do endereço.')
    return redirect('editar_cliente', client_id=client_id)


@login_required
@staff_member_required
def excluir_endereco_cliente_staff(request, client_id, endereco_id):
    cliente = get_empresa_or_404(WfClient, request, client_id=client_id)
    endereco = get_object_or_404(Endereco, id=endereco_id, cliente=cliente)
    if request.method == 'POST':
        endereco.delete()
        messages.success(request, 'Endereço removido.')
    return redirect('editar_cliente', client_id=client_id)


@login_required
@staff_member_required
def editar_endereco_cliente_staff(request, client_id, endereco_id):
    from .models import wefixhub_uf
    cliente = get_empresa_or_404(WfClient, request, client_id=client_id)
    endereco = get_object_or_404(Endereco, id=endereco_id, cliente=cliente)
    ufs = wefixhub_uf.objects.all().order_by('uf_name')
    if request.method == 'POST':
        form = EnderecoForm(request.POST, instance=endereco)
        if form.is_valid():
            form.save()
            messages.success(request, 'Endereço atualizado.')
            return redirect('editar_cliente', client_id=client_id)
    else:
        form = EnderecoForm(instance=endereco)
    return render(request, 'clientes/editar_endereco_staff.html', {
        'cliente': cliente,
        'form': form,
        'endereco': endereco,
        'ufs': ufs,
    })


# ==============================================================================
# IMPERSONAÇÃO E AÇÕES RÁPIDAS (superuser)
# ==============================================================================

@login_required
def impersonar_cliente(request, empresa_id, client_id):
    if not request.user.is_superuser:
        return redirect('home')
    cliente = get_object_or_404(WfClient, client_id=client_id, empresa_id=empresa_id)
    if not cliente.user:
        messages.error(request, 'Este cliente não possui login vinculado.')
        return redirect('detalhe_empresa', empresa_id=empresa_id)
    su_id = request.user.id
    from django.contrib.auth import login as auth_login
    auth_login(request, cliente.user, backend='django.contrib.auth.backends.ModelBackend')
    request.session['impersonando_su_id'] = su_id  # restaura após flush do login
    return redirect('home')


@login_required
def impersonar_membro(request, empresa_id, membro_id):
    if not request.user.is_superuser:
        return redirect('home')
    membro = get_object_or_404(PerfilUsuario, id=membro_id, empresa_id=empresa_id)
    su_id = request.user.id
    from django.contrib.auth import login as auth_login
    auth_login(request, membro.user, backend='django.contrib.auth.backends.ModelBackend')
    request.session['impersonando_su_id'] = su_id  # restaura após flush do login
    return redirect('home')


@login_required
def impersonar_cliente_staff(request, client_id):
    """Permite que membros is_staff façam impersonação de clientes da própria empresa."""
    if not request.user.is_staff:
        return redirect('home')
    cliente = get_object_or_404(WfClient, client_id=client_id, empresa=request.empresa)
    if not cliente.user:
        messages.error(request, 'Este cliente não possui login vinculado.')
        return redirect('listar_clientes')
    staff_id = request.user.id
    from django.contrib.auth import login as auth_login
    auth_login(request, cliente.user, backend='django.contrib.auth.backends.ModelBackend')
    request.session['impersonando_su_id'] = staff_id
    return redirect('home')


@login_required
def sair_impersonacao(request):
    su_id = request.session.get('impersonando_su_id')
    if su_id:
        original = get_object_or_404(User, id=su_id)
        if not (original.is_staff or original.is_superuser):
            return redirect('home')
        del request.session['impersonando_su_id']
        from django.contrib.auth import login as auth_login
        auth_login(request, original, backend='django.contrib.auth.backends.ModelBackend')
        messages.success(request, 'Impersonação encerrada.')
    if request.user.is_superuser:
        return redirect('saas_dashboard')
    return redirect('listar_clientes')


@login_required
def atualizar_plano_empresa(request, empresa_id):
    if not request.user.is_superuser:
        return redirect('home')
    if request.method == 'POST':
        empresa = get_object_or_404(Empresa, id=empresa_id)
        plano = request.POST.get('plano', empresa.plano)
        if plano in dict(Empresa.PLANO_CHOICES):
            empresa.plano = plano
            empresa.save(update_fields=['plano'])
    return redirect('saas_dashboard')


# ==============================================================================
# RELATÓRIO DE FALTAS (staff)
# ==============================================================================

@login_required
@staff_member_required
def relatorio_faltas(request):
    hoje = timezone.localdate()
    mes  = int(request.GET.get('mes',  hoje.month))
    ano  = int(request.GET.get('ano',  hoje.year))
    cliente_id = request.GET.get('cliente', '')

    qs = por_empresa(
        ItemPedidoIgnorado.objects.select_related('cliente', 'pedido'),
        request, campo='cliente__empresa'
    ).filter(
        data_tentativa__year=ano,
        data_tentativa__month=mes,
    )

    if cliente_id:
        qs = qs.filter(cliente__client_id=cliente_id)

    qs = qs.order_by('-data_tentativa')

    # Resumo agregado
    from django.db.models import Count, Sum
    resumo_produtos = (
        qs.values('codigo_produto', 'descricao_produto')
          .annotate(total_qtd=Sum('quantidade_tentada'), ocorrencias=Count('id'))
          .order_by('-ocorrencias')[:10]
    )
    resumo_clientes = (
        qs.values('cliente__client_code', 'cliente__client_name')
          .annotate(total_faltas=Count('id'))
          .order_by('-total_faltas')[:10]
    )
    total_registros   = qs.count()
    total_unidades    = qs.aggregate(t=Sum('quantidade_tentada'))['t'] or 0
    produtos_distintos = qs.values('codigo_produto').distinct().count()

    # Lista de clientes para o filtro
    clientes = por_empresa(
        WfClient.objects.filter(client_is_active=True),
        request
    ).order_by('client_name')

    # Anos disponíveis para o filtro
    anos = list(range(hoje.year, hoje.year - 4, -1))

    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'analise/relatorio_faltas.html', {
        'page_obj': page_obj,
        'mes': mes,
        'ano': ano,
        'cliente_id': cliente_id,
        'clientes': clientes,
        'anos': anos,
        'total_registros': total_registros,
        'total_unidades': total_unidades,
        'produtos_distintos': produtos_distintos,
        'resumo_produtos': resumo_produtos,
        'resumo_clientes': resumo_clientes,
        'meses': [
            (1,'Janeiro'),(2,'Fevereiro'),(3,'Março'),(4,'Abril'),
            (5,'Maio'),(6,'Junho'),(7,'Julho'),(8,'Agosto'),
            (9,'Setembro'),(10,'Outubro'),(11,'Novembro'),(12,'Dezembro'),
        ],
    })


@login_required
@staff_member_required
def exportar_faltas_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    hoje = timezone.localdate()
    mes  = int(request.GET.get('mes',  hoje.month))
    ano  = int(request.GET.get('ano',  hoje.year))
    cliente_id = request.GET.get('cliente', '')

    qs = por_empresa(
        ItemPedidoIgnorado.objects.select_related('cliente', 'pedido'),
        request, campo='cliente__empresa'
    ).filter(data_tentativa__year=ano, data_tentativa__month=mes)

    if cliente_id:
        qs = qs.filter(cliente__client_id=cliente_id)

    qs = qs.order_by('-data_tentativa')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Faltas'

    header_fill = PatternFill('solid', fgColor='6366F1')
    header_font = Font(bold=True, color='FFFFFF')
    headers = ['Data', 'Cód. Cliente', 'Cliente', 'Pedido', 'Cód. Produto', 'Descrição', 'Qtd', 'Motivo']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_num, item in enumerate(qs, 2):
        ws.cell(row=row_num, column=1, value=item.data_tentativa.strftime('%d/%m/%Y %H:%M') if item.data_tentativa else '')
        ws.cell(row=row_num, column=2, value=item.cliente.client_code if item.cliente else '')
        ws.cell(row=row_num, column=3, value=item.cliente.client_name if item.cliente else '')
        ws.cell(row=row_num, column=4, value=str(item.pedido_id) if item.pedido_id else '')
        ws.cell(row=row_num, column=5, value=item.codigo_produto)
        ws.cell(row=row_num, column=6, value=item.descricao_produto or '')
        ws.cell(row=row_num, column=7, value=item.quantidade_tentada or 0)
        ws.cell(row=row_num, column=8, value=item.motivo_erro)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = f'faltas_{ano}_{mes:02d}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    wb.save(response)
    return response


# ==============================================================================
# DASHBOARD SAAS (superuser)
# ==============================================================================

def saas_dashboard(request):
    if not request.user.is_superuser:
        return redirect('home')

    from datetime import date, timedelta
    hoje = date.today()

    empresas = Empresa.objects.prefetch_related('membros').order_by('nome')

    total_empresas = empresas.count()
    total_ativas = empresas.filter(ativo=True).count()
    total_inativas = total_empresas - total_ativas
    total_clientes = WfClient.objects.count()
    total_usuarios = User.objects.count()
    total_pedidos = Pedido.objects.count()

    # Alertas
    alertas_expirando = empresas.filter(
        ativo=True,
        acesso_permanente=False,
        expira_em__isnull=False,
        expira_em__lte=hoje + timedelta(days=7),
        expira_em__gte=hoje,
    )
    alertas_expirados = empresas.filter(
        ativo=True,
        acesso_permanente=False,
        expira_em__isnull=False,
        expira_em__lt=hoje,
    )
    tabela = []
    for emp in empresas:
        ultimo_pedido = Pedido.objects.filter(empresa=emp).order_by('-data_criacao').first()
        ultimo_login = (
            emp.membros.select_related('user')
            .filter(user__last_login__isnull=False)
            .order_by('-user__last_login')
            .values_list('user__last_login', flat=True)
            .first()
        )
        expirado = emp.expira_em and emp.expira_em < hoje
        expirando = emp.expira_em and hoje <= emp.expira_em <= hoje + timedelta(days=7)
        tabela.append({
            'empresa': emp,
            'clientes': WfClient.objects.filter(empresa=emp).count(),
            'clientes_ativos': WfClient.objects.filter(empresa=emp, client_is_active=True).count(),
            'produtos': Product.objects.filter(empresa=emp).count(),
            'pedidos': Pedido.objects.filter(empresa=emp).count(),
            'membros': emp.membros.count(),
            'ultimo_pedido': ultimo_pedido.data_criacao if ultimo_pedido else None,
            'ultimo_login': ultimo_login,
            'expirado': expirado,
            'expirando': expirando,
            'plano_choices': Empresa.PLANO_CHOICES,
        })

    return render(request, 'saas/dashboard_saas.html', {
        'total_empresas': total_empresas,
        'total_ativas': total_ativas,
        'total_inativas': total_inativas,
        'total_clientes': total_clientes,
        'total_usuarios': total_usuarios,
        'total_pedidos': total_pedidos,
        'tabela': tabela,
        'alertas_expirando': alertas_expirando,
        'alertas_expirados': alertas_expirados,
        'plano_choices': Empresa.PLANO_CHOICES,
    })


# ==============================================================================
# STRIPE — CHECKOUT E WEBHOOK
# ==============================================================================

from django.conf import settings as django_settings
from django.views.decorators.csrf import csrf_exempt

def _get_stripe():
    stripe.api_key = django_settings.STRIPE_SECRET_KEY
    return stripe


@login_required
def criar_checkout_stripe(request, empresa_id):
    empresa = get_object_or_404(Empresa, id=empresa_id)
    # Superuser sempre pode; membro ADMIN da própria empresa também pode
    if not request.user.is_superuser:
        try:
            perfil = request.user.perfil
            if perfil.empresa_id != empresa.id or perfil.papel != 'ADMIN':
                return redirect('home')
        except Exception:
            return redirect('home')

    st = _get_stripe()

    # Cria ou recupera o customer no Stripe
    if not empresa.stripe_customer_id:
        customer = st.Customer.create(
            name=empresa.nome,
            email=empresa.email_contato or '',
            metadata={'empresa_id': empresa.id, 'slug': empresa.slug},
        )
        empresa.stripe_customer_id = customer.id
        empresa.save(update_fields=['stripe_customer_id'])
    
    base_url = request.build_absolute_uri('/').rstrip('/')

    session = st.checkout.Session.create(
        customer=empresa.stripe_customer_id,
        payment_method_types=['card'],
        line_items=[{'price': django_settings.STRIPE_PRICE_BASICO, 'quantity': 1}],
        mode='subscription',
        success_url=f"{base_url}/saas/stripe/sucesso/?session_id={{CHECKOUT_SESSION_ID}}&empresa_id={empresa.id}",
        cancel_url=f"{base_url}/saas/empresas/{empresa.id}/",
        metadata={'empresa_id': empresa.id},
    )

    return redirect(session.url, permanent=False)


@login_required
def stripe_sucesso(request):
    session_id = request.GET.get('session_id')
    empresa_id = request.GET.get('empresa_id')

    if session_id and empresa_id:
        st = _get_stripe()
        try:
            session = st.checkout.Session.retrieve(session_id, expand=['subscription'])
            empresa = get_object_or_404(Empresa, id=empresa_id)

            # Verifica permissão: superuser ou ADMIN da própria empresa
            if not request.user.is_superuser:
                try:
                    perfil = request.user.perfil
                    if perfil.empresa_id != empresa.id or perfil.papel != 'ADMIN':
                        return redirect('home')
                except Exception:
                    return redirect('home')

            sub = session.subscription
            empresa.stripe_subscription_id = sub.id
            empresa.plano = 'BASICO'
            empresa.ativo = True
            empresa.expira_em = date.fromtimestamp(sub.current_period_end)
            empresa.save(update_fields=['stripe_subscription_id', 'plano', 'ativo', 'expira_em'])
            messages.success(request, f'Assinatura ativada para "{empresa.nome}"!')
        except Exception as e:
            messages.error(request, f'Erro ao confirmar assinatura: {e}')

    if request.user.is_superuser:
        return redirect('saas_dashboard')
    return redirect('perfil_representante')


@csrf_exempt
def stripe_webhook(request):
    payload = request.body
    sig_header = request.META.get('HTTP_STRIPE_SIGNATURE', '')
    webhook_secret = django_settings.STRIPE_WEBHOOK_SECRET

    st = _get_stripe()

    try:
        event = st.Webhook.construct_event(payload, sig_header, webhook_secret)
    except Exception:
        from django.http import HttpResponse
        return HttpResponse(status=400)

    from django.http import HttpResponse
    data = event['data']['object']

    if event['type'] == 'checkout.session.completed':
        sub_id = data.get('subscription')
        empresa_id = data.get('metadata', {}).get('empresa_id')
        if sub_id and empresa_id:
            empresa = Empresa.objects.filter(id=empresa_id).first()
            if empresa:
                empresa.stripe_subscription_id = sub_id
                empresa.ativo = True
                empresa.save(update_fields=['stripe_subscription_id', 'ativo'])

    elif event['type'] == 'invoice.paid':
        sub_id = data.get('subscription')
        periodo_fim = data.get('lines', {}).get('data', [{}])[0].get('period', {}).get('end')
        if sub_id:
            empresa = Empresa.objects.filter(stripe_subscription_id=sub_id).first()
            if empresa and periodo_fim:
                empresa.expira_em = date.fromtimestamp(periodo_fim)
                empresa.ativo = True
                empresa.save(update_fields=['expira_em', 'ativo'])

    elif event['type'] in ('customer.subscription.deleted', 'invoice.payment_failed'):
        sub_id = data.get('id') or data.get('subscription')
        if sub_id:
            empresa = Empresa.objects.filter(stripe_subscription_id=sub_id).first()
            if empresa:
                empresa.ativo = False
                empresa.save(update_fields=['ativo'])

    return HttpResponse(status=200)


def acesso_bloqueado(request):
    empresa = getattr(request, 'empresa', None)
    return render(request, 'saas/acesso_bloqueado.html', {'empresa': empresa})


def toggle_acesso_permanente(request, empresa_id):
    if not request.user.is_superuser:
        return redirect('home')
    empresa = get_object_or_404(Empresa, id=empresa_id)
    empresa.acesso_permanente = not empresa.acesso_permanente
    empresa.save(update_fields=['acesso_permanente'])
    messages.success(request, f"Acesso permanente {'ativado' if empresa.acesso_permanente else 'desativado'} para {empresa.nome}.")
    return redirect('detalhe_empresa', empresa_id=empresa.id)


@login_required
@staff_member_required
def perfil_representante(request):
    try:
        perfil = request.user.perfil
        empresa = perfil.empresa
    except Exception:
        empresa = getattr(request, 'empresa', None)
        perfil = None

    hoje = date.today()
    dias_restantes = None
    status_assinatura = 'sem_plano'

    if empresa:
        if empresa.expira_em:
            dias_restantes = (empresa.expira_em - hoje).days
            if dias_restantes < 0:
                status_assinatura = 'expirado'
            elif dias_restantes <= 7:
                status_assinatura = 'expirando'
            else:
                status_assinatura = 'ativo'
        elif empresa.plano != 'FREE':
            status_assinatura = 'ativo'

    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        email = request.POST.get('email', '').strip()
        if nome:
            request.user.first_name = nome.split()[0]
            request.user.last_name = ' '.join(nome.split()[1:])
        if email:
            request.user.email = email
        request.user.save()
        messages.success(request, 'Perfil atualizado com sucesso.')
        return redirect('perfil_representante')

    membros = empresa.membros.select_related('user').all() if empresa else []

    return render(request, 'saas/perfil_representante.html', {
        'perfil': perfil,
        'empresa': empresa,
        'dias_restantes': dias_restantes,
        'status_assinatura': status_assinatura,
        'hoje': hoje,
        'membros': membros,
    })


# ==========================================


# ==========================================
# TAREFAS E COLABORAÇÃO
# ==========================================

def _log_atividade(tarefa, usuario, descricao):
    AtividadeTarefa.objects.create(tarefa=tarefa, usuario=usuario, descricao=descricao)


def registrar_log(request, acao, descricao, modelo='', objeto_id=None):
    try:
        ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', ''))
        ip = ip.split(',')[0].strip() or None
        LogAuditoria.objects.create(
            empresa=getattr(request, 'empresa', None),
            usuario=request.user if request.user.is_authenticated else None,
            acao=acao,
            modelo=modelo,
            objeto_id=objeto_id,
            descricao=descricao,
            ip=ip,
        )
    except Exception:
        pass


def _criar_notificacao(tarefa, mensagem, excluir_user=None):
    membros = tarefa.empresa.membros.exclude(user=excluir_user).select_related('user')
    notifs = [
        NotificacaoTarefa(usuario=m.user, tarefa=tarefa, mensagem=mensagem)
        for m in membros
    ]
    if notifs:
        NotificacaoTarefa.objects.bulk_create(notifs)


def _parsear_mencoes(texto, tarefa, autor):
    """Detecta @username no texto e cria notificação para cada mencionado."""
    import re
    usernames = re.findall(r'@(\w+)', texto)
    if not usernames:
        return
    membros_map = {
        m.user.username: m.user
        for m in tarefa.empresa.membros.select_related('user')
    }
    for username in set(usernames):
        user = membros_map.get(username)
        if user and user != autor:
            NotificacaoTarefa.objects.create(
                usuario=user,
                tarefa=tarefa,
                mensagem=f'{autor.get_full_name() or autor.username} mencionou você em "{tarefa.titulo}"'
            )


def _dados_filtros(request):
    membros = request.empresa.membros.select_related('user')
    tags = TagTarefa.objects.filter(empresa=request.empresa)
    return membros, tags


@staff_member_required
def dashboard_produtividade(request):
    if not request.empresa:
        return redirect('home')

    hoje = date.today()

    qs_all = Tarefa.objects.filter(empresa=request.empresa)

    # ── KPIs ──────────────────────────────────────────────────
    total = qs_all.count()
    concluidas_total = qs_all.filter(status='CONCLUIDO').count()
    taxa_conclusao = int(concluidas_total / total * 100) if total > 0 else 0

    inicio_semana = hoje - timedelta(days=hoje.weekday())
    concluidas_semana = qs_all.filter(
        status='CONCLUIDO',
        atualizado_em__date__gte=inicio_semana
    ).count()

    # Atrasadas: não concluídas com prazo passado
    atrasadas = qs_all.exclude(status='CONCLUIDO').filter(prazo__lt=hoje).count()

    em_andamento = qs_all.filter(status='EM_ANDAMENTO').count()

    # ── Concluídas por semana (últimas 8 semanas) ──────────────
    semanas_labels = []
    semanas_data = []
    for i in range(7, -1, -1):
        seg = hoje - timedelta(weeks=i, days=hoje.weekday())
        dom = seg + timedelta(days=6)
        count = qs_all.filter(
            status='CONCLUIDO',
            atualizado_em__date__range=[seg, dom]
        ).count()
        semanas_labels.append(seg.strftime('%d/%m'))
        semanas_data.append(count)

    # ── Por membro ────────────────────────────────────────────
    por_membro = (
        qs_all.filter(responsavel__isnull=False)
        .values('responsavel__id', 'responsavel__first_name',
                'responsavel__last_name', 'responsavel__username')
        .annotate(
            total=Count('id'),
            concluidas=Count('id', filter=Q(status='CONCLUIDO')),
            atrasadas_m=Count('id', filter=Q(prazo__lt=hoje) & ~Q(status='CONCLUIDO')),
        )
        .order_by('-total')[:8]
    )

    membros_labels, membros_total, membros_concluidas = [], [], []
    for m in por_membro:
        nome = (f"{m['responsavel__first_name']} {m['responsavel__last_name']}".strip()
                or m['responsavel__username'])
        membros_labels.append(nome)
        membros_total.append(m['total'])
        membros_concluidas.append(m['concluidas'])

    # ── Por status (donut) ────────────────────────────────────
    status_counts = {s: qs_all.filter(status=s).count() for s, _ in Tarefa.STATUS_CHOICES}
    status_labels = ['A Fazer', 'Em Andamento', 'Revisão', 'Concluído']
    status_data_chart = [
        status_counts.get('A_FAZER', 0),
        status_counts.get('EM_ANDAMENTO', 0),
        status_counts.get('REVISAO', 0),
        status_counts.get('CONCLUIDO', 0),
    ]

    # ── Por prioridade ────────────────────────────────────────
    pri_counts = {p: qs_all.filter(prioridade=p).count() for p, _ in Tarefa.PRIORIDADE_CHOICES}

    # ── Tarefas recentes atrasadas ────────────────────────────
    tarefas_atrasadas = (
        qs_all.exclude(status='CONCLUIDO')
        .filter(prazo__lt=hoje)
        .select_related('responsavel')
        .order_by('prazo')[:10]
    )

    return render(request, 'tarefas/dashboard.html', {
        # KPIs
        'total': total,
        'concluidas_total': concluidas_total,
        'taxa_conclusao': taxa_conclusao,
        'concluidas_semana': concluidas_semana,
        'atrasadas': atrasadas,
        'em_andamento': em_andamento,
        # Charts
        'semanas_labels': json.dumps(semanas_labels),
        'semanas_data': json.dumps(semanas_data),
        'membros_labels': json.dumps(membros_labels),
        'membros_total': json.dumps(membros_total),
        'membros_concluidas': json.dumps(membros_concluidas),
        'status_labels': json.dumps(status_labels),
        'status_data': json.dumps(status_data_chart),
        'pri_alta': pri_counts.get('ALTA', 0),
        'pri_media': pri_counts.get('MEDIA', 0),
        'pri_baixa': pri_counts.get('BAIXA', 0),
        # Lista
        'tarefas_atrasadas': tarefas_atrasadas,
        'hoje': hoje,
    })


@staff_member_required
def tarefas_board(request):
    if not request.empresa:
        return redirect('home')

    qs = Tarefa.objects.filter(empresa=request.empresa).select_related(
        'criado_por', 'responsavel'
    ).prefetch_related('tags').annotate(
        checklist_total=Count('checklist', distinct=True),
        checklist_done=Count('checklist', filter=Q(checklist__concluido=True), distinct=True),
        comentarios_total=Count('comentarios', distinct=True),
    )

    filtro_prioridade = request.GET.get('prioridade', '')
    filtro_responsavel = request.GET.get('responsavel', '')
    filtro_tag = request.GET.get('tag', '')
    filtro_busca = request.GET.get('q', '')
    filtro_minhas = request.GET.get('minhas', '')

    if filtro_prioridade:
        qs = qs.filter(prioridade=filtro_prioridade)
    if filtro_responsavel:
        qs = qs.filter(responsavel_id=filtro_responsavel)
    if filtro_tag:
        qs = qs.filter(tags__id=filtro_tag)
    if filtro_busca:
        qs = qs.filter(titulo__icontains=filtro_busca)
    if filtro_minhas:
        qs = qs.filter(responsavel=request.user)

    STATUS_CONFIG = [
        {'status': 'A_FAZER',      'label': 'A Fazer',      'color': '#7c3aed', 'bg': '#ffffff'},
        {'status': 'EM_ANDAMENTO', 'label': 'Em Andamento', 'color': '#059669', 'bg': '#ffffff'},
        {'status': 'REVISAO',      'label': 'Revisão',      'color': '#d97706', 'bg': '#ffffff'},
        {'status': 'CONCLUIDO',    'label': 'Concluído',    'color': '#dc4f2a', 'bg': '#ffffff'},
    ]

    hoje = date.today()
    tarefas_list = list(qs)
    for t in tarefas_list:
        if t.prazo:
            start = t.criado_em.date()
            total = (t.prazo - start).days
            if total <= 0:
                t.timeline_pct = None
                t.timeline_color = '#6366f1'
            else:
                elapsed = (hoje - start).days
                t.timeline_pct = max(0, min(100, int(elapsed / total * 100)))
                if t.status == 'CONCLUIDO':
                    t.timeline_color = '#22c55e'
                elif hoje > t.prazo:
                    t.timeline_color = '#ef4444'
                elif hoje == t.prazo or t.timeline_pct >= 75:
                    t.timeline_color = '#f59e0b'
                else:
                    t.timeline_color = '#6366f1'
        else:
            t.timeline_pct = None
            t.timeline_color = '#6366f1'

    colunas = []
    for cfg in STATUS_CONFIG:
        colunas.append({
            **cfg,
            'tarefas': [t for t in tarefas_list if t.status == cfg['status']],
        })

    membros, tags = _dados_filtros(request)

    return render(request, 'tarefas/board.html', {
        'colunas': colunas,
        'membros': membros,
        'tags': tags,
        'filtro_prioridade': filtro_prioridade,
        'filtro_responsavel': filtro_responsavel,
        'filtro_tag': filtro_tag,
        'filtro_busca': filtro_busca,
        'filtro_minhas': filtro_minhas,
        'hoje': hoje,
    })


@staff_member_required
def tarefas_lista(request):
    if not request.empresa:
        return redirect('home')

    qs = Tarefa.objects.filter(empresa=request.empresa).select_related(
        'criado_por', 'responsavel'
    ).prefetch_related('tags').annotate(
        checklist_total=Count('checklist', distinct=True),
        checklist_done=Count('checklist', filter=Q(checklist__concluido=True), distinct=True),
        comentarios_total=Count('comentarios', distinct=True),
    ).order_by('criado_em')

    filtro_status     = request.GET.get('status', '')
    filtro_prioridade = request.GET.get('prioridade', '')
    filtro_responsavel = request.GET.get('responsavel', '')
    filtro_tag        = request.GET.get('tag', '')
    filtro_busca      = request.GET.get('q', '')
    filtro_minhas     = request.GET.get('minhas', '')

    if filtro_status:
        qs = qs.filter(status=filtro_status)
    if filtro_prioridade:
        qs = qs.filter(prioridade=filtro_prioridade)
    if filtro_responsavel:
        qs = qs.filter(responsavel_id=filtro_responsavel)
    if filtro_tag:
        qs = qs.filter(tags__id=filtro_tag)
    if filtro_busca:
        qs = qs.filter(titulo__icontains=filtro_busca)
    if filtro_minhas:
        qs = qs.filter(responsavel=request.user)

    hoje = date.today()

    # Calcula % do tempo decorrido entre criação e prazo para barra de cronograma
    tarefas_list = list(qs)
    for t in tarefas_list:
        if t.prazo:
            start = t.criado_em.date()
            total = (t.prazo - start).days
            if total <= 0:
                # Prazo igual ou anterior à criação — dado inválido, omite barra
                t.timeline_pct = None
                t.timeline_color = '#6366f1'
            else:
                elapsed = (hoje - start).days
                t.timeline_pct = max(0, min(100, int(elapsed / total * 100)))
                if t.status == 'CONCLUIDO':
                    t.timeline_color = '#22c55e'        # verde — concluída
                elif hoje > t.prazo:
                    t.timeline_color = '#ef4444'        # vermelho — atrasada (prazo passou)
                elif hoje == t.prazo or t.timeline_pct >= 75:
                    t.timeline_color = '#f59e0b'        # laranja — prazo hoje ou chegando
                else:
                    t.timeline_color = '#6366f1'        # azul — ok
        else:
            t.timeline_pct = None
            t.timeline_color = '#6366f1'

    STATUS_CONFIG = [
        {'status': 'A_FAZER',      'label': 'A Fazer',      'color': '#7c3aed', 'bg': '#ffffff'},
        {'status': 'EM_ANDAMENTO', 'label': 'Em Andamento', 'color': '#059669', 'bg': '#ffffff'},
        {'status': 'REVISAO',      'label': 'Revisão',      'color': '#d97706', 'bg': '#ffffff'},
        {'status': 'CONCLUIDO',    'label': 'Concluído',    'color': '#dc4f2a', 'bg': '#ffffff'},
    ]

    grupos = []
    for cfg in STATUS_CONFIG:
        s = cfg['status']
        if filtro_status and filtro_status != s:
            continue
        grupos.append({**cfg, 'tarefas': [t for t in tarefas_list if t.status == s]})

    membros, tags = _dados_filtros(request)

    return render(request, 'tarefas/lista.html', {
        'grupos': grupos,
        'total_tarefas': len(tarefas_list),
        'membros': membros,
        'tags': tags,
        'filtro_status': filtro_status,
        'filtro_prioridade': filtro_prioridade,
        'filtro_responsavel': filtro_responsavel,
        'filtro_tag': filtro_tag,
        'filtro_busca': filtro_busca,
        'filtro_minhas': filtro_minhas,
        'status_choices': Tarefa.STATUS_CHOICES,
        'prioridade_choices': Tarefa.PRIORIDADE_CHOICES,
        'hoje': hoje,
    })


@staff_member_required
def criar_tarefa(request):
    if not request.empresa:
        return redirect('home')

    membros, tags = _dados_filtros(request)

    if request.method == 'POST':
        titulo = request.POST.get('titulo', '').strip()
        if not titulo:
            messages.error(request, 'O título é obrigatório.')
            return render(request, 'tarefas/form.html', {
                'membros': membros, 'tags': tags,
                'prioridade_choices': Tarefa.PRIORIDADE_CHOICES,
                'status_choices': Tarefa.STATUS_CHOICES,
            })

        responsavel_id = request.POST.get('responsavel') or None
        prazo_str = request.POST.get('prazo') or None
        prazo = None
        if prazo_str:
            try:
                prazo = datetime.strptime(prazo_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        tarefa = Tarefa.objects.create(
            empresa=request.empresa,
            titulo=titulo,
            descricao=request.POST.get('descricao', '').strip() or None,
            prioridade=request.POST.get('prioridade', 'MEDIA'),
            status=request.POST.get('status', 'A_FAZER'),
            criado_por=request.user,
            responsavel_id=responsavel_id,
            prazo=prazo,
        )

        tag_ids = request.POST.getlist('tags')
        if tag_ids:
            tarefa.tags.set(TagTarefa.objects.filter(id__in=tag_ids, empresa=request.empresa))

        _log_atividade(tarefa, request.user, 'Tarefa criada')
        _criar_notificacao(tarefa, f'Nova tarefa: "{tarefa.titulo}"', excluir_user=request.user)

        if responsavel_id and int(responsavel_id) != request.user.id:
            try:
                resp_user = User.objects.get(pk=responsavel_id)
                NotificacaoTarefa.objects.get_or_create(
                    usuario=resp_user, tarefa=tarefa,
                    defaults={'mensagem': f'Você foi atribuído à tarefa "{tarefa.titulo}"'}
                )
            except User.DoesNotExist:
                pass

        messages.success(request, 'Tarefa criada!')
        return redirect('detalhe_tarefa', tarefa_id=tarefa.id)

    return render(request, 'tarefas/form.html', {
        'membros': membros,
        'tags': tags,
        'prioridade_choices': Tarefa.PRIORIDADE_CHOICES,
        'status_choices': Tarefa.STATUS_CHOICES,
    })


@staff_member_required
def detalhe_tarefa(request, tarefa_id):
    if not request.empresa:
        return redirect('home')

    tarefa = get_object_or_404(Tarefa, id=tarefa_id, empresa=request.empresa)
    comentarios = tarefa.comentarios.select_related('autor').all()
    atividades = tarefa.atividades.select_related('usuario').all()[:20]
    checklist = tarefa.checklist.all()
    membros, _ = _dados_filtros(request)
    hoje = date.today()

    nome_membros = [m.user.username for m in membros]

    if request.method == 'POST':
        texto = request.POST.get('texto', '').strip()
        if texto:
            ComentarioTarefa.objects.create(tarefa=tarefa, autor=request.user, texto=texto)
            _log_atividade(tarefa, request.user, 'Adicionou um comentário')
            _criar_notificacao(
                tarefa,
                f'{request.user.get_full_name() or request.user.username} comentou em "{tarefa.titulo}"',
                excluir_user=request.user
            )
            _parsear_mencoes(texto, tarefa, request.user)
        return redirect('detalhe_tarefa', tarefa_id=tarefa.id)

    anexos = tarefa.anexos.select_related('enviado_por').all()

    return render(request, 'tarefas/detalhe.html', {
        'tarefa': tarefa,
        'comentarios': comentarios,
        'atividades': atividades,
        'checklist': checklist,
        'anexos': anexos,
        'membros': membros,
        'nome_membros': nome_membros,
        'hoje': hoje,
        'status_choices': Tarefa.STATUS_CHOICES,
    })


@staff_member_required
@require_POST
def upload_anexo_tarefa(request, tarefa_id):
    tarefa = get_object_or_404(Tarefa, id=tarefa_id, empresa=request.empresa)
    arquivo = request.FILES.get('arquivo')
    if not arquivo:
        messages.error(request, 'Nenhum arquivo selecionado.')
        return redirect('detalhe_tarefa', tarefa_id=tarefa_id)

    nome = arquivo.name
    AnexoTarefa.objects.create(tarefa=tarefa, arquivo=arquivo, nome=nome, enviado_por=request.user)
    _log_atividade(tarefa, request.user, f'Anexou o arquivo "{nome}"')
    registrar_log(request, 'ANEXO_UPLOAD', f'Arquivo "{nome}" anexado à tarefa "{tarefa.titulo}"',
                  'Tarefa', tarefa.id)
    messages.success(request, f'Arquivo "{nome}" enviado com sucesso.')
    return redirect('detalhe_tarefa', tarefa_id=tarefa_id)


@staff_member_required
@require_POST
def excluir_anexo_tarefa(request, anexo_id):
    anexo = get_object_or_404(AnexoTarefa, id=anexo_id, tarefa__empresa=request.empresa)
    tarefa_id = anexo.tarefa_id
    nome = anexo.nome
    anexo.arquivo.delete(save=False)
    anexo.delete()
    _log_atividade(anexo.tarefa, request.user, f'Removeu o arquivo "{nome}"')
    registrar_log(request, 'ANEXO_EXCLUIDO', f'Arquivo "{nome}" removido da tarefa',
                  'Tarefa', tarefa_id)
    messages.success(request, f'Arquivo "{nome}" removido.')
    return redirect('detalhe_tarefa', tarefa_id=tarefa_id)


@staff_member_required
def logs_auditoria(request):
    if not request.empresa:
        return redirect('home')

    qs = LogAuditoria.objects.filter(empresa=request.empresa).select_related('usuario')

    filtro_acao    = request.GET.get('acao', '')
    filtro_usuario = request.GET.get('usuario', '')
    filtro_data    = request.GET.get('data', '')

    if filtro_acao:
        qs = qs.filter(acao=filtro_acao)
    if filtro_usuario:
        qs = qs.filter(usuario_id=filtro_usuario)
    if filtro_data:
        try:
            from datetime import datetime as dt
            d = dt.strptime(filtro_data, '%Y-%m-%d').date()
            qs = qs.filter(criado_em__date=d)
        except ValueError:
            pass

    # Paginação simples
    from django.core.paginator import Paginator
    paginator = Paginator(qs, 50)
    page = paginator.get_page(request.GET.get('page', 1))

    membros, _ = _dados_filtros(request)

    return render(request, 'auditoria.html', {
        'page_obj': page,
        'acao_choices': LogAuditoria.ACAO_CHOICES,
        'membros': membros,
        'filtro_acao': filtro_acao,
        'filtro_usuario': filtro_usuario,
        'filtro_data': filtro_data,
    })


@staff_member_required
def editar_tarefa(request, tarefa_id):
    if not request.empresa:
        return redirect('home')

    tarefa = get_object_or_404(Tarefa, id=tarefa_id, empresa=request.empresa)
    membros, tags = _dados_filtros(request)

    if request.method == 'POST':
        titulo = request.POST.get('titulo', '').strip()
        if not titulo:
            messages.error(request, 'O título é obrigatório.')
        else:
            responsavel_id = request.POST.get('responsavel') or None
            prazo_str = request.POST.get('prazo') or None
            prazo = None
            if prazo_str:
                try:
                    prazo = datetime.strptime(prazo_str, '%Y-%m-%d').date()
                except ValueError:
                    pass

            alteracoes = []
            antigo_responsavel = tarefa.responsavel_id
            antigo_status = tarefa.status
            antigo_prazo = tarefa.prazo

            tarefa.titulo = titulo
            tarefa.descricao = request.POST.get('descricao', '').strip() or None
            novo_prioridade = request.POST.get('prioridade', tarefa.prioridade)
            novo_status = request.POST.get('status', tarefa.status)

            if novo_status != antigo_status:
                alteracoes.append(f'Status alterado para {tarefa.get_status_display()}')
            if novo_prioridade != tarefa.prioridade:
                alteracoes.append(f'Prioridade alterada para {novo_prioridade}')
            if prazo != antigo_prazo:
                alteracoes.append(f'Prazo alterado para {prazo}')

            tarefa.prioridade = novo_prioridade
            tarefa.status = novo_status
            tarefa.responsavel_id = responsavel_id
            tarefa.prazo = prazo
            tarefa.save()

            tag_ids = request.POST.getlist('tags')
            tarefa.tags.set(TagTarefa.objects.filter(id__in=tag_ids, empresa=request.empresa))

            for alt in alteracoes:
                _log_atividade(tarefa, request.user, alt)

            if responsavel_id and str(responsavel_id) != str(antigo_responsavel):
                try:
                    resp_user = User.objects.get(pk=responsavel_id)
                    if resp_user != request.user:
                        NotificacaoTarefa.objects.create(
                            usuario=resp_user, tarefa=tarefa,
                            mensagem=f'Você foi atribuído à tarefa "{tarefa.titulo}"'
                        )
                    _log_atividade(tarefa, request.user, f'Responsável alterado para {resp_user.get_full_name() or resp_user.username}')
                except User.DoesNotExist:
                    pass

            messages.success(request, 'Tarefa atualizada.')
            return redirect('detalhe_tarefa', tarefa_id=tarefa.id)

    return render(request, 'tarefas/form.html', {
        'tarefa': tarefa,
        'membros': membros,
        'tags': tags,
        'prioridade_choices': Tarefa.PRIORIDADE_CHOICES,
        'status_choices': Tarefa.STATUS_CHOICES,
    })


@staff_member_required
@require_POST
def atualizar_status_tarefa(request, tarefa_id):
    if not request.empresa:
        return redirect('home')

    tarefa = get_object_or_404(Tarefa, id=tarefa_id, empresa=request.empresa)
    novo_status = request.POST.get('status')
    status_validos = [s[0] for s in Tarefa.STATUS_CHOICES]

    if novo_status in status_validos and novo_status != tarefa.status:
        tarefa.status = novo_status
        tarefa.save(update_fields=['status', 'atualizado_em'])
        _log_atividade(tarefa, request.user, f'Status alterado para {tarefa.get_status_display()}')
        _criar_notificacao(
            tarefa,
            f'"{tarefa.titulo}" movida para {tarefa.get_status_display()}',
            excluir_user=request.user
        )

    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or 'tarefas_board'
    return redirect(next_url)


@staff_member_required
@require_POST
def excluir_tarefa(request, tarefa_id):
    if not request.empresa:
        return redirect('home')

    tarefa = get_object_or_404(Tarefa, id=tarefa_id, empresa=request.empresa)
    tarefa.delete()
    messages.success(request, 'Tarefa excluída.')
    next_url = request.POST.get('next') or request.GET.get('next', '')
    if next_url and next_url.startswith('/'):
        return redirect(next_url)
    return redirect('tarefas_board')


@staff_member_required
@require_POST
def toggle_checklist_item(request, item_id):
    item = get_object_or_404(ChecklistItem, id=item_id, tarefa__empresa=request.empresa)
    item.concluido = not item.concluido
    item.save(update_fields=['concluido'])
    return redirect('detalhe_tarefa', tarefa_id=item.tarefa_id)


@staff_member_required
@require_POST
def adicionar_checklist_item(request, tarefa_id):
    tarefa = get_object_or_404(Tarefa, id=tarefa_id, empresa=request.empresa)
    texto = request.POST.get('texto', '').strip()
    if texto:
        ultimo = tarefa.checklist.order_by('-ordem').first()
        ordem = (ultimo.ordem + 1) if ultimo else 0
        ChecklistItem.objects.create(tarefa=tarefa, texto=texto, ordem=ordem)
    return redirect('detalhe_tarefa', tarefa_id=tarefa.id)


@staff_member_required
@require_POST
def excluir_checklist_item(request, item_id):
    item = get_object_or_404(ChecklistItem, id=item_id, tarefa__empresa=request.empresa)
    tarefa_id = item.tarefa_id
    item.delete()
    return redirect('detalhe_tarefa', tarefa_id=tarefa_id)


@staff_member_required
def notificacoes_tarefas(request):
    if not request.empresa:
        return redirect('home')

    notifs = NotificacaoTarefa.objects.filter(
        usuario=request.user
    ).select_related('tarefa').order_by('-criado_em')

    notifs.filter(lida=False).update(lida=True)

    paginator = Paginator(notifs, 20)
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    return render(request, 'tarefas/notificacoes.html', {'page_obj': page_obj})


@staff_member_required
def gerenciar_tags_tarefa(request):
    if not request.empresa:
        return redirect('home')

    tags = TagTarefa.objects.filter(empresa=request.empresa)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'criar':
            nome = request.POST.get('nome', '').strip()
            cor = request.POST.get('cor', '#6366f1')
            if nome:
                TagTarefa.objects.get_or_create(empresa=request.empresa, nome=nome, defaults={'cor': cor})
        elif action == 'excluir':
            tag_id = request.POST.get('tag_id')
            TagTarefa.objects.filter(id=tag_id, empresa=request.empresa).delete()
        return redirect('gerenciar_tags_tarefa')

    return render(request, 'tarefas/tags.html', {'tags': tags})


@staff_member_required
def criar_tarefa_rapida(request):
    if request.method != 'POST' or not request.empresa:
        from django.http import HttpResponseBadRequest
        return HttpResponseBadRequest()

    titulo = request.POST.get('titulo', '').strip()
    status = request.POST.get('status', 'A_FAZER')
    valid = [s[0] for s in Tarefa.STATUS_CHOICES]
    if status not in valid:
        status = 'A_FAZER'
    if not titulo:
        from django.http import HttpResponseBadRequest
        return HttpResponseBadRequest()

    tarefa = Tarefa.objects.create(
        empresa=request.empresa,
        titulo=titulo,
        status=status,
        prioridade='MEDIA',
        criado_por=request.user,
    )
    _log_atividade(tarefa, request.user, 'Tarefa criada')

    tarefa.checklist_total = 0
    tarefa.checklist_done = 0
    tarefa.comentarios_total = 0
    tarefa.timeline_pct = None
    tarefa.timeline_color = '#6366f1'

    # HTMX requests (Kanban board) get the card fragment; plain POSTs redirect back
    if request.headers.get('HX-Request'):
        return render(request, 'tarefas/_card.html', {'tarefa': tarefa, 'col_status': status})

    next_url = request.POST.get('next', '')
    if next_url and next_url.startswith('/'):
        return redirect(next_url)
    return redirect('tarefas_lista')


# ==========================================
# NOTIFICAÇÕES DE PEDIDOS
# ==========================================

def _notificar_novo_pedido(pedido):
    """Cria notificações para todos os membros staff da empresa quando um pedido é criado."""
    empresa = pedido.empresa or getattr(pedido.cliente, 'empresa', None)
    if not empresa:
        return
    membros = empresa.membros.select_related('user')
    nome_cliente = pedido.cliente.client_name or str(pedido.cliente)
    mensagem = f'Novo pedido de {nome_cliente} — #{pedido.id}'
    notifs = [
        NotificacaoPedido(empresa=empresa, usuario=m.user, pedido=pedido, mensagem=mensagem)
        for m in membros
    ]
    if notifs:
        try:
            NotificacaoPedido.objects.bulk_create(notifs, ignore_conflicts=True)
        except Exception:
            pass


@staff_member_required
def notificacoes_pedidos(request):
    if not request.empresa:
        return redirect('home')

    notifs = NotificacaoPedido.objects.filter(
        usuario=request.user, empresa=request.empresa
    ).select_related('pedido', 'pedido__cliente').order_by('-criado_em')

    if request.method == 'POST':
        notif_id = request.POST.get('marcar_lida')
        marcar_todas = request.POST.get('marcar_todas')
        if marcar_todas:
            notifs.filter(lida=False).update(lida=True)
        elif notif_id:
            NotificacaoPedido.objects.filter(id=notif_id, usuario=request.user).update(lida=True)
        return redirect('notificacoes_pedidos')

    paginator = Paginator(notifs, 20)
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    return render(request, 'notificacoes_pedidos.html', {'page_obj': page_obj})
